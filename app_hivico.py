"""
app_hivico.py · HIVICO / HIDALGO VIGUERAS CONSULTORES S.A. DE C.V.
Ejecutar: streamlit run app_hivico.py
"""
import streamlit as st
import pandas as pd
import re, os, io
from datetime import datetime
from collections import defaultdict

# Imports pesados — se cargan una sola vez con cache
@st.cache_resource
def _cargar_openpyxl():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    return Workbook, Font, PatternFill, Alignment, Border, Side, get_column_letter

@st.cache_resource
def _cargar_pdfplumber():
    import pdfplumber
    return pdfplumber

# Cargar al inicio pero solo una vez
pdfplumber = _cargar_pdfplumber()
Workbook, Font, PatternFill, Alignment, Border, Side, get_column_letter = _cargar_openpyxl()

try:
    import psycopg2
    PSYCOPG2_OK = True
except ImportError:
    PSYCOPG2_OK = False

try:
    import reportlab
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

LOGO_B64  = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAMCAgICAgMCAgIDAwMDBAYEBAQEBAgGBgUGCQgKCgkICQkKDA8MCgsOCwkJDRENDg8QEBEQCgwSExIQEw8QEBD/2wBDAQMDAwQDBAgEBAgQCwkLEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD/wAARCACKAYkDASIAAhEBAxEB/8QAHgABAAICAwEBAQAAAAAAAAAAAAcIBQYDBAkCAQr/xABXEAABAwQBAgMDBggHCwgLAAABAgMEAAUGEQcSIQgTMUFRYQkUIjJxgRUWIzdikaGzNkJSc3WCsRckMzVydIOSssLSQ1N2lKK0wdMYJUVUVVeElqPR4f/EABkBAQEBAQEBAAAAAAAAAAAAAAABAgUEA//EACQRAAICAgICAwADAQAAAAAAAAABAhEDEgQxIUETIlEUI2Hw/9oADAMBAAIRAxEAPwD09pSlAKUrBZhmtgwi2Kud9lhAOw0yju48r3JH/iew9poDOkgDZPatBy7m3BsUK2DP/CMtPbyYZCwD+kv6o+Otke6oG5B5syfNHHYUV5VutZJAjsq0Vp/TV6q+z0+HtqOlKUolSiST6k1aJZMOQeJfLpylIsUKJbGj9VXT5rg+9X0T/q1oly5Nz27KJmZXciD6pbkKQk/1UkD9laxSqQ7My5T7ioLnTHX1D0LiidV+w7ncbcSqDNeYKvUtrI3XVpVIbRbeTs+tR3Dyu5dvRLkhTiR219VRI/ZW9WDxMZbBUlF9gxLk0PrK6fKcP3p+iP8AVqHaVClvMT5ywTKClhc42yUo68qZpKSfgv6v69H4VIKVJUkKSQQRsEehFUDBKSCkkEehFbzgnMOWYOtEdiSZtuT2MN8koA/RPqj7u3fuDUotlw6VqmCclY1n8TzLXI8qWhO3YbpAcR7yP5SfiPhvXpW11CilKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUrF5PkduxOxyr7dHOlmMjYSD9JxX8VI+JP/AO/QUBheR+RrTx5ZzLllL018ERIoPdxQ9p9yR22fuqo+T5Ve8vurt3vk1b7zh7AnSUJ9iUj0AG/T7faTXNmeX3TNr9Ivl0dJU4dNtg/RabBPShPwG/7T6kmsFWkjLYpSlUgpSlAKUpQClKUApSlAdq2XSfZ5rNxtkpyPIYWFtuNqIII+yrRcPcwt50j8CXlsNXhhHV1oT9CQkDurt2Sr3j0PqPaBVVttbziWm0lSlHQA9pq0HGuPWfiHHWbhkjMxN1uyCt4swnpHkITo+WfLSrp9QTvWz279O6jKiW6VoKuduL0EpXkTiSPUKgyAR/2K7ELmbj65u+RbbpLluevSxbZLh/Y2ayaN2pXwy6h9lt9vq6XEhaepJSdEb7g9wfga1W78p4ZYX3I93lzopbcU0VLtknoKknR6V+X0qHxBIoDbaVoH93fi09hkiv8AqT//AAVnsez7GspeSzZHpj/WkqS4qA+22QP01oCf20BsNKwuQZjZMYUlN3VMQFI8zrZgvvICdkd1NoIHp7TWtHnfi0djkah/9E//AMFAb/StMgcv4JdVFNsnzpZHr5FrlLA+0hvtWwZBklqxeEm4Xhx9thS+jqajOPaOie4QkkDt6nt+ugMnStAPO/Fw7HI1j7YT/wDwVyxubeN5rqWId7kPuq7JQ1b5ClH7gj4igN6pX42tLraXE76VgKGwQdH4H0r9oBStKuXMnHtomOwLneH4z7KilaHIEgaIOv5Hf09a/LfzNx5dpbUC23l+Q+8oJQhuDIJJ/wBTt60Bu1Kxt8yC3Y7HRJuQleWtXSPm8R18g/ENpUQPia1NXOvF6FFK8icSQSCDBkAgj+pQG/UrSIfNHHlxd8m33WXKc1voYtslw6+xLZrcYUtmfEZmxw4Gn0BxHmNqbVojY2lQBSfgQDQHNSldK73q02GGqfebixDjp7dbywkE+4e8/Ad6A7tKjKX4gMND/wA2s0G63Zz2GPHASe+u3UQr1/R9tc8XnLGUupYv1ovNlUpXSFS4v0Tv0+qSr3/xfYatEskalYWTmOPM45JyiPcmJcGM0pwqYcSrqIHZA7/WJ0AD7TXes1ybvVog3hlpbbc6M1JSheupIWkKAOu2xuoU7lKV8uutMNKefdQ22gdSlrUAlI95J9KA+qVoF45wwS2SFQ4kmTdX0kjpgtBYJ0T2UopSfT2E1j1c9WaG4kXnFMgt7S9acdjpA7nXoVA6+zdWhZJ9KwWMZxi2YNFzH7uzIWkbWydodQPig6Ovj6fGs7UApSlAKq/4hc+VkGQ/ixb39wLSopX0ns5I9FH+r9UfYr2Gp95EyhOHYfcb4FAPNt+XHGt7dV2Sde3RPUfgDVKH3nJDy33VKUtaiolR2T99VEZ8UpXatlsn3mezbLZFckSZCwhttA2VE1oydZKSo6SCT8KkXD+Cs2yoIlPRBbIS9Hzpe0lQ/RTrqP6gD76mfjDhCzYeyzdb601PvBSCeodTUc+ukj0Kv0vh2952LkflvjfiO0fhvkXL7fZY6gfKS+5t58j1DTSdrcPwSDUVt0i+ErZpll8M+IwkpXeLpOnuAaUEdLTZP2fSV/2q2iNwjxlFb8pONBY3va5LxP8At1U3kP5U3Ebc87D4x48nXjpV0pm3WQIrR/SS0gLUofaUH4VCN8+Uy8RVzcdNrj4vaELT0oEa3KcKP0gXVr2ftGvhXojxcsv8Pi+Rjiejszg3jGYkBWOlop3otynR/aoitUvXhjxmUFrsl6mQ3COyXkpdQD8NdJH7a884Pyi3iiiSA9Iyu1zUAEFl+zxgg/e2lKv21IOIfKm8oW9xtvNsAx68sAgLXCW7CeI77OyXE79P4o9Pj2r4eVdEXJxsnrLOCs6xhK5LcEXKKnZLsMlZA+KdBQ+3Wh76jxba21FDiClQ9hFTNxJ8oBwHyc8xa7ndX8QuzxCRHvIShhaydaTISS37e3X0E+6pZzjiPEM/jmchtuJOdR1tTowBC9jsVAdlg9u/rr218JRlB1JUfaLjNXFlPqVs+c8eZFgVwMS7xiWFk+RJR3bdHvB/tB0R7vSuxxjgUzP8katrfU3DZ07LeA+o2D317Nn0HxPuBrIJB8PfGX4SlDN73GBixV6gtrH+EdB+v9iT6fpf5JBsbXXt9vh2qCxbbewliNGbS002n0SkDQFdiss2vBVzxKWyPBztmVHabbM2C2850p0VL6lpJPvOkCpj4JtsOBxrbH4zCUOSy668sAbWoOqSNn26SkCoo8UH8MLb/Rrf712pg4V/NjZP8h79+5VfRF2bvXVu1ui3a2SrZNaQ6xJaU0tK07BBGvSu1X4v6ivsNQpSjA7bFuWeWe3zWkuMPT2UuI0OlSS4naSD7CDqrrpSlCQhCQlKRoADQA91Uy41/OTY/wCkWP3iaudVZEKqJztao9r5JuKIrLbTUjy3kobToAqbSVdvioqP31buqp+In84738yz/sJouwywXFcCLb+PLC3FYQ0HoTUhwJGupxaQpRP2k1tVa7xz/AHHf6Ljfu01sVQqK4eJDBI1qnRcttURDLE4lmUltOkh4dwrQ9qhv70E+prP+FyRb1WW8REob+fIkocWvQ6i0pOkp37QFJUf61SfyBjCMvxG42LpBedaK45PbTye6O/sBI0fgTVY+HcncwvkCKmWtTcaUswpIVsaSogAn3aUEqPwSavaJ0y3tKUqFIr8Rirb+IzbMmO25LfloRFUU7Wg6JUR7daGj/lD4VsfFmCRMGxeNFMZKbjIbS7NcIHUVnv0b9yd6+3Z9tadfR/dA5tgWMflLbi7fzh/sekughR7+h+kW0kH+SqpiqkFVn8TduhxcrgzY8ZDbkqGkuqSNdagtY2fjoAb+AqzFVx8Uf8Aj+0f5n/vqogyQfDxb40XjmPMaYQl2ZIeW4sD6SulXQNn+r6fGpNqPOAvzXWv+dk/vl1IdRlRrmfZvbcCx929Tx5jhPlx2ArRdc9g+AHqT7vedAwPhliyPnbKHr1lU94WqEfyiUdkgE7DTY9B6d/XQHfZIJxnP2WPZHnD1qZc3EtBMVpIPYub/KH7eoEfYlNT3w9j7WPcfWlhLXQ7LaEx4lOlFTg2N/EJ6U/dV6J2bHZcfsuOQ0wLJbWIbCBrpbToq+Kj6qPxJJruSYsaYyqNLjtPsrGlNuICkqHxB7GuSlQpXjnDiCNY4astxRlTURCgJcROyGtnQWj3J2dEezY129Ji40cU7gFgUtfWRAaRv4JTof2VnbjAi3WBJtk5oOR5bS2XUn2pUNEfqNYbj6x3DG8QgWS6eWZMXzUqLZ2kgurKSP6pHb2VbIZqdOiW2G/cJz6WY8ZtTrrij2SkDZNVuyLMsi5tzBjEcfdciWgu9k7I2hPcuua9ewJA9B2A79ztviYyt632aDi0RwpNwUX5Gv8Am0n6CT8CrZ+1ArpeF2wtpjXfJHEArUpMNpWu6R9ZY+/8n+qi8eR/hLWJYTj2FW5ECyQUIUEgOSFAF10+0qV/4eg9grOOtNPtqZfbQ42sdKkLAIUPcQfWvqlQpAvMXGhxLo5CwPrt6oriTJYjkpS2SdJcQB6DfYge8a0N1unD/KrHIFvVAuBS1eYaAXUjQD6PTzEj7dbHp3BHrob3drbHvNrmWmWnqZmMLYcH6KkkH+2qZ43eZ+BZuxPRtLtvldDzaf4yd6Wj7wSPvq9k6Lr0r4YeakstyGFhbbqQtCh6FJGwa+6hSA/FDkRAtOLNOD0VNeTvv32hs/sc/XVfqkHna7G68k3MBQLcQpjI+HQkBQ/1gr9dR9WkZZ+gFRCQNk9hVpeCONGsXsreSXSODdbi2Fthae8dkjsBv0UoHZ9w0O3fcJcN4ijMM4hw5LfXEi7lSQR2LaNdj8CopSfgqt+8aviO/uAcZeVYJCRluSdcS0DsTGSAPNkkH+QFAJ/TWnsQDWoxc5KMSOSgnJmmeL7xzWzhd2Rx5xqIt1zPo6ZUhenI1p36BQ/5R7Xfo9E9irf1T5i5hmuWZ/fpGT5pf5t4uko7dkynStR9yR7EpG+yRoD0AFYqVKkzpLs2bIcfkPrU6666sqW4tR2VKJ7kkkkk1xV18WGOFUuzmZcssr89ClKV9T5ClKUA9KsN4a/GfyPwFMj2aXIeyHDisB60SXSVR0+1UZZ35Z9vT9Q+0AnqFeaVmUVNVI1GTg7ie7GL5Rx1z5x7Hv8AYJbN5sN2b7fxXGXB6oUPVt1BPceo+IIJyOCYHZ8AtK7XaetwuuFx19wDrc7npB17AO36z7TXkv4N/ErcOAORmGbrMdXh1+dRGvEYqJSzs6TKQn2LR7dfWR1D16SPYll5mSy3IjuodadSFoWhQUlSSNggjsQR7a5GfC8Mq9HTw5Vljfs+6UpXwPsVo8UH8MLb/Rrf712tm405hx3HcHtdmmWq8OvRkuBS2WGyg7dWoaJWD6H3VrPig/hhbf6Nb/eu1MHCv5sbJ/kPfv3K16IuzGf3e8T/APgt+/6u1/5lfiue8UKSPwLfu4/93a/8ypLr8V2SSPdUKUx4yUF8j2JQB0bgx6/ziaufVMeNfzk2P+kWP3iaudRkQqqfiJ/OO9/Ms/7CatZVU/ET+cd7+ZZ/2E0j2GWK45/gDjv9Fxv3aa2Ktd45/gDjv9Fxv3aa2KoyoVU7nrFDjGdOz4zQTFuv99t6HYLJ/KJ+P0tn7FCrY1GvPmJjI8GenstgyrOTJQfb5Xo4N+7WlH/IqojM3xTlYzDCIFycd65TKfmsrZ2fMQB3PxUnpV/WrPZHeo+OWGffJXduEwp3p3rqIH0U/aToffVefDZlptmRyMWlO6ZuiNtA+x5AJH2bT1b95CRUgc5z5V1Nk46tatyb3KQt7Q30tBWk9Q/k9W1f6M0ryL8HY4GskhqwzsvuQ6p2QSlvKWU6JbSo9/vWVn4jVSfXWtlvjWm3RbXCR0MRGUMNJ9yUgAf2V2ahRVcfFH/j+0f5n/vqqx1Vx8Uf+P7R/mf++qqiMkvgL811r/nZP75dSGSACT6Co84C/Nda/wCdk/vl1IfrUZUUSu8tcu+Spziwpbz6nVK1rZJ2T+uryWxpuPbYjDSAhDbDaEpHoAEgAVR6/QXLfkU23PJCVx5K2FjWgClXSf2irvWSW1Ps0CawdtyIzTqD8FJBH9tVmUdylKVDQpSlAVZ8Sby3eQEpWezUJptI+HdX9qjUq+HKMhjjvrSSS/NccVv2HoQn/dFRp4mra7HzSJcOk+VLhI0SfVaVKCh+oI/XUheGmb84wSTFUslUeesAEeiS2jX7QqtPonslqlKVkoqk/IiEsZ/fktJCAm6SdBI1r8qqrsVSXPnW5ee3x1snocucggkewuqO/wBVVEZbjjqQuXgWPSHFFS1WyP1E+pIbA3+ytirEYdbnLRidmtboIciQGGVgjX0ktgH9tZeoVFHc2lLm5feZayol6c+4Oo7IClkgftrCVySX3JT7kl07W4oqUfeTX0xDlSXEtR47ji19kpSnZV9lbMFhfC7ZktWq8X5ST1PPIioJHoEjqVr7epH6q84/HLyg/wAm+InIi3J8y2425+AYKQraUpYJDpHs+k8XTv3FPur1D4WgTMa4wU5Nt70WQFPylNutqSpWkDR6T39Ej09dV4j3W4ybvc5d2mq6pE19ch1WydrWoqUe/f1Jr2cKNycjzcuVRUTq0pSuieAUpSgFKUoBSlKAV68fJ9cmzuRvDxb4V1UtyZicpywqdV38xltCFsn+q24lH+j+NeQ9en3yWX5lcn/6Tuf91j15eYk8dnp4rrJRc+lKxOQZZj2LIaXfrmiIHgot9SFKKunW9BIJ9o/XXJOkV+8UH8MLb/Rrf712pg4V/NjZP8h79+5UA8z5D+PWXm5WaLJXCYjojtLWgp6wCVFXSe47qPr39PT0qTuHeTsdtGIQsbyN9y3y4inEILjKyhxKllYPUkEA/SI763rtWn0ZXZM9fi/qK+w18sPNSGW5DCwtt1IWhQ9qSNg1pWWcsYlY4k+IxcVSbmyl1pEZlpZV5qQRoq10jR95+zdZNFaeNfzk2P8ApFj94mrnVSXGH7jj+T2+/G3POIhy231I6e6kpWCR9uhqrb45yFieVPJi2e5lclSSosOMrQtOvXexo/aCR7jVZEbHVU/ET+cd7+ZZ/wBhNWRyHNcYxVaW79dUxVrR5iUeWtalJ2R2CQSe4NVT5OuknMs1uF+gwZIjPKSlkODv0oQlIPw30718aIMtLxz/AABx3+i437tNbFUVcZcqYpHxO12O8yl26bBjIjFDzSylYQOkKCgCO412Ojvfs71KtRlQr5eZakMrYfbC23ElC0qGwpJGiDX1SgKYZTbJ3G/IL7EVakOW6Wl6M4dnadhbaj7+xST8dipp4wkq5I5FunJL8ZxuJBaTFgtuaJQop0dEe5PWSPe5XS8TGIGVEgZfFbHWwfmck6/iHakKPuAPWCfimpG4qxYYlg9utq2uiS6j5zJ2NHzFgHR+KR0p/q1b8GV2bbSlKhoVXHxR/wCP7R/mf++up8vuSWPGo6JV8uCIjTiulKlJUdn3dgarLzdk7Od5KxJsMR9yHEjpZS8pspLp6lKJAPcD6Wu4B7VURk08Bfmutf8AOyf3y6kOoM4T5IsGO4sjGslW9bnWHlqacW0pTa0q7nukHpO9+vbuO++1TbCmxbjDZnwng7HkIS40seikkbB70YRWTxE4W9Zcp/GWM0fmV2+mVAdkPj66fv8ArfHqPuqUeAMzYyHDm7I86kTbOPJKPQqZ/iKA9w+r9w94rfclxy1ZXZpFjvDHmR3x6jXUhXsUk+wj/wDh2CRVbL3x3yDxDfxkGNl6TEaUS3KjpKh0e1Lqe+hr132+Jp2Oi01KhfGvEpZpLDbWU2aVDf1pT0dIW0T79Egj7B1elbA/z7goSlNvTcp77h6W2WY2lKPw6iP2bpQskGdOiW2G9cJ76WY8dBcdcV6JSB3NY7FL1NyG1fhiVAERmQ6tURBVta4/8RavcVdzobGiO9ajGtOU8kSmZmY2/wDBGPsqDrVp6iXZSgdpU92BCR/JIHf2dgqpFSlKEhCEhKUjQAGgBUKRrz1hLmV4gZ8FkuTrQVPoSkbUtogeYkfqSr+qQPWo08NWUN2vJJmNTHOhNzbBa6j/AMq3shP3gr+8Ae2rLVBnJXBs5u5/jdx3pqShwPrhoPQUrB31NH7e/T6j2ewCojJzpUSYpzvCbaTaeQoEu03SOkB1wx1FCz7ygDqSfhrX2egz9w5w45gMKebvLstQT1BtiK51Ht71JCR6e0ilCzZMwyKNimNT79KWkCKyothR11uHshP3qIFVe4fwyVnWbNy5SFOQYTglS3FDsoA7CD7yojWvd1H2Vu15Z5A54uTLLFuds2NsL6kLfBCVfp+zzFaJ7DsO42NkmZ8Nw6zYPZW7LZmiEA9brqvrvL1oqV+r09gp0OzOUpSoUjqycB8c2foW7bn7g6gfXlPH4fxUdIPp7QfWt3tdhslkbDNntEOEgDWo7CW/7B8a71KA/FJStJSpIKSNEEdiK8HeVcdOIcm5Zi3lFsWm9TYSUkAEJbeUkfV7egHp2r3jrxT8X8KNA8THIbEVvoQu9OvEb39NwBaz96lE/fXu4L+zR4+YvqmQ9Slckdh6U+3GjtqcddUEIQkbKlE6AA99dI8Bx0r1C4g+TY4jtmFQl8tMXC9ZLLZQ9MDM5cdiItQ2WWw2QVdO9FSidkbAA7VVrxo+EWP4c7jbcjxK5SZ2J3x1bDQlFJfhSQOryVKGgtJTspVoH6KgfQFXwhycc5ao+0sE4R2ZWKlKmrwj8F2HxCcrnBMlu063QG7ZIuDjsJbaXiWyhISnrSoHusb7egNfWUlBOTPnGLk6RCtK9P0fJZ8JpUCvOs2Ukeo86IN//grMR/kyPDq00EPXLMXlD1Wq4sjf3BmvP/MxH3/i5Dyor0++Sy/Mrk//AEnc/wC6x68+uZ8LtWD8xZZgeO+eYFmvcm3RPnDgW4W23ShPUoAAnt66FevPhi4AtXh141aw+FcX58+c6LhdZLhHQuWptCVBtI+qgBAA9SdbJ71nlzXxpfpeNB/Jf4S3SlVy8U3N+Y8bZVhuGWLKrRhFtyRMpyXld1ty5rMZbQBSwlA+iFK33KvYR6aJrmwg5vVHvlJQVssbSoq8O+V8k5Xis5/kO64tfhGmqZteQY7LbdjXWLrs4pCFENOD2p7fWA122du5G5JxDinGHctzW4qiQEOtx2w20p1199Z0hpttAKlrUfQAewk6AJo4NS1CkmtjZ6VpPGHMGHctRbivGTcYs2zPIj3O2XSC5DmwXFp6kJdZcAKeobII2Do6PY1keReRsR4qxSXmmb3QQbZEKUFQQpxbjijpDaEJ2VKUfQD4k6AJqau9a8l2VWbLSoz458RHHXJ+UvYPYfw3DyGJbvwpMtl1tD8J+Kz1pQA4HEgBR60KABO0qBGxW45rmNi4+xS55pk8hxi1WhgyZTjbSnFJbBGyEp7n19lVxknTXkKSatGbpWlNcxYK9kGHYwifINwzu3uXSyo+arAejtsh5SlHWmz0EHStHfb1ra7rc4lltcy8T1qRFgR3JLykpKiG0JKlEAdz2B7VGmuwmmdqlRvxLz/x7za9MGALu8qNCYbfVNkW12PHc6lFPS2tYAUpKkqBA9CD66NfPOOdysCt+KyYuXw8f/C2TwrUtcm1OThLS6F/3ukII8pSukacV2GiPaKuj21fZNlWyJKpSqu+Ibxn41h2GZlH4wmzZWSY7JZgJuKrK8/aUTQ+2HoypAHl+aGvMOiQNjsSe1WEJZHURKagrZZybBhXGMqHcIjMlhZBU08gLSSCCNg9uxAP3VzV17a+5Kt0WS7rrdZQtWhobKQTUUZ34p+KuPMluON3z8YXjZCyL1Pg2WRIg2ougFv5y8hPSjq6hrW+/b12KkYuTpIrko+WS/SviO+zKYblRnUusvIDja0HaVJI2CD7QRUI80eKbBuORlWJ22VcpeU2OyvTHVw7S/MiWx9TKlRvni20lLYUvo9ewBHUUg7pGLm6QlJRVsnGlaPwfl14z3h/D80yBbS7lerPGmylNN9CC4tAKtJ9g2fSt4qNU6YTtWKUql9x5859uOQcuy7Nybx9YrZx1eJsSJAvcLpcmNNeYpCQ4Fg7IR0713JrcMbyXRmc1DsuhStJ4Tz+bylxPi/INytzcCXfLeiS/HbJ6EL2Qro336SQSN77Edz61huSvEbxrxVk7GF5K7eH77LgouMaBbbU/LdfYU4pslPlpIJBQtRBIISnfuBzpLbVLyXZVZvc7FcYuiiu5Y7bJSidlT0RtZJ+0iue32Oy2ntarPCh9tfkI6G+39UCu4lXUkK0Rsb0Roio64LziTn2NXe6Ssuh5CqHfptvEiNbHIKWQ0UjyShwkrUnZ2sdjv4VEm1ZbV0SNStWyLkzE8VzPGsEvkt+PdMtMhFq3HWWXlso61oLgHSlXTrQJG9gDdMJ5LxPkOXkMTFJj0v8WLo5Zrg6WFIaTLb/AMI2hZGl9Oxsj02PfTV1dC1dG00rTOTuWsT4ngQJeSIucqRdpPzK22+1wHJkya/0lXQ202CSdD1Oh3A33FcnF/KmIcvY67kmHSJSmYsx23TI8yKuPIiSm9eYy62sApWnqTsd/Wrq62rwNldezZptsttyQG7jb40pI9EvtJWB+sV0YeIYnbnPNgYzao69g9TUNtJ2PiBWE5S5bwzh+xxr3mMqUPn8pEGBDhRlyZU2SrZS000gbUo6PuHx7jcccDc93PmDmDk+wMLUMbxlu0fgpmRbnIkxhx5lfzlt9DgCwtLrak6IGtdtjvVUJOLl6I5xT19k8+lK1vkLkLFuL8Wk5hmE12Pboy22j5LC3nXHFqCUIQ2gFSlFRA0B8ToAmsLxbzZhnLb14t9gYvNuu2PrZRdLTebc5CmxPNSVNFbax2CglRGj7PsqatravBdldG/UpSslFKUoBXiR4q70xkHiN5DuUboLX4elR0lC+oKDS/K6t/Ho399ewfMvJlo4f4zyDkO8OIDdpiLcYaUrXnyT9FlofFSykfAEn0FeF9zuEu73GVdZ7ynpMx5ch5xXqta1FSlH4kkmuhwYu3I8XMl4UTrVtnEfflXDQR/7ft//AHhFanWZwy+t4vmFjyZ6OqQ3abjGnKaSrpLgadSspB9m+nVe99HiXhnvpVIflVLg03xlhdqIPmSL67ISfZ0tsKSf2upq4GDZzi/JGK2/NMOurVwtVzZDrDzZ7j3oUPVK0nYUk9wQQao58q/c3ExeN7OEDy3F3SSVe3aRGSB/2jXI4y/uSZ1M7/qbPPOt84R5fyDgzki18jY4wzJfgFaHorxIbksLSUuNkjuNg9ld9EA6OtVodK67SkqZy02naPSUfKsYNob4mve/b/6wa/4amjwzeLq0+Ji8Xy2WTBrhZ2bFGafekSZaHQpTiilCAlIB2QlZ3+jXjlXpR8lTZ2WeP84v4SrzZd3jwySka6WmSsaP+mO/urw58GPHjckvJ7MOac5pNlHPETc0XbnzkK5MoKEuZPcikH1GpKx/4V60+Evk7+6zwDimTyJHm3CPFFsuJJ2r5zH/ACalK+K0hLn9cV5A8yPNyeXc3kMrC23cjuS0KHoUmS4QRVxvks+Tvml8yriOfJ03PZTerehR7ec3pt8D4qQWj9jRrfJx7Yk/wxgnrla/T0XqEee7BzTLyG0XbB8Xx/PsRMRyLe8LvBjspkO7JbktvPNqAUAekgnWkj6J6iRLb+T4/FyKJiUi7R27xOjOS40Mq/KOstkBawPcCoA/bWKj8ocfS2re/Gy23uN3W5uWaEpLnZ+cgqC2E+9YKFdvhXNjLV2dCS2VESeFDhnMOMpmdZNk+OWrEY+X3CPJg4pa5fzmNa22kKSSFj6PUsqGwntpA9BpKcr4suHsp5dwmxDC1ly8YtkES/MQ/n6oRmBoKSppMhPdlzS9pc7dJHqPWtwt/PnDF0y1eCQeSrC5f0SnIRgGUEuGQhXSpodWgpYUNdIJNbBc89w2zZVa8IumSQYt+vSFuW+3uO6ekpQFFRSPboJV+o1r5Zb7+zPxrXQibw28W5Nil7y7OsyxGVYblkKosdtE/KX75PdYYQQFSX1qLZIJ0jo9E7B9ATx+NGzx7hxRbbwrI7TZpeN5Lbb1AXdipMKRJZUoJYeWlJ6EqC1fSOk7ABIB2JlhZbjVyyO5YjBvUV+82dtl6fCQvbsdDo22pY9gUO4rTsj5m4FfyF3izKc3xiTc5TyIL1oluIeSp5RHSy4lQKOsq1pCu+/ZuiyPdTY0+uqIJ4GyrJeQPGLk+V5DCx6M83gceM5HsV2Tc2Ym5SFIadkISELdICl/RGgkpHfRNWD51wi68kcPZdg1jWym43m1PRonnL6UF0jaApWjoEgDeu26wTueeGjgK4fiSi8YXhMqR0POwIrTUU/S+qt0NpATvfYr12Purf8AIcxxfE7KnI8jvsSBa1rabTMdc/JFTqglv6Q7fSKgAfTuKs8lyUo+iRhUafsrZx/xvz1cuWOHspzfjy249ZuPMfl2SSpu9tS33XFQ/JDxSgABK1JR0pSVEfSKj6VZPM7bLvOH320W9CVyp1slRmEqUEhTi2lJSCT6dyO9cF+z7DMYmu26/wCRwoMpm2SLy6065pSILH+GkKHsbT7VGsZk/MvFmGWK15LlOd2m3W29toetrzz4/vttSQtKmkj6Sx0qSdgduob9RWZTcmn+Gow1VGJ8OWE37jjhHEsIyeI1FutphKZlNNOJcSlZcWrspPY7Cge3vrA+J3jTLuTbRg0PEILMlyx5ta73NDj6WuiIx5nmKHUfpEdQ+iO5rfbXynx1e4+PSrPmVrmM5Wt5uyrZfCxOW0CXUtkepR0q6h6ggg964rxy7xpYLZebzec0tsSFj01NuujzjukxJKunTS+3ZR6k9vjTd7b+xotdTbqpJmHAfiLY4y5J4DxjCLBdLHlGQPX+FkTl5QwtTa5DT3zf5sRvzQWkjqKgjXV6/R3Zh3xCcKs4qzm73I1oRYpE021qcpxQbXKCOstDtsq6e+tVsFu5IwK8YY7yHa8ttkrGmGHZLtzakJUw222CXCpQ9OnR2D3GqsMjx9EnBT7M1bWXI1uix3gA40whCgDvRCQDVO+euEPEXyZeeQ7TIs0rILbdi0cWk/jeYFtt8dGiW1wEEec+e6QpzaNqJJGhu0r/AChx5FwdPJcnMbW1iy2w6m6qkAR1JKugaV7SVfR1677a3Xew7NcU5AsTOTYZfYt3tb6loRJjq2kqQelSTvuCCNEEbpjyPG7QnBTVM5MQgy7Zidkts9gMyolujMPNhQUEOJaSlSdjsdEEbFVszLifnWwZ1y8zx9iNkyCx8wW9to3GZdhEVZ3RFcYWHGikqeBDiino0N9OyO+p0sfNfE2TZc9geP8AINkuF/jl1K4MeUlbnU3/AIRKddlKT36gCSNHfoa69h574ZyjKfxKx7kqw3C9Fa20RGJSVKdWjfUltX1XCACSEknsakZuLbLKGyo++CsUvWDcOYbh2RsIYulms0aHLbQ4FpS6hACgFJ7Hv7RW9VqaOWeNXEZI43mlrUMPUUX7TwJtxBUD5o9U90LH2pPurq3vmzibHL/a8Wvmf2eHdryhlyDEckDzHUOqCWla/ihZICSrWz6VlvZ2VKlRu1Vf438JWPTuSOUcu5n44st2RfcoeuNhekOJfJiLWtXdKT9He0kpUKsArkDC0RMhnKyOEI+KFxN6c6+0Eob8xfme7SD1H4VwtcmYC/NRbmcrt65TljGSpaS5tRtZOhKA/wCb2QN1qM5QTS9klBSqzYYsSLAiswYMZqPHjtpaZZaQEIbQkaSlKR2AAAAA9KiKdxtlcjxZ27lhEFk43FwdyyKkeenrExUwuBPl/W10H63p31WbxfxG8G5pfY2M4ryZZbldJnX5EVh0la+hClq9R7EpUe/sBrt4zzxw3mWSKxDFuSLFc7wCsJix5QUp0o+t5Z9HNaJPQT2BPoKkZOJXGzfKh7wv8cZZxjheQWbMITMaVcMquV1YS0+l0KjvKSUKJSdAnR7etbPbedeHrxmLnH9t5Fsj2RNSXYarcJAD3ntEhxoA6BWkg7SO/Y1+2/nHiS65ovjq257apOSNyHYira24VOpeaCi4g9tdSehWxv2Gik0nH9Djbs1zxN8YZNyTgEWRx+tlrNMWu0S/Y866sISJTKx1IUo9ulTZWNHsSE7rKeHfjGTxJxJZMSuq0u3pSFz70+F9ZduD6i4+Sr+NpSugK9oQKzOH8v8AGOf3qfjuFZxarzcbWkrlMRHwsoSFdBUCOykhWhtJI2R76/JXMPF8LOmuM5Oc2lGUPqShFr88F7qUnrSkgdkqKe4SSCR3Aq7vXT0TRbbGieJbGOYclZxVvjVNwmWaPPdXklrtd8TZ5s1goAaDcs6KEpV1lQCgVbA+I6nhI4tzXibGcws2aWCNa13bK5d6hIYuhnj5u+20EoLqvyi1JKCCpf0lep9al/LMuxnBbFIyfML3EtNqiFAelyl9DaCtQQkE+8qUAPtrIQZ0O5wY9yt0luTFltIfYebV1IcbUApKkkeoIIIPxq/I9NCaLbYhjxJ8aZzllwwDkHjq2xrxecAvn4TFnkzBFRPZWlIWlLigUpWOhOioaAKvb2ON8P8AgPLFr5j5T5S5NxS22BObotCoUSFcUSw2I7LjakLUACVgFHUrQBUVdOxUmZzzTxVxncI1pz3OrVY5ktrz2WZbvSpbfUU9fp6bBGz7q+8o5j4twyyW3Jcmzu0QrVef8XzDIC2pXbe21I2FDRB2KfI9df8Av0fGttjAeI+x8sZDxo7beHJrjF6VNjqktsTUw5EmECfOZZkKBDLivo6X20Ae/sMeeGPh/P8Aj3lHkHMMoxJdmteVxLZ8xbkZEq8S21sIWlxL76yVrcUV9ZOykfVSSAKn205bjN9xprMrXe4j1jejqlIn+YEs+SnfU4VK0AkaOyda1WDwrmTi3kb8IDBs6tN6NqR5kxMV7qLSO/0yPUp+ifpDY7etFkai4BwTlsblStZTybgC8ZtuZIy23Ksl4faiwJqXdtyXnF9CG0H2qKwU69dg+6tmr5mxUb8oeIvhnh6G7IzjO7bHktpJTbo7ofmuH2BLKNqG9jurSe/citnzzAcW5LxuRieY29cy2ySFKQiQ4wtKh9VSVtqSpJHwNUU5e+S7nh1+68M5siQhRKxa74elY2d6RIQNK+AUlPp3Ua+2KOOT+7o+WSU4r6KyAPFh4uMj8SN3j2+LCcs2I2p1TkC3FYU464djz3yOxX0nQSOyASATsqNfalDOvDFz5xy6tGU8W35plCikyo0Yyo/b2+az1IAPxIqM3Y77Cy08yttY9UqSQR91dfGoqNQ6OZNybuXZx0rtQLXcrrJRCtdvky5DhAQ0w0pxaifYABs1ZjgjwA8w8o3CLcc1tcnDMb6krekXBromPI2CUssK+kFEeilgJHr9LWqs5xgrkxGEpuook35LnCMxmZPkOerutzjYtbmDCTFRIWmNNnudJ2pAPSsttjZ2NguN6rh+VYlqXneCwC9sM2iS8G/5PW8AT9/lj/Vr0B484+xTi3D7bguF2xMG1Wtry2kDupaj3U4tX8ZaiSpR9pNUc+VM4vuT7uK8vwWnHYjDKrFcNbIYPWp1hWvYFFbwJ94SPaK5+PKsnI2PbkxvHg1PPilKV0jnivUH5LaYy7wlkkFKCHI+TOLUfeFxmNf7Jry+q8PyXfKEKxZzkfFtzlJa/GWO3NtwWrQVJj9XW2P0lNrKvsarz8qO2J0ffjSrIrOzcvkveUbzfrhdJXIuLRmpkp19IbTIcUEqWSNgtp76Pvrd+Evk9+SuGuU8d5Igco2N78Dyw4+wmI8C/HUCh5sd9bU2pYBPYEg+yr2UrnvlZWqbPauPjTtIrtzPhmdZb4jcIGE5lcsRcZxi6hy8xbW1MSnbzH5FQeBbHVrfv+j2qMsDxXKoGK8RQrlBukyVA5dub8yS7BU2tbfVL/vhaANISokHf1e/arrUrznosoRJj3LKOO+QODLJhORP5zduTp1ytUk2V9Ee2sm4tufPVS1JDbafLacGwrZ6gNaNZHlu0ct5bnGcc3Y5xybgjBb5bmbDPky1R5jTNoKly0sRvKJfakLfe2QtIISNBWu95qUFlWsXz+14v4rc1vN6tGRMQ86tONs2aQ3ZZLra1lnRC1IQQ30lxIV1a0QretGo9wvIc+40xfDuIolrvELMIeVCPk1texJc+PfGnp/Wq4pnlBR0paOwve+3sCQavPSgsqFJvVv45uHOuF59x9kF3ved3aXLsnzWySJrd6hvxUNxoyX0IUhPlrCgQsgI6jr0qVIfDd0v/hLh8LZY2gXZzEmbc4FL8xLE1DILXf0PlupR6dvodu1TTShCmWENZzydxVy7y/m+MXSDe18fqwy3RJUVaJKzHt63Ji0I1tQdlOnp0O/lgeor6Sm98dX/AA7O8ih3y1wJ3FdosVqv8XH3LsbHcWylbzbkVKVKQpaCBsp76I9h1culC2U5h33OXcS4J5n5Gxm6JTYb/e3r8qBj7jbrDEgSGY8lcRpJWlKx5alEJP199yRvB5nAv984c5Rzq22nILbAyXkm3Xe0yF2hwzPmbTkVJlpiqT1lILalJCkjfT6d6vFSgsqLl72U8iYhxIvF+Rcgvd0icghC8luGJ/NHox+avkLXDU22noSHEJ6iAD79itbjWvNl8KO8NwcLnzcwyvkqczlKHwuJDmpaeEuU+26G+hmO802yhPSkp0tYAVoirvUoQpZZrR+D+NM84b5csGRYjEx7L495sknHIrtzbtDEtZlxi0tDP5Zpl4LQshvt5iQQDoiaOBL5yjnPDF4XmaVouq5Nxg2a7u282566xQnpjz3IxALKlkk60OyQfbU10oCpfDV6hPcFscAW/j7IrLyHAsF1gLS/Y3mWYU8x3kmSZhT5aQ6ojS0rJJcSPQ1grXKjZ5gPC3DuJcd3+35dhd8s0y8/O7LIhNWZMMf34+qQtAQouqCtBKiVlzv30KujSgPPDkvj7PbHauY+VMPxq5Sn7xluQ4re4AjOlUy1SUMqiy2ka+l5Ujf0kg7StXcAE1JXIMq+4TyJar7xUvJ/x7nRLDa7rjEzHXn7XfYiEJIcTK6Ohgsha+pfWOlTahodybi0oWynmdXeZhZ8QXG91xPJZF45GU87jHzC0PymLj86gJjpCXW0lCChwHrCynQ9N13I93Twjz/iE/OrRevmkfiKBjjj9ttUic2JyJaSpvbKFextR39nvFW3pQWVmOH5Peb34pLTYbZIZn32LFjWdwoLKX3VWXy0+W5oD650SD2PrWp264w+SLbwVxzguBX+1X7A73arjfTMssiC1Z40RopltKeWhKFl5egEpUes91d6uLSgsorY8PzCJnUPJsli3OZhMfnO7SVWlm19D8OWp1fzS4+cElxcbrUOtPZHZPc+zZ+CbpOx3mjkbGsgybKYka8ZlkEhrFk4s4uHPQtJKZQnhv6JUEKAT5gSelOvrd7h0oLKteF2bdLHnkjjzBZuQXnjGBZ3ZMVy/Y89b5VimLkAi3h51CC+CCtR7HRSO/bv1LtcrxiviTQ5xI5ksqVlGTR2cwxq6Y86IKYyGwhV1jzSgIQEoQkpAWesq9O3TVsKUBV7mbILzzvM49xbDON5l8tDkydkF5tl9LlqadZhLXFaadcLbgSFPLLgT0kqDaTob2Mz4Zc/vWLYBjPEnIWOX1rI7Tdp+Jh1uE47GCIifMadU9oJ8osqQhC9aUUHQGjqxFKEK28p5RbOPvFTasuyiw3ubZZXHsm0INvsz9wD0pdwbWGOlpCu5QhR79vTfqKh664tmOCcecTXCYjIMAQ1l1+uzKoNlNzlY/CkoeMdpUUIWkEhYSU6+j1k6BB1fOlAQNzFbMm5P8Id2t2GXC6ZFd7jZo5Q+9CVAl3Lynm1P7j9KS2txDbg8sJG+rQGjWRxLnfjOXZk3SwYBlURFsZgWtxH4rvsLjec4G0RUlaU9YbJ2oI2lIG/aNzRSgKU41j+Q4rygviu9Ypd14HxJfrzyDFfYgrcbkxlspdgRY4H+EdbdlSFdI9Sga7g6lz/ANMjj/8A+X/Jv/2nIqeqUApSlAK6Myw2O4Ol+fZoMlxQAK3o6FqI92yK71KA6kK0Wm2qUu3WuJFUsaUWGEoJHuOhXbpSgFYPN8LxzkTE7phOW25E603eOqPJZV66PcKSf4qkkBSVDuFAEelZylE68odnkD4ivBDylwlcJV1slulZRiPUpbNzhslbkdvfZMltOyggeqwOg+8H6IrgQR6iv6Darj4n+L+M3bOi9O8d4wu4uvoDktVojl5YPVva+jqP666WDlSn9ZI8GXjqPmLPICu9Yr5eMZvMLIMfuL8C5W99EiLJYWUuNOJO0qSR6EEVm7tEiN8gPQ24rSWBMKQ0lACAN+nT6ar0U8NHFHFt4kpcu/GuKzlobirSqTZo7pSo+pBUg6Pxr1zlqrPNCOzPvw5fKF2vPo0TF+SsRvTOQoSltU6y252bHlHsOtTLQLjSiT3CUqTs+qR2FxoE1m4w2Z0dDyW30haQ8ytlevihYCkn4EA1w2myWWwREwLFaIVujJACWYkdDLY0NDSUgCu7XGyOLf1VHVxqSX2dilKV8zYpSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQH//Z"
LOGO_MIME = "jpeg"

st.set_page_config(
    page_title="HIVICO · Control de Nómina",
    page_icon="🛡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════
# CSS — CORPORATIVO CLARO CON AZUL MARINO
# ══════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

*, html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}
#MainMenu, footer { visibility: hidden; }

/* ─── Fondo principal: blanco roto cálido ─── */
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background: #F4F6F9;
}
[data-testid="stMain"] > div {
    padding-top: 1.2rem;
}

/* ─── Sidebar ─── */
section[data-testid="stSidebar"] {
    background: #0F2744 !important;
    border-right: none;
    box-shadow: 4px 0 20px rgba(0,0,0,.15);
}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span {
    color: #A8C4E0 !important;
}
section[data-testid="stSidebar"] .stSelectbox > div > div,
section[data-testid="stSidebar"] .stNumberInput > div > div > input {
    background: #1A3A5C !important;
    border: 1px solid #2A4F78 !important;
    color: #E8F2FC !important;
    border-radius: 6px !important;
}

/* ─── Sidebar logo area ─── */
.sb-brand {
    padding: 1.6rem 1.2rem 1.2rem;
    border-bottom: 1px solid rgba(255,255,255,.08);
    margin-bottom: 1.2rem;
}
.sb-logo-wrap {
    background: white;
    border-radius: 8px;
    padding: .6rem 1rem;
    display: flex;
    align-items: center;
    justify-content: center;
    margin-bottom: .7rem;
}
.sb-logo-wrap img {
    height: 36px;
    width: auto;
    display: block;
}
.sb-tagline {
    font-size: .6rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #4A90D9;
    text-align: center;
}

/* ─── Sidebar labels ─── */
.sb-lbl {
    font-size: .58rem;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #3A6A9A;
    margin-bottom: .4rem;
    padding: 0 .1rem;
}

/* ─── Header principal ─── */
.main-hdr {
    background: white;
    border-radius: 12px;
    border: 1px solid #E2E8F0;
    border-left: 5px solid #1A3A6B;
    padding: 1.1rem 1.6rem;
    margin-bottom: 1.4rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 2px 8px rgba(0,0,0,.06);
}
.hdr-left {
    display: flex;
    align-items: center;
    gap: 1.4rem;
}
.hdr-logo-bg {
    background: #F0F5FF;
    border-radius: 8px;
    padding: .5rem .8rem;
    border: 1px solid #D0DCF0;
}
.hdr-logo-bg img {
    height: 38px;
    width: auto;
    display: block;
}
.hdr-sep {
    width: 1px;
    height: 38px;
    background: #E2E8F0;
}
.hdr-title-tag {
    font-size: .58rem;
    font-weight: 700;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    color: #4A90D9;
    margin-bottom: .2rem;
}
.hdr-title-main {
    font-size: .95rem;
    font-weight: 700;
    color: #1A2A3A;
    letter-spacing: -.2px;
}
.hdr-title-sub {
    font-size: .7rem;
    color: #8A9BB0;
    margin-top: .1rem;
}
.hdr-right {
    text-align: right;
}
.hdr-periodo {
    font-family: 'JetBrains Mono', monospace;
    font-size: .95rem;
    font-weight: 500;
    color: #1A3A6B;
    letter-spacing: .5px;
}
.hdr-fecha {
    font-size: .65rem;
    color: #8A9BB0;
    margin-top: .15rem;
}
.hdr-modo {
    display: inline-block;
    margin-top: .4rem;
    padding: 2px 10px;
    background: #EEF4FF;
    border: 1px solid #C0D4F0;
    border-radius: 20px;
    font-size: .6rem;
    font-weight: 700;
    letter-spacing: .6px;
    text-transform: uppercase;
    color: #1A3A6B;
}

/* ─── Tabs ─── */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    background: transparent;
    gap: .3rem;
    border-bottom: 2px solid #E2E8F0;
    padding-bottom: 0;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    background: transparent !important;
    border: none !important;
    color: #8A9BB0 !important;
    font-size: .78rem !important;
    font-weight: 600 !important;
    padding: .55rem 1rem !important;
    border-radius: 0 !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -2px !important;
    letter-spacing: .2px;
}
[data-testid="stTabs"] [aria-selected="true"] {
    color: #1A3A6B !important;
    border-bottom: 2px solid #1A3A6B !important;
    background: transparent !important;
}

/* ─── Sección header ─── */
.sec-hdr {
    display: flex;
    align-items: center;
    gap: .8rem;
    margin: 1.6rem 0 1.1rem;
    padding-bottom: .65rem;
    border-bottom: 2px solid #E8EFF8;
}
.sec-badge {
    background: #1A3A6B;
    color: white;
    font-size: .58rem;
    font-weight: 700;
    letter-spacing: 1.5px;
    padding: 3px 8px;
    border-radius: 4px;
}
.sec-title {
    font-size: .82rem;
    font-weight: 700;
    color: #1A2A3A;
    letter-spacing: .8px;
    text-transform: uppercase;
}
.sec-desc {
    font-size: .7rem;
    color: #8A9BB0;
    margin-left: auto;
}

/* ─── Upload zone ─── */
.up-lbl {
    font-size: .62rem;
    font-weight: 700;
    letter-spacing: 1.8px;
    text-transform: uppercase;
    color: #4A6A8A;
    margin-bottom: .35rem;
}
.up-ok {
    font-size: .72rem;
    color: #1A7A3C;
    font-weight: 600;
    margin-top: .2rem;
}

/* ─── KPI cards ─── */
.kpi-row { display:flex; gap:.85rem; margin:.8rem 0 1.3rem; }
.kpi {
    flex: 1;
    background: white;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 1rem 1.1rem;
    border-top: 3px solid #E2E8F0;
    box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
.kpi.k-blue   { border-top-color: #1A3A6B; }
.kpi.k-green  { border-top-color: #16A34A; }
.kpi.k-orange { border-top-color: #D97706; }
.kpi.k-red    { border-top-color: #DC2626; }
.kpi.k-purple { border-top-color: #7C3AED; }
.kpi-v {
    font-family: 'JetBrains Mono', monospace;
    font-size: 1.75rem;
    font-weight: 500;
    line-height: 1;
    color: #1A2A3A;
}
.kpi.k-green  .kpi-v { color: #16A34A; }
.kpi.k-orange .kpi-v { color: #D97706; }
.kpi.k-red    .kpi-v { color: #DC2626; }
.kpi.k-purple .kpi-v { color: #7C3AED; }
.kpi-l {
    font-size: .6rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1.2px;
    color: #8A9BB0;
    margin-top: .35rem;
}
.kpi-s { font-size: .64rem; color: #B0BCC8; margin-top: .1rem; }

/* ─── Resumen ejecutivo ─── */
.exec {
    background: white;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    margin: .8rem 0;
    box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
.exec-t {
    font-size: .6rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 2px;
    color: #1A3A6B;
    margin-bottom: .85rem;
    padding-bottom: .6rem;
    border-bottom: 2px solid #EEF4FF;
}
.e-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: .48rem 0;
    border-bottom: 1px solid #F4F6F9;
}
.e-row:last-child { border-bottom: none; }
.e-row.hl {
    background: #F0F5FF;
    border-radius: 6px;
    padding: .6rem .75rem;
    margin: .15rem -.5rem;
    border-bottom: none;
}
.e-lbl { font-size: .77rem; color: #5A7A9A; }
.e-lbl b { color: #1A2A3A; font-weight: 600; }
.e-val {
    font-family: 'JetBrains Mono', monospace;
    font-size: .8rem;
    font-weight: 500;
    color: #1A2A3A;
}
.e-val.pos { color: #16A34A; }
.e-val.neg { color: #DC2626; }
.e-val.neu { color: #D97706; }

/* ─── Tabla ─── */
.t-wrap {
    background: white;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    overflow: hidden;
    overflow-x: auto;
    box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
.t {
    width: 100%;
    border-collapse: collapse;
    font-size: .78rem;
}
.t thead tr { background: #1A3A6B; }
.t thead th {
    padding: 10px 13px;
    color: #C8DCF0;
    font-size: .62rem;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
    white-space: nowrap;
    border: none;
}
.t thead th.r { text-align: right; }
.t thead th.c { text-align: center; }
.t tbody tr { border-bottom: 1px solid #F0F4F8; }
.t tbody tr:last-child { border-bottom: none; }
.t tbody tr:hover { background: #F5F8FF !important; }
.t td { padding: 8px 13px; color: #2A3A4A; vertical-align: middle; }
.t td.mono {
    font-family: 'JetBrains Mono', monospace;
    font-size: .73rem;
    color: #4A6A8A;
}
.t td.r {
    text-align: right;
    font-family: 'JetBrains Mono', monospace;
    font-size: .73rem;
}
.t td.c { text-align: center; }
.t tbody tr:nth-child(even) { background: #FAFBFD; }
.t tfoot tr { background: #EEF4FF; border-top: 2px solid #C0D4F0; }
.t tfoot td {
    padding: 9px 13px;
    font-weight: 700;
    font-size: .78rem;
    color: #1A3A6B;
    font-family: 'JetBrains Mono', monospace;
}

/* ─── Log de archivos ─── */
.log-wrap {
    background: white;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,.04);
}

/* ─── Badges ─── */
.badge {
    display: inline-block;
    padding: 3px 9px;
    border-radius: 4px;
    font-size: .62rem;
    font-weight: 700;
    letter-spacing: .4px;
    text-transform: uppercase;
    white-space: nowrap;
}
.b-ok    { background: #DCFCE7; color: #15803D; }
.b-no    { background: #FEE2E2; color: #B91C1C; }
.b-dif   { background: #FEF3C7; color: #92400E; }
.b-doble { background: #EDE9FE; color: #6D28D9; }

/* ─── Modo pill ─── */
.modo {
    padding: .6rem 1rem;
    border-radius: 8px;
    font-size: .75rem;
    font-weight: 600;
    margin-bottom: .8rem;
    display: flex;
    align-items: center;
    gap: .5rem;
}
.modo-s {
    background: #EEF4FF;
    border: 1px solid #C0D4F0;
    color: #1A3A6B;
}
.modo-u {
    background: #F0FDF4;
    border: 1px solid #BBF7D0;
    color: #15803D;
}

/* ─── Separadores de sección ─── */
.mini-sep {
    height: 1px;
    background: linear-gradient(90deg, #1A3A6B 0%, #E2E8F0 60%, transparent 100%);
    margin: 1.6rem 0 1.1rem;
    border: none;
}

/* ─── Botón primario override ─── */
[data-testid="stButton"] > button[kind="primary"] {
    background: #1A3A6B !important;
    border: none !important;
    color: white !important;
    font-weight: 600 !important;
    letter-spacing: .3px !important;
    border-radius: 7px !important;
    padding: .5rem 1.4rem !important;
    transition: background .2s !important;
}
[data-testid="stButton"] > button[kind="primary"]:hover {
    background: #0F2744 !important;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════
MESES = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
         "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]

with st.sidebar:
    st.markdown(f'''
<div class="sb-brand">
  <div class="sb-logo-wrap">
    <img src="data:image/jpeg;base64,{LOGO_B64}" alt="HIVICO">
  </div>
  <div class="sb-tagline">Control de Nómina</div>
</div>
'''.replace("{LOGO_B64}", LOGO_B64), unsafe_allow_html=True)

    st.markdown('<div class="sb-lbl">Periodo</div>', unsafe_allow_html=True)
    mes_sel  = st.selectbox("Mes", MESES, index=0, label_visibility="collapsed")
    anio_sel = st.number_input("Año", 2024, 2030, 2026, label_visibility="collapsed")
    periodo  = f"{mes_sel} {anio_sel}"

    st.markdown('<div style="height:.7rem"></div>', unsafe_allow_html=True)
    st.markdown('<div class="sb-lbl">Tolerancia importe ($)</div>', unsafe_allow_html=True)
    tolerancia = st.number_input("Tolerancia", value=1.00, step=0.50,
                                  min_value=0.0, label_visibility="collapsed")

    st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)

    if PSYCOPG2_OK:
        with st.expander("🗄  PostgreSQL", expanded=False):
            db_host = st.text_input("Host", "localhost")
            db_port = st.text_input("Puerto", "5432")
            db_name = st.text_input("Base de datos", "hivico_nomina")
            db_user = st.text_input("Usuario", "postgres")
            db_pass = st.text_input("Contraseña", "hivico2026", type="password")
            usar_bd = st.checkbox("Activar BD", value=True)
        DB = dict(host=db_host, port=db_port, database=db_name,
                  user=db_user, password=db_pass)
    else:
        usar_bd = False
        DB = {}
        st.markdown('<div style="font-size:.68rem;color:#3A6A9A;margin-top:.5rem">'
                    'psycopg2 no instalado</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# HEADER PRINCIPAL
# ══════════════════════════════════════════════════════════════
modo_lbl = "Servidor · PostgreSQL activo" if (PSYCOPG2_OK and usar_bd) else "Sin base de datos"
st.markdown(f'''
<div class="main-hdr">
  <div class="hdr-left">
    <div class="hdr-logo-bg">
      <img src="data:image/jpeg;base64,{LOGO_B64}" alt="HIVICO">
    </div>
    <div class="hdr-sep"></div>
    <div>
      <div class="hdr-title-tag">Sistema de Control</div>
      <div class="hdr-title-main">Control de Nómina y Dispersión Bancaria</div>
      <div class="hdr-title-sub">HIDALGO VIGUERAS CONSULTORES S.A. DE C.V.</div>
    </div>
  </div>
  <div class="hdr-right">
    <div class="hdr-periodo">{periodo}</div>
    <div class="hdr-fecha">{datetime.now().strftime('%d/%m/%Y  ·  %H:%M')}</div>
    <div class="hdr-modo">{modo_lbl}</div>
  </div>
</div>
'''.replace("{LOGO_B64}", LOGO_B64)
   .replace("{periodo}", periodo)
   .replace("{modo_lbl}", modo_lbl), unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════
TAB1, TAB2, TAB3, TAB4 = st.tabs([
    "  01 · Análisis de Sábanas  ",
    "  02 · Layouts de Dispersión  ",
    "  03 · Cruce Nómina vs Banco  ",
    "  04 · Cruce IMSS  ",
])


# ╔══════════════════════════════════════════════════════════════╗
# ║  LÓGICA COMPARTIDA                                           ║
# ╚══════════════════════════════════════════════════════════════╝

# ── Constantes (script 02) ─────────────────────────────────────
BANCOS_MAP = {"BBVA":"BBVA","BANCOMER":"BBVA","BANORTE":"BANORTE","SANTANDER":"SANTANDER"}
COLS_CLABE  = ["CLABE","CUENTA INTERBANCARIA","CLAVE INTERBANCARIA","NUMERO DE EMPLEADO"]
COLS_CUENTA = ["NUMERO DE CUENTA","CUENTA","BBVA","BANORTE","SANTANDER"]
COLS_TOTAL  = ["TOTAL","TOTAL QUINCENAL","DIF","DIFERENCIAL","IMPORTE DIFERENCIAL"]
NOMBRES_HOJA_SAB = [
    "SABANA PAGO (4)","SABANA DE PAGO. (4)","SABANA DE PAGO (4)",
    "ADM APOYOS Y BAJAS (4)","SABANA(4)","SABANA 4",
    "SABANA PAGO (5)","SABANA DE PAGO (5)","SABANA(5)",
    "SABANA PAGO (3)","SABANA PAGO (2)","SABANA PAGO","SABANA DE PAGO",
    "FW HIDALGO ","FW HIDALGO  (2)","FW HIDALGO  (3)","FW HIDALGO  (4)",
    "turno","ADM APOYOS Y BAJAS",
]
TRABAJADO_MD  = {1, 4, 5, 16}
FALTA_MD      = {2, 15, 52}
MAX_DIAS_QNA  = 16

def encontrar_col(cols, opciones):
    cu = {str(c).upper().strip(): c for c in cols}
    for op in opciones:
        if op.upper() in cu: return cu[op.upper()]
    return None

def limpiar_clabe(valor):
    if valor is None: return None, None
    if isinstance(valor, float):
        if pd.isna(valor): return None, None
        try: s = str(int(valor))
        except: return None, None
    else:
        s = re.sub(r'[\s\-]', '', str(valor).replace("CLABE","").strip())
        if s.endswith(".0"): s = s[:-2]
    if not s or s == "nan": return None, None
    if len(s) == 18: return s, "CLABE"
    if len(s) >= 10: return s, "CUENTA"
    return None, None

def inferir_banco(clabe):
    if not clabe: return None
    return {"012":"BBVA","006":"BANORTE","014":"SANTANDER"}.get(str(clabe)[:3])

def limpiar_fecha(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ["%d/%m/%Y","%d/%m/%y","%Y-%m-%d"]:
        try: return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except: pass
    return None

def limpiar_importe(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return 0.0
    try: return float(str(v).replace(",","").replace("$","").strip())
    except: return 0.0

def nss_norm(v):
    return re.sub(r"[^0-9]", "", str(v)).strip() if v else ""

def xl_fill(c):  return PatternFill("solid", fgColor=c)
def xl_brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def xl_fn(bold=False, color="000000", size=9):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def xl_mfmt(ws, fila, cols):
    for c in cols: ws.cell(fila, c).number_format = '$#,##0.00'
def xl_titulo(ws, texto, fila, ncols, bg="1F4E79"):
    ws.merge_cells(f"A{fila}:{get_column_letter(ncols)}{fila}")
    c = ws.cell(fila, 1, texto)
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.fill = xl_fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 28

# ── Leer sábanas ──────────────────────────────────────────────
def encontrar_hoja_sabana(xl_obj):
    candidatas = []
    for hoja in xl_obj.sheet_names:
        tiene_num = any(p in hoja for p in ["(4)","(5)","(3)","(2)"])
        tiene_kw  = any(p in hoja.upper() for p in ["SABANA","PAGO","APOYOS","HIDALGO","TURNO","FW "])
        if tiene_num or tiene_kw:
            try:
                df_enc = pd.read_excel(xl_obj, sheet_name=hoja, header=None, nrows=15)
                enc = " ".join(str(v).upper() for v in df_enc.values.flatten() if pd.notna(v))
                tiene_banco = any(b in enc for b in ["BBVA","SANTANDER","BANORTE","BANCOMER"])
                tiene_total = "TOTAL" in enc
                candidatas.append((hoja, tiene_banco, tiene_total, tiene_num))
            except:
                candidatas.append((hoja, False, False, tiene_num))
    for h, banco, total, num in candidatas:
        if banco and total and num: return h
    for nombre in NOMBRES_HOJA_SAB:
        if nombre in xl_obj.sheet_names: return nombre
    for h, banco, total, _ in candidatas:
        if banco and total: return h
    for h, _, _, num in candidatas:
        if num: return h
    if candidatas: return candidatas[0][0]
    return xl_obj.sheet_names[0] if xl_obj.sheet_names else None

def leer_sabana_bytes(contenido_bytes, nombre_archivo, qna):
    try:
        xl = pd.ExcelFile(io.BytesIO(contenido_bytes))
    except Exception as e:
        return None, f"No se pudo abrir: {e}"

    hoja = encontrar_hoja_sabana(xl)
    if not hoja:
        return None, f"No se encontró hoja de sábana en {nombre_archivo}"

    df_raw = pd.read_excel(io.BytesIO(contenido_bytes), sheet_name=hoja, header=None)

    meta = {"periodo": "", "sucursal": ""}
    for i in range(8):
        row = [str(v) for v in df_raw.iloc[i].tolist()]
        rs  = " ".join(row).upper()
        if "PERIODO" in rs:
            for j, v in enumerate(row):
                if "PERIODO" in v.upper() and j+1 < len(row) and row[j+1] != "nan":
                    meta["periodo"] = row[j+1].strip()
        if "SUCURSAL" in rs:
            for j, v in enumerate(row):
                if "SUCURSAL" in v.upper() and j+1 < len(row) and row[j+1] != "nan":
                    meta["sucursal"] = row[j+1].strip()

    fila_enc = 8
    for i in range(15):
        row_s = df_raw.iloc[i].astype(str).str.upper()
        if row_s.str.contains("NSS").any():
            fila_enc = i; break

    df = pd.read_excel(io.BytesIO(contenido_bytes), sheet_name=hoja,
                       header=fila_enc, dtype=object)
    cols_l, seen = [], {}
    for col in df.columns:
        c = str(col).strip()
        seen[c] = seen.get(c, 0) + 1
        cols_l.append(c if seen[c] == 1 else f"{c}_{seen[c]}")
    df.columns = cols_l

    col_nss     = encontrar_col(df.columns, ["NSS","N.S.S."])
    col_curp    = encontrar_col(df.columns, ["CURP","curp"])
    col_clabe   = encontrar_col(df.columns, COLS_CLABE)
    col_cuenta  = encontrar_col(df.columns, COLS_CUENTA)
    col_bbva    = encontrar_col(df.columns, ["BBVA","BANCOMER"])
    col_banorte = encontrar_col(df.columns, ["BANORTE"])
    col_san     = encontrar_col(df.columns, ["SANTANDER"])
    col_banco   = encontrar_col(df.columns, ["BANCO"])
    col_total   = encontrar_col(df.columns, COLS_TOTAL)
    col_sueldo  = encontrar_col(df.columns, ["SUELDO MENSUAL"])
    col_fecha   = encontrar_col(df.columns, ["FECHA ALTA"])
    col_turno   = encontrar_col(df.columns, ["rool","ROOL","tno","TNO"])
    col_serv    = encontrar_col(df.columns, ["NOMBRE SERVICIO"])
    col_clave   = encontrar_col(df.columns, ["CLAVE SERVICIO"])
    col_ap      = encontrar_col(df.columns, ["APELLIDO PATERNO","AP. PATERNO"])
    col_am      = encontrar_col(df.columns, ["APELLIDO MATERNO","AP. MATERNO"])
    col_noms    = encontrar_col(df.columns, ["NOMBRES ELEMENTO","NOMBRES","NOMBRE"])
    col_nombcom = encontrar_col(df.columns, ["NOMBRE COMPLETO","TRABAJADOR"])
    col_mediored= encontrar_col(df.columns, ["MEDIO RED"])
    # Columna X (índice 23) — días trabajados directos, sin encabezado
    col_x = df.columns[23] if len(df.columns) > 23 else None

    if not col_nss:
        return None, f"Columna NSS no encontrada en {nombre_archivo}"

    df = df[df[col_nss].notna()].copy()
    df[col_nss] = df[col_nss].astype(str).str.replace('-','').str.replace(' ','').str.strip()
    df = df[~df[col_nss].str.upper().isin(["NSS","N.S.S.","NAN","NONE",""])].copy()
    df = df[df[col_nss].str.len() >= 3].copy()

    enc_str = " ".join(str(v).upper() for v in df_raw.values.flatten() if pd.notna(v))
    if "HIVICO" in enc_str:    empresa_excel = "HIVICO"
    elif "FIREWALL" in enc_str: empresa_excel = "FIREWALL"
    else:                       empresa_excel = "SIN EMPRESA"

    _pares = {}
    for c in df.columns:
        u = str(c).strip().upper()
        m = re.match(r'^MD(\d+)$', u)
        if m: _pares.setdefault(int(m.group(1)), [None, None])[0] = c
        m = re.match(r'^MN(\d+)$', u)
        if m: _pares.setdefault(int(m.group(1)), [None, None])[1] = c
    pares_md = [(md, mn) for _, (md, mn) in
                sorted(_pares.items(), key=lambda x: (x[0]==0, x[0])) if md]

    registros = []
    for _, row in df.iterrows():
        nss  = str(row.get(col_nss, "") or "").strip().replace(".0","")
        curp = str(row.get(col_curp, "") or "").strip() if col_curp else ""
        if curp == "nan": curp = ""

        clabe_raw  = row.get(col_clabe) if col_clabe else None
        banco_val  = None
        cuenta_raw = None
        for _c, _b in [(col_bbva,'BBVA'),(col_banorte,'BANORTE'),(col_san,'SANTANDER')]:
            if _c:
                _v = row.get(_c)
                if _v and str(_v).strip() not in ('','nan','None'):
                    cuenta_raw = _v; banco_val = _b; break
        if not cuenta_raw and col_cuenta:
            cuenta_raw = row.get(col_cuenta)

        clabe_val, tipo = limpiar_clabe(clabe_raw)
        cuenta_val, _   = limpiar_clabe(cuenta_raw)
        if tipo == "CUENTA":
            cuenta_val = clabe_val; clabe_val = None

        if col_ap:
            ap  = str(row.get(col_ap,"") or "").strip()
            am  = str(row.get(col_am,"") or "").strip() if col_am  else ""
            nn  = str(row.get(col_noms,"") or "").strip() if col_noms else ""
            ap  = "" if ap == "nan" else ap
            am  = "" if am == "nan" else am
            nn  = "" if nn == "nan" else nn
            nombre = " ".join(x for x in [ap,am,nn] if x).strip()
        elif col_nombcom:
            nombre = str(row.get(col_nombcom,"") or "").strip()
        elif col_noms:
            nombre = str(row.get(col_noms,"") or "").strip()
        elif col_mediored:
            nombre = str(row.get(col_mediored,"") or "").strip()
        else:
            nombre = ""
        if nombre in ("nan","None"): nombre = ""

        servicio = str(row.get(col_serv,"") or "").strip() if col_serv else meta["sucursal"]
        if servicio == "nan": servicio = meta["sucursal"]

        if col_banco:
            b = str(row.get(col_banco,"")).upper()
            for k, v in BANCOS_MAP.items():
                if k in b: banco_val = v; break
        if not banco_val: banco_val = inferir_banco(clabe_val)
        if not banco_val and cuenta_val: banco_val = inferir_banco(cuenta_val)

        total  = limpiar_importe(row.get(col_total))  if col_total  else 0.0
        sueldo = limpiar_importe(row.get(col_sueldo)) if col_sueldo else 0.0
        turno  = str(row.get(col_turno,"") or "").strip() if col_turno else ""
        if turno == "nan": turno = ""
        fecha  = limpiar_fecha(row.get(col_fecha)) if col_fecha else None

        dias = faltas = 0
        # Primero intentar columna X (días directos, tope 16)
        if col_x is not None:
            try:
                _dx = float(str(row.get(col_x) or 0))
                dias = min(int(round(_dx)), MAX_DIAS_QNA)
            except:
                dias = 0
        # Fallback: calcular desde códigos MD si col_x no dio resultado
        if dias == 0 and pares_md:
            for col_md, col_mn in pares_md:
                try: vd = int(float(str(row.get(col_md) or 0)))
                except: vd = 0
                try: vn = int(float(str(row.get(col_mn) or 0))) if col_mn else 0
                except: vn = 0
                if vd in TRABAJADO_MD or vn in TRABAJADO_MD: dias += 1
                elif vd in FALTA_MD or vn in FALTA_MD: faltas += 1

        tiene_algo = (nss and len(nss) >= 3) or clabe_val or cuenta_val
        if not tiene_algo: continue

        registros.append({
            "nss": nss, "curp": curp, "nombre_completo": nombre,
            "clave_interbancaria": clabe_val, "numero_cuenta": cuenta_val,
            "banco": banco_val, "nombre_servicio": servicio,
            "estado": meta["sucursal"], "total_quincenal": total,
            "sueldo_mensual": sueldo, "turno": turno, "fecha_alta": fecha,
            "nss_curp_completo": bool(nss and curp and len(curp) == 18),
            "archivo_origen": nombre_archivo, "periodo": meta["periodo"],
            "empresa": empresa_excel, "dias_trabajados": dias,
            "faltas": faltas, "quincena": qna, "hoja_detectada": hoja,
        })

    if not registros:
        return None, f"Sin registros válidos en {nombre_archivo} (hoja: {hoja})"
    return pd.DataFrame(registros), None


# ── Helper: extraer archivos Excel de ZIP o lista directa ──────
def extraer_archivos_excel(uploaded_files):
    """
    Acepta lista de archivos subidos (xlsx/xls) o un ZIP con xlsx adentro.
    Devuelve lista de tuplas (nombre, bytes).
    """
    import zipfile
    resultado = []
    for f in (uploaded_files or []):
        if f.name.lower().endswith('.zip'):
            try:
                with zipfile.ZipFile(io.BytesIO(f.read())) as z:
                    for nombre in z.namelist():
                        if nombre.lower().endswith(('.xlsx','.xls')) and not nombre.startswith('__'):
                            with z.open(nombre) as xf:
                                resultado.append((os.path.basename(nombre), xf.read()))
            except Exception as e:
                st.warning(f"Error leyendo ZIP {f.name}: {e}")
        elif f.name.lower().endswith(('.xlsx','.xls')):
            resultado.append((f.name, f.read()))
    return resultado

# ── Exportar sábanas Excel ─────────────────────────────────────
def exportar_sabanas_excel(df_all, periodo):
    wb = Workbook(); C_H="1B3A6B"; C_S="2E5FA3"
    ws1 = wb.active; ws1.title = "RESUMEN POR ARCHIVO"
    xl_titulo(ws1, f"RESUMEN DE SÁBANAS — {periodo}", 1, 6, C_H)
    ws1.merge_cells("A2:F2")
    ws1["A2"].value = (f"Total empleados: {len(df_all):,}  |  Con CLABE: {df_all['clave_interbancaria'].notna().sum():,}  |  "
                       f"Sin CLABE: {df_all['clave_interbancaria'].isna().sum():,}  |  Total nómina: ${df_all['total_quincenal'].sum():,.2f}")
    ws1["A2"].font = xl_fn(size=9, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="center")
    hdrs = ["ARCHIVO","QUINCENA","EMPRESA","EMPLEADOS","TOTAL QUINCENAL","SIN CLABE"]
    ws1.append(hdrs); fr = ws1.max_row
    for c in ws1[fr]:
        c.font = xl_fn(bold=True, color="FFFFFF"); c.fill = xl_fill(C_S); c.border = xl_brd()
        c.alignment = Alignment(horizontal="center")
    gran_total = 0.0
    for arch in sorted(df_all["archivo_origen"].unique()):
        sub = df_all[df_all["archivo_origen"] == arch]
        t = sub["total_quincenal"].sum(); gran_total += t
        sc = sub["clave_interbancaria"].isna().sum()
        ws1.append([arch, "Q"+str(sub["quincena"].iloc[0]), sub["empresa"].iloc[0],
                    len(sub), t, int(sc)])
        fr = ws1.max_row
        for c in ws1[fr]: c.border = xl_brd(); c.font = xl_fn(size=9)
        ws1.cell(fr,5).number_format = '$#,##0.00'
        for c in ws1[fr]: c.fill = xl_fill("EAF7EE" if sc==0 else "FFFBEB")
    ws1.append(["TOTAL GENERAL","","",len(df_all),gran_total,
                int(df_all["clave_interbancaria"].isna().sum())])
    fr = ws1.max_row
    for c in ws1[fr]: c.font=xl_fn(bold=True,size=10); c.fill=xl_fill("E8EFF8"); c.border=xl_brd()
    ws1.cell(fr,5).number_format='$#,##0.00'
    for col,w in zip("ABCDEF",[48,8,12,10,18,10]): ws1.column_dimensions[col].width=w
    ws1.freeze_panes="A4"
    # Hoja detalle
    ws2=wb.create_sheet("DETALLE"); xl_titulo(ws2,f"DETALLE EMPLEADOS — {periodo}",1,12,C_H)
    h2=["NSS","CURP","NOMBRE","BANCO","CLABE","CUENTA","SERVICIO",
        "TOTAL Q","SUELDO","DÍAS","FALTAS","ARCHIVO"]
    ws2.append(h2); fr=ws2.max_row
    for c in ws2[fr]: c.font=xl_fn(bold=True,color="FFFFFF",size=8); c.fill=xl_fill(C_S); c.border=xl_brd()
    BCOL={"BBVA":"D6E4F0","BANORTE":"E8F5E9","SANTANDER":"FFF3E0"}
    for _,r in df_all.iterrows():
        ws2.append([r.get("nss"),r.get("curp"),r.get("nombre_completo"),
                    r.get("banco"),r.get("clave_interbancaria"),r.get("numero_cuenta"),
                    r.get("nombre_servicio"),r.get("total_quincenal"),r.get("sueldo_mensual"),
                    r.get("dias_trabajados"),r.get("faltas"),r.get("archivo_origen")])
        fr=ws2.max_row
        for c in ws2[fr]: c.border=xl_brd(); c.font=xl_fn(size=8)
        xl_mfmt(ws2,fr,[8,9])
        bg=BCOL.get(str(r.get("banco","")).upper()[:6],"FFFFFF")
        for c in ws2[fr]: c.fill=xl_fill(bg)
    for i,w in enumerate([14,20,35,12,22,14,35,14,14,8,8,40],1):
        ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.freeze_panes="A3"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ── BD helpers ─────────────────────────────────────────────────
def bd_conectar(): return psycopg2.connect(**DB)

def bd_get_banco(cur, n):
    if not n: return None
    cur.execute("SELECT id_banco FROM bancos WHERE nombre_banco=%s",(n,))
    r=cur.fetchone(); return r[0] if r else None

def bd_get_servicio(cur, nombre, estado):
    if not nombre: return None
    nombre=str(nombre).strip()[:60]; estado=str(estado or "SIN ESTADO").strip()[:40]
    cur.execute("""INSERT INTO servicios(nombre_cliente,estado) VALUES(%s,%s)
        ON CONFLICT(nombre_cliente,estado) DO UPDATE SET nombre_cliente=EXCLUDED.nombre_cliente
        RETURNING id_servicio""",(nombre,estado))
    return cur.fetchone()[0]

def bd_insertar_periodo(cur,nombre_per,fecha_pago,tipo,mes,anio):
    cur.execute("""INSERT INTO periodos(nombre_periodo,fecha_pago,tipo_pago,mes,anio,autorizado)
        VALUES(%s,%s,%s,%s,%s,TRUE)
        ON CONFLICT(fecha_pago,tipo_pago) DO UPDATE SET nombre_periodo=EXCLUDED.nombre_periodo,autorizado=TRUE
        RETURNING id_periodo""",(nombre_per,fecha_pago,tipo,mes,anio))
    return cur.fetchone()[0]

def bd_insertar_empleado(cur,row):
    nss=str(row.get("nss","") or "").strip()[:20]
    curp=str(row.get("curp","") or "").strip()
    if not curp or len(curp)!=18:
        ident=str(row.get("clave_interbancaria") or row.get("numero_cuenta") or nss or "SIN")
        curp=f"SINCURP_{ident[:10].strip()}"
    clabe=str(row.get("clave_interbancaria") or "").strip() or None
    cuenta=str(row.get("numero_cuenta") or "").strip() or None
    nombre=str(row.get("nombre_completo","") or "").strip()[:120]
    turno=str(row.get("turno","") or "").strip()[:10] or None
    sueldo=float(row.get("sueldo_mensual") or 0)
    fecha=row.get("fecha_alta")
    nss_ok=bool(row.get("nss_curp_completo",False))
    id_banco=bd_get_banco(cur,row.get("banco"))
    id_serv=bd_get_servicio(cur,row.get("nombre_servicio"),row.get("estado"))
    if not nss and not clabe and not cuenta: return None
    cur.execute("""INSERT INTO empleados
        (nss,curp,nombre_completo,clave_interbancaria,numero_cuenta,
         id_banco,id_servicio,turno,sueldo_mensual,fecha_alta,nss_curp_completo)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT(nss,curp) DO UPDATE SET
            nombre_completo=CASE WHEN EXCLUDED.nombre_completo!=''
                THEN EXCLUDED.nombre_completo ELSE empleados.nombre_completo END,
            clave_interbancaria=COALESCE(EXCLUDED.clave_interbancaria,empleados.clave_interbancaria),
            numero_cuenta=COALESCE(EXCLUDED.numero_cuenta,empleados.numero_cuenta),
            sueldo_mensual=EXCLUDED.sueldo_mensual
        RETURNING id_empleado""",
        (nss,curp,nombre,clabe,cuenta,id_banco,id_serv,turno,sueldo,fecha,nss_ok))
    return cur.fetchone()[0]

def bd_insertar_nomina(cur,id_emp,id_per,row):
    id_banco=bd_get_banco(cur,row.get("banco"))
    id_serv=bd_get_servicio(cur,row.get("nombre_servicio"),row.get("estado"))
    clabe=str(row.get("clave_interbancaria") or "").strip() or None
    cuenta=str(row.get("numero_cuenta") or "").strip() or None
    total=float(row.get("total_quincenal") or 0)
    sueldo=float(row.get("sueldo_mensual") or 0)
    turno=str(row.get("turno","") or "").strip()[:10] or None
    arch=str(row.get("archivo_origen","") or "").strip()[:100]
    cur.execute("""INSERT INTO nomina
        (id_empleado,id_periodo,id_banco,id_servicio,
         clave_interbancaria,numero_cuenta,total_quincenal,sueldo_mensual,turno,archivo_origen)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT(id_empleado,id_periodo) DO UPDATE SET
            total_quincenal=EXCLUDED.total_quincenal,
            clave_interbancaria=COALESCE(EXCLUDED.clave_interbancaria,nomina.clave_interbancaria),
            numero_cuenta=COALESCE(EXCLUDED.numero_cuenta,nomina.numero_cuenta)""",
        (id_emp,id_per,id_banco,id_serv,clabe,cuenta,total,sueldo,turno,arch))

# ── Cruce banco (lógica exacta script 08) ─────────────────────
def norma_c(s): return re.sub(r'[^\d]','',str(s or '')).strip()

def banco_hex(banco):
    b=str(banco or '').upper()
    if 'BBVA' in b:      return "D6E4F0"
    if 'BANORTE' in b:   return "E8F5E9"
    if 'SANTANDER' in b: return "FFF3E0"
    return "F2F2F2"

def qna_pdf(nombre_archivo, r1a, r2a):
    base=os.path.basename(str(nombre_archivo or ''))
    m=re.match(r'^(\d{2})(\d{2})(\d{2})',base)
    if not m: return None
    try: fecha=(int(m.group(1))+2000,int(m.group(2)),int(m.group(3)))
    except: return None
    if r1a['desde']<=fecha<=r1a['hasta']: return '1A'
    if r2a['desde']<=fecha<=r2a['hasta']: return '2A'
    return None

def hacer_cruce_banco(nomina_rows, pdf_rows, tolerancia, rango_1a, rango_2a):
    totales_pdf=defaultdict(lambda:{'banco':'','registros':0,'importe':0.0})
    pdf_por_cuenta=defaultdict(list)
    for banco_p,cuenta,imp,nombre_e,estatus,arch in pdf_rows:
        n=norma_c(cuenta)
        if not n: continue
        reg={'banco':str(banco_p or ''),'cuenta':cuenta or '','imp':float(imp or 0),
             'nombre':nombre_e or '','estatus':estatus or '','arch':arch or ''}
        pdf_por_cuenta[n].append(reg)
        totales_pdf[arch]['banco']=str(banco_p or '')
        totales_pdf[arch]['registros']+=1
        totales_pdf[arch]['importe']+=float(imp or 0)
    dobles_raw={k:v for k,v in pdf_por_cuenta.items() if len(v)>=2}

    idx={'1A':{'exact':defaultdict(list),'sufijo':defaultdict(list)},
         '2A':{'exact':defaultdict(list),'sufijo':defaultdict(list)},
         None:{'exact':defaultdict(list),'sufijo':defaultdict(list)}}
    for banco_p,cuenta,imp,nombre_e,estatus,arch in pdf_rows:
        n=norma_c(cuenta)
        if not n: continue
        q=qna_pdf(arch,rango_1a,rango_2a)
        reg={'banco':str(banco_p or ''),'cuenta':cuenta or '','imp':float(imp or 0),
             'nombre':nombre_e or '','estatus':estatus or '','arch':arch or '','qna_pdf':q}
        for dest in [q,None]:
            for clave in [n,n.lstrip('0')]: idx[dest]['exact'][clave].append(reg)
            for largo in [16,11,10]:
                if len(n)>=largo: idx[dest]['sufijo'][n[-largo:]].append(reg)

    def buscar(cuenta,clabe,importe,qna_nom):
        for dest in [qna_nom,None]:
            ix=idx[dest]
            for campo,metodo in [(cuenta,'CUENTA'),(clabe,'CLABE')]:
                nc=norma_c(campo)
                if not nc: continue
                for clave in [nc,nc.lstrip('0')]:
                    hits=ix['exact'].get(clave,[])
                    if hits: return min(hits,key=lambda x:abs(x['imp']-importe)),metodo
                for largo in [16,11,10]:
                    if len(nc)>=largo:
                        hits=ix['sufijo'].get(nc[-largo:],[])
                        if hits: return min(hits,key=lambda x:abs(x['imp']-importe)),metodo
        return None,None

    from collections import OrderedDict as _OD
    _grp = _OD()
    for row in nomina_rows:
        nss,nombre,cuenta,clabe,banco_e,imp_nom,quincena,arch_orig=row
        imp_nom=float(imp_nom or 0)
        _k=(norma_c(nss or ''),norma_c(cuenta or '') or norma_c(clabe or ''))
        if _k not in _grp:
            _grp[_k]={'nss':nss,'nombre':nombre,'cuenta':cuenta,'clabe':clabe,
                      'banco_e':banco_e,'imp_nom':imp_nom,'quincena':quincena,
                      'archs':[arch_orig or ''],'n_sab':1}
        else:
            _grp[_k]['imp_nom']+=imp_nom; _grp[_k]['n_sab']+=1
            if arch_orig and arch_orig not in _grp[_k]['archs']:
                _grp[_k]['archs'].append(arch_orig)
            if not _grp[_k]['nombre'] and nombre: _grp[_k]['nombre']=nombre
            if not _grp[_k]['cuenta'] and cuenta: _grp[_k]['cuenta']=cuenta
            if not _grp[_k]['clabe']  and clabe:  _grp[_k]['clabe'] =clabe
    resultados=[]; stats=defaultdict(int)
    for _k, _g in _grp.items():
        nss=_g['nss']; nombre=_g['nombre']; cuenta=_g['cuenta']; clabe=_g['clabe']
        banco_e=_g['banco_e']; imp_nom=_g['imp_nom']; quincena=_g['quincena']
        arch_orig=' | '.join(_g['archs']); n_sab=_g['n_sab']
        qna_nom='1A' if '1' in str(quincena) else '2A'
        match,metodo=buscar(cuenta,clabe,imp_nom,qna_nom)
        # Verificar doble abono antes de calcular diff
        clave_norm = norma_c(cuenta or '') or norma_c(clabe or '')
        tiene_doble = clave_norm in dobles_raw
        imp_pdf = None
        arch_pdf_final = ''
        if match:
            if tiene_doble and clave_norm:
                # Sumar TODOS los importes PDF de esa cuenta
                imp_pdf       = sum(a['imp'] for a in dobles_raw[clave_norm])
                arch_pdf_final = ' | '.join(sorted(set(a['arch'] for a in dobles_raw[clave_norm])))
                match = dict(match)
                match['arch'] = arch_pdf_final
            else:
                imp_pdf        = match['imp']
                arch_pdf_final = match['arch']
            diff   = imp_nom - imp_pdf
            estatus= "OK" if abs(diff)<=tolerancia else f"DIFIERE ({diff:+,.2f})"
            stats['OK' if abs(diff)<=tolerancia else 'DIFIERE']+=1
        elif tiene_doble and clave_norm:
            # Tiene doble abono en PDF pero buscar() no lo encontró por índice — sumar directo
            imp_pdf        = sum(a['imp'] for a in dobles_raw[clave_norm])
            arch_pdf_final = ' | '.join(sorted(set(a['arch'] for a in dobles_raw[clave_norm])))
            match          = {'banco': dobles_raw[clave_norm][0]['banco'],
                              'nombre': dobles_raw[clave_norm][0]['nombre'],
                              'estatus': '', 'arch': arch_pdf_final}
            metodo = 'CUENTA'
            diff   = imp_nom - imp_pdf
            estatus= "OK" if abs(diff)<=tolerancia else f"DIFIERE ({diff:+,.2f})"
            stats['OK' if abs(diff)<=tolerancia else 'DIFIERE']+=1
        else:
            imp_pdf=None; arch_pdf_final=''
            match={'banco':'','nombre':'','estatus':'','arch':''}
            metodo='—'; diff=None; estatus="NO EN PDF"; stats['NO EN PDF']+=1
        _obs = f'En {n_sab} sábanas: {arch_orig}' if n_sab>1 else ''
        if tiene_doble and not _obs: _obs='⚠ Doble abono PDF'
        resultados.append({
            'qna':qna_nom,'quincena':quincena or '',
            'nss':nss or '','nombre':nombre or '',
            'banco':banco_e or '','cuenta':cuenta or '','clabe':clabe or '',
            'imp_nom':imp_nom,'imp_pdf':imp_pdf,
            'diff':diff,'banco_pdf':match['banco'],'nombre_pdf':match['nombre'],
            'estatus_pdf':match['estatus'],'arch_pdf':arch_pdf_final,
            'metodo':metodo,'estatus':estatus,'arch_origen':arch_orig or '',
            'doble_abono':tiene_doble,'n_sab':n_sab,'obs':_obs,
        })
    return resultados,stats,dobles_raw,totales_pdf

# ── Exportar nómina vs PDF Excel (estructura exacta script 08) ─
def exportar_nomina_pdf_excel(resultados,stats,dobles_raw,totales_pdf,empresa,reg_pat,periodo,fecha_pago,tolerancia):
    wb=Workbook()
    CV="C6EFCE";CR="FFC7CE";CA="FFEB9C";CG="F2F2F2";CH="1F4E79";CN="FCE4D6"
    hdrs_det=["QNA","QUINCENA","NSS","NOMBRE","BANCO NÓM","CUENTA","CLABE",
              "IMPORTE NÓMINA","IMPORTE PDF","DIFERENCIA","BANCO PDF","NOMBRE PDF",
              "ESTATUS BANCO","MÉTODO","ESTATUS","DOBLE ABONO","ARCHIVO SÁBANA","ARCHIVO PDF","OBSERVACIONES"]
    tot_nom=sum(r['imp_nom'] for r in resultados)
    tot_pdf=sum(r['imp_pdf'] for r in resultados if r['imp_pdf'])
    total=len(resultados) or 1

    # HOJA 1 — RESUMEN
    ws1=wb.active; ws1.title="RESUMEN"
    xl_titulo(ws1,f"CRUCE NÓMINA vs PDF BANCO — {empresa}",1,10,CH)
    ws1.merge_cells("A2:J2")
    ws1["A2"].value=(f"Reg. Patronal: {reg_pat}  |  Periodo: {fecha_pago}  |  "
                     f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Tolerancia: ${tolerancia:.2f}")
    ws1["A2"].font=xl_fn(size=9,color="555555"); ws1["A2"].alignment=Alignment(horizontal="center")
    ws1.append([])
    ws1.append(["","NÓMINA","PDF BANCO","DIFERENCIA","OK","DIFIERE","NO EN PDF","% OK","DOBLE ABONO",""])
    fr=ws1.max_row
    for c in ws1[fr]: c.font=xl_fn(bold=True,color="FFFFFF"); c.fill=xl_fill(CH); c.border=xl_brd()
    ws1.append(["TOTAL GENERAL",tot_nom,tot_pdf,tot_nom-tot_pdf,
                stats['OK'],stats['DIFIERE'],stats['NO EN PDF'],
                stats['OK']/total*100,len(dobles_raw),""])
    fr=ws1.max_row
    for c in ws1[fr]: c.font=xl_fn(bold=True); c.border=xl_brd()
    xl_mfmt(ws1,fr,[2,3,4]); ws1.cell(fr,8).number_format='0.0"%"'
    bg=CV if abs(tot_nom-tot_pdf)<5000 else CR
    for c in ws1[fr]: c.fill=xl_fill(bg)
    ws1.append([])
    ws1.append(["BANCO","EMPLEADOS","NÓMINA","PDF","DIFERENCIA","OK","DIFIERE","NO EN PDF","% OK",""])
    fr=ws1.max_row
    for c in ws1[fr]: c.font=xl_fn(bold=True,color="FFFFFF"); c.fill=xl_fill(CH); c.border=xl_brd()
    grp=defaultdict(lambda:{'emp':0,'nom':0.0,'pdf':0.0,'ok':0,'dif':0,'nopdf':0})
    for r in resultados:
        k=r['banco'] or 'SIN BANCO'
        grp[k]['emp']+=1; grp[k]['nom']+=r['imp_nom']; grp[k]['pdf']+=r['imp_pdf'] or 0
        if r['estatus']=='OK': grp[k]['ok']+=1
        elif 'DIFIERE' in r['estatus']: grp[k]['dif']+=1
        else: grp[k]['nopdf']+=1
    for banco,v in sorted(grp.items()):
        diff_b=v['nom']-v['pdf']; pct=v['ok']/v['emp']*100 if v['emp'] else 0
        ws1.append([banco,v['emp'],v['nom'],v['pdf'],diff_b,v['ok'],v['dif'],v['nopdf'],pct,""])
        fr=ws1.max_row
        for c in ws1[fr]: c.border=xl_brd(); c.font=xl_fn(size=9)
        xl_mfmt(ws1,fr,[3,4,5]); ws1.cell(fr,9).number_format='0.0"%"'
        for c in ws1[fr]: c.fill=xl_fill(banco_hex(banco))
        ws1.cell(fr,5).fill=xl_fill(CV if abs(diff_b)<1000 else CA if abs(diff_b)<50000 else CR)
    for col,w in zip("ABCDEFGHIJ",[22,10,16,16,14,7,7,9,8,6]):
        ws1.column_dimensions[col].width=w

    # HOJA 2 — TOTALES POR PDF
    ws_pdf=wb.create_sheet("TOTALES POR PDF")
    xl_titulo(ws_pdf,f"TOTALES POR ARCHIVO PDF — {fecha_pago}",1,5,"1A5276")
    gran_total_pdf=sum(v['importe'] for v in totales_pdf.values())
    ws_pdf.merge_cells("A2:E2")
    ws_pdf["A2"].value=(f"Archivos: {len(totales_pdf)}  |  "
                        f"Registros: {sum(v['registros'] for v in totales_pdf.values()):,}  |  "
                        f"Total: ${gran_total_pdf:,.2f}")
    ws_pdf["A2"].font=xl_fn(bold=True,size=10,color="1A5276")
    ws_pdf["A2"].alignment=Alignment(horizontal="center")
    ws_pdf.append(["ARCHIVO PDF","BANCO","REGISTROS","IMPORTE TOTAL","% DEL TOTAL"])
    fr=ws_pdf.max_row
    for c in ws_pdf[fr]:
        c.font=xl_fn(bold=True,color="FFFFFF",size=9)
        c.fill=xl_fill("1A5276"); c.border=xl_brd(); c.alignment=Alignment(horizontal="center")
    for arch,v in sorted(totales_pdf.items(),key=lambda x:x[1]['banco']):
        pct=v['importe']/gran_total_pdf*100 if gran_total_pdf else 0
        ws_pdf.append([arch,v['banco'],v['registros'],v['importe'],pct])
        fr=ws_pdf.max_row
        for c in ws_pdf[fr]: c.border=xl_brd(); c.font=xl_fn(size=9)
        ws_pdf.cell(fr,4).number_format='$#,##0.00'; ws_pdf.cell(fr,5).number_format='0.0"%"'
        for c in ws_pdf[fr]: c.fill=xl_fill(banco_hex(v['banco']))
    ws_pdf.append(["TOTAL","",sum(v['registros'] for v in totales_pdf.values()),gran_total_pdf,100.0])
    fr=ws_pdf.max_row
    for c in ws_pdf[fr]: c.font=xl_fn(bold=True,size=10); c.fill=xl_fill(CG); c.border=xl_brd()
    ws_pdf.cell(fr,4).number_format='$#,##0.00'
    ws_pdf.append([]); ws_pdf.append([])
    fs=ws_pdf.max_row
    ws_pdf.merge_cells(f"A{fs}:E{fs}")
    ws_pdf.cell(fs,1,"TOTALES POR ARCHIVO DE NÓMINA").font=Font(bold=True,color="FFFFFF",size=11,name="Calibri")
    ws_pdf.cell(fs,1).fill=xl_fill("145A32")
    ws_pdf.cell(fs,1).alignment=Alignment(horizontal="center",vertical="center")
    ws_pdf.row_dimensions[fs].height=22
    ws_pdf.append(["ARCHIVO NÓMINA","EMPLEADOS","IMPORTE NÓMINA","% DEL TOTAL",""])
    fr=ws_pdf.max_row
    for c in ws_pdf[fr]:
        c.font=xl_fn(bold=True,color="FFFFFF",size=9)
        c.fill=xl_fill("145A32"); c.border=xl_brd(); c.alignment=Alignment(horizontal="center")
    totales_nom=defaultdict(lambda:{'empleados':0,'importe':0.0})
    for r in resultados:
        k=r['arch_origen'] or 'SIN ARCHIVO'
        totales_nom[k]['empleados']+=1; totales_nom[k]['importe']+=r['imp_nom']
    gran_total_nom=sum(v['importe'] for v in totales_nom.values())
    for arch,v in sorted(totales_nom.items()):
        pct=v['importe']/gran_total_nom*100 if gran_total_nom else 0
        ws_pdf.append([arch,v['empleados'],v['importe'],pct,""])
        fr=ws_pdf.max_row
        for c in ws_pdf[fr]: c.border=xl_brd(); c.font=xl_fn(size=9); c.fill=xl_fill(CV)
        ws_pdf.cell(fr,3).number_format='$#,##0.00'; ws_pdf.cell(fr,4).number_format='0.0"%"'
    ws_pdf.append(["TOTAL",sum(v['empleados'] for v in totales_nom.values()),gran_total_nom,100.0,""])
    fr=ws_pdf.max_row
    for c in ws_pdf[fr]: c.font=xl_fn(bold=True,size=10); c.fill=xl_fill(CG); c.border=xl_brd()
    ws_pdf.cell(fr,3).number_format='$#,##0.00'
    for col,w in zip("ABCDE",[40,14,12,16,10]): ws_pdf.column_dimensions[col].width=w
    ws_pdf.freeze_panes="A4"

    # HOJA 3 — DETALLE COMPLETO
    def hoja_det(wb_o,regs,nom,bg_h):
        if not regs: return
        ws=wb_o.create_sheet(nom)
        xl_titulo(ws,f"{nom} — NÓMINA vs PDF",1,len(hdrs_det),bg_h)
        ws.merge_cells(f"A2:{get_column_letter(len(hdrs_det))}2")
        ws["A2"].value=(f"Total: {len(regs):,}  |  Nómina: ${sum(r['imp_nom'] for r in regs):,.2f}  |  "
                        f"OK: {sum(1 for r in regs if r['estatus']=='OK'):,}  |  "
                        f"No en PDF: {sum(1 for r in regs if r['estatus']=='NO EN PDF'):,}  |  "
                        f"Doble: {sum(1 for r in regs if r['doble_abono']):,}")
        ws["A2"].font=xl_fn(bold=True,size=10,color=bg_h)
        ws["A2"].alignment=Alignment(horizontal="center")
        ws.append(hdrs_det); fr=ws.max_row
        for c in ws[fr]:
            c.font=xl_fn(bold=True,color="FFFFFF",size=8); c.fill=xl_fill(bg_h); c.border=xl_brd()
            c.alignment=Alignment(horizontal="center",wrap_text=True,vertical="center")
        ws.row_dimensions[fr].height=35
        for r in sorted(regs,key=lambda x:(x['banco'],x['nombre'])):
            est=r['estatus']
            _obs_v = r.get('obs','') or ''
            if r['doble_abono'] and not _obs_v: _obs_v = '⚠ Doble abono PDF'
            row=[r['qna'],r['quincena'],r['nss'],r['nombre'],r['banco'],r['cuenta'],r['clabe'],
                 r['imp_nom'],r['imp_pdf'],r['diff'],r['banco_pdf'],r['nombre_pdf'],r['estatus_pdf'],
                 r['metodo'],est,'⚠ DOBLE' if r['doble_abono'] else '',r['arch_origen'],r['arch_pdf'],_obs_v]
            ws.append(row); fr=ws.max_row
            for c in ws[fr]: c.border=xl_brd(); c.font=xl_fn(size=8)
            xl_mfmt(ws,fr,[8,9,10])
            for c in ws[fr]: c.fill=xl_fill(banco_hex(r['banco']))
            ce=hdrs_det.index('ESTATUS')+1
            ws.cell(fr,ce).fill=xl_fill(CV if est=='OK' else CR if est=='NO EN PDF' else CA)
            ws.cell(fr,ce).font=xl_fn(bold=True,size=8)
            if r['doble_abono']:
                cd=hdrs_det.index('DOBLE ABONO')+1
                ws.cell(fr,cd).fill=xl_fill("FF0000")
                ws.cell(fr,cd).font=xl_fn(bold=True,size=8,color="FFFFFF")
        ws.append(["TOTAL","","","","","","",
                   sum(r['imp_nom'] for r in regs),
                   sum(r['imp_pdf'] for r in regs if r['imp_pdf']),
                   "","","","","","","","","",""])
        fr=ws.max_row
        for c in ws[fr]: c.font=xl_fn(bold=True,size=9); c.fill=xl_fill(CG); c.border=xl_brd()
        xl_mfmt(ws,fr,[8,9])
        for i,w in enumerate([6,18,14,30,14,16,22,14,14,12,14,28,14,10,20,14,40,36],1):
            if i<=ws.max_column: ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A4"

    hoja_det(wb,resultados,'DETALLE COMPLETO',CH)

    # HOJA 4 — DOBLE ABONO
    if dobles_raw:
        ws_da=wb.create_sheet("DOBLE ABONO")
        xl_titulo(ws_da,"CUENTAS CON DOBLE ABONO EN PDF",1,12,"9C0006")
        ws_da.merge_cells("A2:L2")
        ws_da["A2"].value=f"{len(dobles_raw):,} cuentas con 2 o más pagos — revisar PDFs"
        ws_da["A2"].font=xl_fn(bold=True,size=10,color="9C0006")
        ws_da["A2"].alignment=Alignment(horizontal="center")
        hda=["CUENTA","NSS","NOMBRE","BANCO","# ABONOS","TOTAL ABONADO","PDF 1","IMP 1","PDF 2","IMP 2","PDF 3","IMP 3"]
        ws_da.append(hda); fr=ws_da.max_row
        for c in ws_da[fr]:
            c.font=xl_fn(bold=True,color="FFFFFF",size=9); c.fill=xl_fill("9C0006"); c.border=xl_brd()
        nom_por_c={}
        for r in resultados:
            ck=norma_c(r['cuenta'] or r['clabe'] or '')
            if ck and ck not in nom_por_c:
                nom_por_c[ck]={'nss':r['nss'],'nombre':r['nombre'],'banco':r['banco']}
        for cn,abonos in sorted(dobles_raw.items()):
            nom=nom_por_c.get(cn,{})
            row=[abonos[0]['cuenta'],nom.get('nss',''),
                 nom.get('nombre',abonos[0]['nombre']),nom.get('banco',abonos[0]['banco']),
                 len(abonos),sum(a['imp'] for a in abonos)]
            for i in range(3):
                row.append(abonos[i]['arch'] if i<len(abonos) else '')
                row.append(abonos[i]['imp']  if i<len(abonos) else '')
            ws_da.append(row); fr=ws_da.max_row
            for c in ws_da[fr]: c.border=xl_brd(); c.font=xl_fn(size=9)
            for c in ws_da[fr]: c.fill=xl_fill(banco_hex(nom.get('banco','')))
            ws_da.cell(fr,6).fill=xl_fill(CR); ws_da.cell(fr,6).font=xl_fn(bold=True,size=9)
            for ci in [6,8,10,12]: ws_da.cell(fr,ci).number_format='$#,##0.00'

    # HOJA 5 — NO EN PDF
    hoja_det(wb,[r for r in resultados if r['estatus']=='NO EN PDF'],'NO EN PDF',"9C0006")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ── PDF ejecutivo (reportlab) ──────────────────────────────────
def generar_pdf_ejecutivo(res,stats,dobles,tot_nom,tot_pdf,periodo,fecha_pago):
    if not REPORTLAB_OK: return None
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle,HRFlowable
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER,TA_RIGHT

        buf=io.BytesIO()
        doc=SimpleDocTemplate(buf,pagesize=A4,
                               leftMargin=2.2*cm,rightMargin=2.2*cm,
                               topMargin=2*cm,bottomMargin=2*cm)
        CA=colors.HexColor
        AZ=CA('#0B1929'); AZ2=CA('#1B3A6B'); AZ3=CA('#4A90D9')
        VD=CA('#27AE60'); RJ=CA('#E74C3C'); AM=CA('#E67E22')
        GR=CA('#E8EFF8'); GR2=CA('#F5F8FC'); BD=CA('#D0DCF0')

        ST=lambda n,**kw: ParagraphStyle(n,fontName='Helvetica',fontSize=9,
                                          textColor=CA('#4A7A9B'),leading=13,**kw)
        s_titulo =ST('t',fontName='Helvetica-Bold',fontSize=20,textColor=AZ,spaceAfter=3,leading=24)
        s_sub    =ST('s',fontSize=8,spaceAfter=2)
        s_sect   =ST('sc',fontName='Helvetica-Bold',fontSize=7,textColor=AZ3,
                     spaceBefore=14,spaceAfter=6,letterSpacing=1.5)
        s_footer =ST('f',fontSize=6.5,textColor=CA('#8A9BB0'),alignment=TA_CENTER)

        total_r=len(res) or 1
        pct_ok=stats['OK']/total_r*100
        diff=tot_nom-tot_pdf

        story=[]
        story.append(Paragraph("HIVICO / FIREWALL SEGURIDAD PRIVADA S.A. DE C.V.",
                                ST('c',fontSize=7,textColor=AZ3,letterSpacing=1.2)))
        story.append(Paragraph("Reporte de Dispersión de Nómina",s_titulo))
        story.append(Paragraph(
            f"Periodo: {periodo}  ·  Fecha de pago: {fecha_pago}  ·  "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",s_sub))
        story.append(HRFlowable(width="100%",thickness=2,color=AZ2,spaceAfter=16))

        # Resumen
        story.append(Paragraph("RESUMEN EJECUTIVO",s_sect))
        datos=[
            ['CONCEPTO','IMPORTE'],
            ['Total nómina dispersada',f"${tot_nom:,.2f}"],
            ['Total confirmado en PDFs bancarios',f"${tot_pdf:,.2f}"],
            ['Diferencia nómina vs PDFs',f"${diff:+,.2f}"],
            ['',''],
            ['Empleados procesados',f"{total_r:,}"],
            ['Confirmados OK',f"{stats['OK']:,}  ({pct_ok:.1f}%)"],
            ['No encontrados en PDF',f"{stats['NO EN PDF']:,}"],
            ['Importe difiere',f"{stats['DIFIERE']:,}"],
            ['Cuentas con doble abono',f"{len(dobles):,}"],
        ]
        ts=[
            ('BACKGROUND',(0,0),(-1,0),AZ2),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8.5),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GR2]),
            ('GRID',(0,0),(-1,-1),.5,BD),
            ('FONTNAME',(0,3),(-1,3),'Helvetica-Bold'),
            ('TEXTCOLOR',(1,3),(1,3),VD if abs(diff)<1000 else (AM if abs(diff)<50000 else RJ)),
            ('TEXTCOLOR',(1,6),(1,6),VD),
            ('TEXTCOLOR',(1,7),(1,7),RJ if stats['NO EN PDF']>0 else VD),
            ('SPAN',(0,4),(-1,4)),
            ('ALIGN',(1,0),(1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
            ('LEFTPADDING',(0,0),(-1,-1),10),
        ]
        story.append(Table(datos,[11*cm,5.5*cm],style=TableStyle(ts)))
        story.append(Spacer(1,.5*cm))

        # Por banco
        story.append(Paragraph("DESGLOSE POR BANCO",s_sect))
        grp=defaultdict(lambda:{'nom':0.,'pdf':0.,'ok':0,'no':0,'dif':0})
        for r in res:
            k=r['banco'] or 'SIN BANCO'
            grp[k]['nom']+=r['imp_nom']; grp[k]['pdf']+=r['imp_pdf'] or 0
            if r['estatus']=='OK': grp[k]['ok']+=1
            elif r['estatus']=='NO EN PDF': grp[k]['no']+=1
            else: grp[k]['dif']+=1
        bdata=[['BANCO','NÓMINA','PDF','DIFERENCIA','OK','NO PDF','DIF']]
        for banco,v in sorted(grp.items()):
            d=v['nom']-v['pdf']
            bdata.append([banco,f"${v['nom']:,.2f}",f"${v['pdf']:,.2f}",
                           f"${d:+,.2f}",str(v['ok']),str(v['no']),str(v['dif'])])
        bts=[('BACKGROUND',(0,0),(-1,0),AZ),('TEXTCOLOR',(0,0),(-1,0),colors.white),
             ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),
             ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GR2]),
             ('GRID',(0,0),(-1,-1),.5,BD),
             ('ALIGN',(1,0),(-1,-1),'RIGHT'),('ALIGN',(0,0),(0,-1),'LEFT'),
             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
             ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
             ('LEFTPADDING',(0,0),(-1,-1),8)]
        story.append(Table(bdata,[4*cm,3.3*cm,3.3*cm,3*cm,1.5*cm,1.5*cm,1.5*cm],
                           style=TableStyle(bts)))
        story.append(Spacer(1,.5*cm))

        # No en PDF
        no_pdf=[r for r in res if r['estatus']=='NO EN PDF']
        if no_pdf:
            story.append(Paragraph(f"EMPLEADOS NO ENCONTRADOS EN PDF  ({len(no_pdf):,})",s_sect))
            nd=[['NSS','NOMBRE','BANCO','IMPORTE']]+\
               [[r['nss'],r['nombre'][:42],r['banco'],f"${r['imp_nom']:,.2f}"]
                for r in sorted(no_pdf,key=lambda x:x['banco'])]
            nts=[('BACKGROUND',(0,0),(-1,0),CA('#991B1B')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                 ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),
                 ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,CA('#FFF8F8')]),
                 ('GRID',(0,0),(-1,-1),.5,BD),('ALIGN',(3,0),(3,-1),'RIGHT'),
                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                 ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
                 ('LEFTPADDING',(0,0),(-1,-1),8)]
            story.append(Table(nd,[3.5*cm,7*cm,3*cm,3*cm],style=TableStyle(nts)))
            story.append(Spacer(1,.2*cm))
            story.append(Paragraph(
                f"Total no encontrado: ${sum(r['imp_nom'] for r in no_pdf):,.2f}",
                ST('b2',fontSize=8,textColor=RJ)))

        story.append(Spacer(1,1*cm))
        story.append(HRFlowable(width="100%",thickness=1,color=BD))
        story.append(Spacer(1,.2*cm))
        story.append(Paragraph(
            f"Documento generado el {datetime.now().strftime('%d de %B de %Y a las %H:%M')}  ·  "
            f"Sistema de Control de Nómina  ·  HIVICO / FIREWALL",s_footer))
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        return None

# ══════════════════════════════════════════════════════════════
# ══ TAB 1 · SÁBANAS ══════════════════════════════════════════
# ══════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════
# ══ TAB 1 · SÁBANAS ══════════════════════════════════════════
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
# PARSERS DE LAYOUTS DE DISPERSIÓN BANCARIA
# ══════════════════════════════════════════════════════════════
def _limpiar_imp_lay(v):
    try: return float(str(v).replace(',','').replace('$','').replace(' ','').strip())
    except: return 0.0

def _detectar_banco_contenido(rows):
    """Detecta el banco examinando el contenido cuando el nombre no lo indica.
    Prioridad: Santander > Banorte > BBVA."""
    contenido = ' '.join(
        str(v or '').upper()
        for r in rows[:25] for v in (r if r else [])
    )
    # SANTANDER: apellidos separados en columnas distintas (PATERNO + MATERNO)
    if 'PATERNO' in contenido and 'MATERNO' in contenido:
        return 'SANTANDER'
    # Verificar Santander por MCRO o MACRO en cualquier celda
    if 'MCRO' in contenido or 'MACRO PAGO' in contenido:
        return 'SANTANDER'
    # BANORTE: tiene EMPLEADO en header (layout de dispersión Banorte)
    if 'EMPLEADO' in contenido:
        return 'BANORTE'
    # BBVA: tiene TDP (Tarjeta De Pago) o BENEFICIARIO
    if 'TDP' in contenido or 'BENEFICIARIO' in contenido:
        return 'BBVA'
    # BBVA fallback: CUENTA + IMPORTE en misma fila sin apellidos
    for r in rows[:20]:
        vals = [str(v or '').upper().strip() for v in (r or [])]
        tiene_cta = any('CUENTA' in v or 'TDP' in v for v in vals)
        tiene_imp = any('IMPORTE' in v or 'MONTO' in v for v in vals)
        if tiene_cta and tiene_imp:
            return 'BBVA'
    return 'DESCONOCIDO'

def _imp_banorte(imp_raw):
    """Convierte importe Banorte: soporta formato ceros (000000000063008→630.08)
    y formato decimal normal (1234.56).
    IMPORTANTE: openpyxl devuelve enteros como 30649480.0 (con .0),
    por eso NO usamos 'if . in s' — miramos si los centavos son cero."""
    s = str(imp_raw or '').replace(',','').replace('$','').strip()
    if not s or s == 'nan': return 0.0
    try:
        f = float(s)
    except:
        return 0.0
    iv = int(f)
    # Detectar si tiene centavos REALES (no sólo .0 de float)
    centavos_reales = abs(f - iv) > 0.001
    if centavos_reales:
        # Ya es decimal real (ej. 1234.56) → devolver directo
        return f
    # Es entero (o .0 de openpyxl) → intentar conversión de ceros
    if iv >= 100:
        s2 = str(iv).zfill(15)
        converted = float(s2[:-2] + '.' + s2[-2:])
        if 0 < converted < 500000:
            return converted
    return f

def leer_layout_bytes(raw_bytes, nombre_archivo):
    """Lee layout de dispersión bancaria. Detecta banco por nombre y contenido.
    Soporta .xls y .xlsx/.csv"""
    import io as _io
    nombre_up = nombre_archivo.upper()
    if 'BANORTE' in nombre_up or 'CCM' in nombre_up or 'PAGO_NOM' in nombre_up.replace(' ','_'):
        banco = 'BANORTE'
    elif 'BBVA' in nombre_up or 'BANCOMER' in nombre_up:
        banco = 'BBVA'
    elif 'SANTANDER' in nombre_up or 'MCRO' in nombre_up or 'MACRO' in nombre_up:
        banco = 'SANTANDER'
    else:
        banco = 'DESCONOCIDO'

    registros = []
    total_lay = 0.0
    err = None

    try:
        if nombre_archivo.lower().endswith('.csv'):
            import csv as _csv
            lines = raw_bytes.decode('utf-8-sig', errors='replace').splitlines()
            reader = _csv.reader(lines)
            rows = list(reader)
        else:
            # Intentar leer con openpyxl (xlsx) o xlrd (xls)
            try:
                import openpyxl as _ox
                wb = _ox.load_workbook(_io.BytesIO(raw_bytes), data_only=True, read_only=True)
                ws = wb.active
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                wb.close()
            except Exception:
                try:
                    import xlrd as _xl
                    wb = _xl.open_workbook(file_contents=raw_bytes)
                    ws = wb.sheet_by_index(0)
                    rows = [[ws.cell_value(r,c) for c in range(ws.ncols)] for r in range(ws.nrows)]
                except Exception as e2:
                    return [], 0.0, f"No se pudo leer: {e2}"

        # ── Si banco sigue DESCONOCIDO, detectar por contenido ──
        if banco == 'DESCONOCIDO':
            banco = _detectar_banco_contenido(rows)

        # ── Detectar header y parsear según banco ──────────────
        if banco == 'BANORTE':
            # Cols: Empleado | Ref.Leyenda | Importe(con ceros) | CodBco | TipoCta | Cuenta
            # Buscar fila header con "Empleado" o "Importe" — hasta fila 20
            hdr_row = 0
            for i, r in enumerate(rows[:20]):
                vals = [str(v or '').upper() for v in r]
                if any('EMPLEADO' in v or 'IMPORTE' in v for v in vals):
                    hdr_row = i; break
            hdr = [str(v or '').upper().strip() for v in rows[hdr_row]]
            ci_emp  = next((i for i,h in enumerate(hdr) if 'EMPLEADO' in h), 0)
            ci_imp  = next((i for i,h in enumerate(hdr) if 'IMPORTE' in h), 2)
            ci_cta  = next((i for i,h in enumerate(hdr) if 'CUENTA' in h), 5)
            ci_ref  = next((i for i,h in enumerate(hdr) if 'REF' in h or 'LEYENDA' in h), 1)
            for r in rows[hdr_row+1:]:
                if not r or not any(r): continue
                emp  = str(r[ci_emp] or '').strip() if ci_emp < len(r) else ''
                imp_raw = r[ci_imp] if ci_imp < len(r) else 0
                cta  = str(r[ci_cta] or '').strip() if ci_cta < len(r) else ''
                ref  = str(r[ci_ref] or '').strip() if ci_ref < len(r) else ''
                # Soporta formato ceros (000000000063008→630.08) y decimal normal
                imp = _imp_banorte(imp_raw)
                if not emp and not cta: continue
                if imp <= 0: continue
                registros.append({
                    'banco': banco, 'cuenta': cta, 'importe': imp,
                    'nombre': '', 'referencia': ref,
                    'empleado_id': emp, 'archivo': nombre_archivo,
                })
                total_lay += imp

        elif banco == 'BBVA':
            # Cols: CUENTA/TDP | IMPORTE | NOMBRE
            # Buscar header: exigir que en la misma fila estén keywords de CUENTA e IMPORTE
            # Y que la celda de importe sea texto (etiqueta), no un número
            KW_CTA = ['CUENTA','TDP','CLABE','TARJETA']
            KW_IMP = ['IMPORTE','NETO','MONTO','CANTIDAD']
            hdr_row = 0
            for i, r in enumerate(rows[:25]):
                vals = [str(v or '').upper().strip() for v in r]
                tiene_cta = any(k in v for v in vals for k in KW_CTA)
                tiene_imp = any(k in v for v in vals for k in KW_IMP)
                if tiene_cta and tiene_imp:
                    # Verificar que la celda de importe sea etiqueta, no número
                    imp_idx = next((j for j,v in enumerate(vals)
                                    if any(k in v for k in KW_IMP)), -1)
                    try:
                        float(str(r[imp_idx] or '').replace(',','').replace('$',''))
                        continue  # es número → es fila de totales, no header
                    except (ValueError, TypeError):
                        hdr_row = i; break  # es texto → es el header real

            hdr = [str(v or '').upper().strip() for v in rows[hdr_row]]
            ci_cta = next((i for i,h in enumerate(hdr) if any(k in h for k in KW_CTA)), None)
            ci_imp = next((i for i,h in enumerate(hdr) if any(k in h for k in KW_IMP)), None)
            ci_nom = next((i for i,h in enumerate(hdr)
                           if any(k in h for k in ['NOMBRE','BENEFICIARIO'])), None)

            # Fallback por tipo de valor si header no alcanzó
            if ci_cta is None or ci_imp is None:
                for r in rows[hdr_row+1:hdr_row+6]:
                    if not r or not any(r): continue
                    for j, v in enumerate(r):
                        sv = str(v or '').strip().replace(',','').replace('$','')
                        sv = re.sub(r'\.0$', '', sv)
                        if ci_cta is None and re.match(r'^\d{10,18}$', sv):
                            ci_cta = j
                        elif ci_imp is None and re.match(r'^\d{1,8}\.\d{2}$', sv):
                            try:
                                if float(sv) > 0: ci_imp = j
                            except: pass
                    if ci_cta is not None and ci_imp is not None: break

            ci_cta = ci_cta if ci_cta is not None else 0
            ci_imp = ci_imp if ci_imp is not None else 1
            ci_nom = ci_nom if ci_nom is not None else 2
            if ci_imp == ci_cta: ci_imp = ci_cta + 1

            for r in rows[hdr_row+1:]:
                if not r or not any(r): continue
                cta = str(r[ci_cta] or '').strip().replace('.0','') if ci_cta < len(r) else ''
                imp = _limpiar_imp_lay(r[ci_imp]) if ci_imp < len(r) else 0
                nom = str(r[ci_nom] or '').strip() if ci_nom < len(r) else ''
                if not cta or imp <= 0: continue
                registros.append({
                    'banco': banco, 'cuenta': cta, 'importe': imp,
                    'nombre': nom, 'referencia': '',
                    'empleado_id': '', 'archivo': nombre_archivo,
                })
                total_lay += imp

        elif banco == 'SANTANDER':
            # Cols: Num.Empleado | Ap.Paterno | Ap.Materno | Nombre | CUENTA | IMPORTE
            # El layout tiene filas de cabecera del documento (empresa, cuenta cargo, etc.)
            # antes del header real de empleados — buscar fila con AMBAS keywords como texto
            KW_CTA2 = ['CUENTA','CLABE','TARJETA']
            KW_IMP2 = ['IMPORTE','MONTO','NETO','CANTIDAD']
            KW_NOM2 = ['APELLIDO','PATERNO','NOMBRE']
            hdr_row = 0
            for i, r in enumerate(rows[:30]):
                vals = [str(v or '').upper().strip() for v in r]
                tiene_cta2 = any(k in v for v in vals for k in KW_CTA2)
                tiene_imp2 = any(k in v for v in vals for k in KW_IMP2)
                tiene_nom2 = any(k in v for v in vals for k in KW_NOM2)
                if (tiene_cta2 and tiene_imp2) or (tiene_nom2 and tiene_cta2):
                    # Verificar que la celda de importe/cuenta sea texto, no número
                    imp_idx2 = next((j for j,v in enumerate(vals)
                                     if any(k in v for k in KW_IMP2)), -1)
                    cta_idx2 = next((j for j,v in enumerate(vals)
                                     if any(k in v for k in KW_CTA2)), -1)
                    # Si la celda de cuenta contiene un número → es fila de datos, no header
                    cta_val = str(r[cta_idx2] or '').strip() if cta_idx2 >= 0 and cta_idx2 < len(r) else ''
                    try:
                        float(cta_val.replace(',',''))
                        continue  # es número → no es header
                    except (ValueError, TypeError):
                        hdr_row = i; break

            hdr = [str(v or '').upper().strip() for v in rows[hdr_row]]
            ci_num = next((i for i,h in enumerate(hdr)
                           if any(k in h for k in ['NUMERO','NUM','ID EMPLEADO','NO.EMPLEADO','NO. EMPLEADO'])), None)
            ci_pat = next((i for i,h in enumerate(hdr) if 'PATERNO' in h), None)
            ci_mat = next((i for i,h in enumerate(hdr) if 'MATERNO' in h), None)
            ci_nom = next((i for i,h in enumerate(hdr)
                           if 'NOMBRE' in h and 'APELLIDO' not in h and 'SERVICIO' not in h), None)
            ci_cta = next((i for i,h in enumerate(hdr)
                           if any(k in h for k in ['CUENTA','CLABE','TARJETA'])), None)
            ci_imp = next((i for i,h in enumerate(hdr)
                           if any(k in h for k in ['IMPORTE','MONTO','NETO','CANTIDAD'])), None)

            # Fallback por tipo de valor
            if ci_cta is None or ci_imp is None:
                for r in rows[hdr_row+1:hdr_row+6]:
                    if not r or not any(r): continue
                    for j, v in enumerate(r):
                        sv = str(v or '').strip().replace(',','').replace('$','')
                        sv = re.sub(r'\.0$', '', sv)
                        if ci_cta is None and re.match(r'^\d{10,18}$', sv):
                            ci_cta = j
                        elif ci_imp is None and re.match(r'^\d{1,8}\.\d{2}$', sv):
                            try:
                                if float(sv) > 0: ci_imp = j
                            except: pass
                    if ci_cta is not None and ci_imp is not None: break

            ci_num = ci_num if ci_num is not None else 0
            ci_pat = ci_pat if ci_pat is not None else 1
            ci_mat = ci_mat if ci_mat is not None else 2
            ci_nom = ci_nom if ci_nom is not None else 3
            ci_cta = ci_cta if ci_cta is not None else 4
            ci_imp = ci_imp if ci_imp is not None else 5
            if ci_imp == ci_cta: ci_imp = ci_cta + 1

            for r in rows[hdr_row+1:]:
                if not r or not any(r): continue
                pat = str(r[ci_pat] or '').strip() if ci_pat < len(r) else ''
                mat = str(r[ci_mat] or '').strip() if ci_mat < len(r) else ''
                nom = str(r[ci_nom] or '').strip() if ci_nom < len(r) else ''
                nombre = f"{pat} {mat} {nom}".strip()
                cta = str(r[ci_cta] or '').strip().replace('.0','') if ci_cta < len(r) else ''
                imp = _limpiar_imp_lay(r[ci_imp]) if ci_imp < len(r) else 0
                emp_id = str(r[ci_num] or '').strip() if ci_num < len(r) else ''
                if not cta or imp <= 0: continue
                # Filtrar cuenta de cargo de la empresa (suelen ser 11 dígitos o muy distintos)
                if not re.match(r'^\d{8,18}$', re.sub(r'\.0$', '', cta)): continue
                registros.append({
                    'banco': banco, 'cuenta': cta, 'importe': imp,
                    'nombre': nombre, 'referencia': '01PAGO DE NOMINA',
                    'empleado_id': emp_id, 'archivo': nombre_archivo,
                })
                total_lay += imp

        else:
            # Fallback genérico
            import re as _re
            for r in rows[1:]:
                if not r or not any(r): continue
                row_str = ' '.join(str(v or '') for v in r)
                nums = _re.findall(r'\b(\d{10,18})\b', row_str)
                imps = _re.findall(r'(\d{1,8}\.\d{2})\b', row_str)
                if nums and imps:
                    imp = _limpiar_imp_lay(imps[0])
                    if imp > 0:
                        registros.append({'banco': banco, 'cuenta': nums[0],
                                          'importe': imp, 'nombre': '',
                                          'referencia': '', 'empleado_id': '',
                                          'archivo': nombre_archivo})
                        total_lay += imp

    except Exception as e:
        err = str(e)

    return registros, total_lay, err


# ══════════════════════════════════════════════════════════════
# TAB 2 — LAYOUTS DE DISPERSIÓN
# ══════════════════════════════════════════════════════════════
with TAB1:
    st.markdown("""
<div class="sec-hdr">
  <span class="sec-badge">01</span>
  <span class="sec-title">Sábanas de Nómina</span>
  <span class="sec-desc">Lectura, procesamiento y carga · Q1 y Q2</span>
</div>""", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="up-lbl">Primera quincena — Q1</div>', unsafe_allow_html=True)
        ups_q1 = st.file_uploader("Q1  (xlsx individuales o un ZIP)", type=None,
                                   accept_multiple_files=True,
                                   key="up_q1", label_visibility="collapsed")
        if ups_q1:
            st.markdown(f'<div class="up-ok">📎 {len(ups_q1)} archivo(s) cargado(s)</div>',
                        unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="up-lbl">Segunda quincena — Q2</div>', unsafe_allow_html=True)
        ups_q2 = st.file_uploader("Q2  (xlsx individuales o un ZIP)", type=None,
                                   accept_multiple_files=True,
                                   key="up_q2", label_visibility="collapsed")
        if ups_q2:
            st.markdown(f'<div class="up-ok">📎 {len(ups_q2)} archivo(s) cargado(s)</div>',
                        unsafe_allow_html=True)

    st.markdown('<div style="height:.5rem"></div>', unsafe_allow_html=True)

    if ups_q1 or ups_q2:
        if st.button("▶  Procesar sábanas", type="primary", use_container_width=True):
            dfs, log_det = [], []
            bar = st.progress(0)
            i = 0
            archivos_q1 = extraer_archivos_excel(ups_q1)
            archivos_q2 = extraer_archivos_excel(ups_q2)
            total_arch  = len(archivos_q1) + len(archivos_q2)
            for nombre, datos in archivos_q1:
                df, err = leer_sabana_bytes(datos, nombre, 1)
                if err:
                    log_det.append(("ERR", nombre, "Q1", 0, 0.0, err))
                    st.error(f"Q1 · {nombre}: {err}")
                elif df is not None and len(df) > 0:
                    dfs.append(df)
                    sc = int(df["clave_interbancaria"].isna().sum())
                    log_det.append(("OK", nombre, "Q1", len(df), df['total_quincenal'].sum(),
                                    f"Hoja: {df['hoja_detectada'].iloc[0]}  ·  Sin CLABE: {sc}"))
                i += 1; bar.progress(i / max(total_arch, 1))
            for nombre, datos in archivos_q2:
                df, err = leer_sabana_bytes(datos, nombre, 2)
                if err:
                    log_det.append(("ERR", nombre, "Q2", 0, 0.0, err))
                    st.error(f"Q2 · {nombre}: {err}")
                elif df is not None and len(df) > 0:
                    dfs.append(df)
                    sc = int(df["clave_interbancaria"].isna().sum())
                    log_det.append(("OK", nombre, "Q2", len(df), df['total_quincenal'].sum(),
                                    f"Hoja: {df['hoja_detectada'].iloc[0]}  ·  Sin CLABE: {sc}"))
                i += 1; bar.progress(i / max(total_arch, 1))
            bar.empty()
            if dfs:
                st.session_state["df_sabanas"] = pd.concat(dfs, ignore_index=True)
                st.session_state["log_sabanas"] = log_det

    if "df_sabanas" in st.session_state:
        df_all = st.session_state["df_sabanas"]
        log    = st.session_state.get("log_sabanas", [])

        total_nom  = df_all['total_quincenal'].sum()
        con_clabe  = int(df_all['clave_interbancaria'].notna().sum())
        sin_clabe  = int(df_all['clave_interbancaria'].isna().sum())
        n_arch     = df_all['archivo_origen'].nunique()
        pct_c      = con_clabe / len(df_all) * 100 if len(df_all) else 0

        # KPIs
        st.markdown(f"""
<div class="kpi-row">
  <div class="kpi k-blue">
    <div class="kpi-v">{len(df_all):,}</div>
    <div class="kpi-l">Total empleados</div>
    <div class="kpi-s">{n_arch} archivo(s) procesado(s)</div>
  </div>
  <div class="kpi k-green">
    <div class="kpi-v">{con_clabe:,}</div>
    <div class="kpi-l">Con CLABE</div>
    <div class="kpi-s">{pct_c:.1f}% del total</div>
  </div>
  <div class="kpi k-red">
    <div class="kpi-v">{sin_clabe:,}</div>
    <div class="kpi-l">Sin CLABE</div>
    <div class="kpi-s">Requieren revisión</div>
  </div>
  <div class="kpi k-blue">
    <div class="kpi-v">${total_nom:,.0f}</div>
    <div class="kpi-l">Total nómina</div>
    <div class="kpi-s">Q1 + Q2</div>
  </div>
</div>""", unsafe_allow_html=True)

        try:
            _xl_top = exportar_nomina_pdf_excel(cb["res"],cb["stats"],cb["dobles"],cb["tot_pdf"],
                "HIVICO / FIREWALL SEGURIDAD PRIVADA SA DE CV","B61-79454-10-7",
                periodo,cb["fecha"],tolerancia)
            st.download_button("⬇  Descargar reporte Excel",_xl_top,
                f"CRUCE_NOMINA_PDF_{cb['fecha'].replace('-','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",use_container_width=True,key="dl_excel_top")
        except Exception as _e_xl: st.warning(f"No se pudo generar Excel: {_e_xl}")

        # Resumen ejecutivo
        q1 = df_all[df_all['quincena']==1]
        q2 = df_all[df_all['quincena']==2]
        bd_dist = df_all['banco'].fillna('SIN BANCO').value_counts().to_dict()
        bd_str  = "  ·  ".join(f"{b}: {n:,}" for b,n in sorted(bd_dist.items()))
        pc_cls  = "pos" if pct_c>=95 else ("neu" if pct_c>=80 else "neg")

        st.markdown(f"""
<div class="exec">
  <div class="exec-t">Resumen ejecutivo &nbsp;·&nbsp; {periodo}</div>
  <div class="e-row hl">
    <span class="e-lbl"><b>Total nómina del mes</b></span>
    <span class="e-val">${total_nom:,.2f}</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">Primera quincena Q1 &nbsp;·&nbsp; {len(q1):,} empleados</span>
    <span class="e-val">${q1['total_quincenal'].sum():,.2f}</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">Segunda quincena Q2 &nbsp;·&nbsp; {len(q2):,} empleados</span>
    <span class="e-val">${q2['total_quincenal'].sum():,.2f}</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">Distribución por banco</span>
    <span class="e-val" style="font-size:.72rem">{bd_str}</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">Cobertura CLABE</span>
    <span class="e-val {pc_cls}">{pct_c:.1f}% &nbsp;({con_clabe:,} / {len(df_all):,})</span>
  </div>
</div>""", unsafe_allow_html=True)

        # Log archivos
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.4rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">LOG</span>
  <span class="sec-title">Detalle por archivo</span>
</div>""", unsafe_allow_html=True)

        rows_log = ""
        gran = 0.0
        for est, arch, qna, emps, tot_a, det in log:
            gran += tot_a
            ok  = '<span style="color:#16A34A;font-weight:700">✔</span>' if est=="OK" else                   '<span style="color:#DC2626;font-weight:700">✘</span>'
            bg  = "" if est=="OK" else ' style="background:#FEF9F9"'
            qb  = (f'<span style="background:{"#EEF4FF" if qna=="Q1" else "#FFFBEB"};'
                   f'color:{"#1A3A6B" if qna=="Q1" else "#92400E"};padding:2px 8px;'
                   f'border-radius:4px;font-size:.62rem;font-weight:700">{qna}</span>')
            rows_log += (f'<tr{bg}><td class="c">{ok}</td>'
                        f'<td class="mono">{arch}</td>'
                        f'<td class="c">{qb}</td>'
                        f'<td class="r">{emps:,}</td>'
                        f'<td class="r">${tot_a:,.2f}</td>'
                        f'<td style="font-size:.71rem;color:#5A7A9A">{det}</td></tr>')

        rows_log += (f'<tr style="background:#EEF4FF"><td colspan="3" '
                    f'style="color:#1A3A6B;font-weight:700;padding:9px 13px">'
                    f'TOTAL GENERAL — {n_arch} archivo(s)</td>'
                    f'<td class="r" style="color:#1A3A6B;font-weight:700">{len(df_all):,}</td>'
                    f'<td class="r" style="color:#1A3A6B;font-weight:700">${gran:,.2f}</td>'
                    f'<td></td></tr>')

        st.markdown(f"""
<div class="t-wrap"><table class="t">
<thead><tr>
  <th class="c"></th><th>Archivo</th><th class="c">QNA</th>
  <th class="r">Empleados</th><th class="r">Total quincenal</th>
  <th>Observaciones</th>
</tr></thead>
<tbody>{rows_log}</tbody>
</table></div>""", unsafe_allow_html=True)

        # Exportar
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.4rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">↓</span>
  <span class="sec-title">Exportar</span>
</div>""", unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            xl = exportar_sabanas_excel(df_all, periodo)
            st.download_button("⬇  Reporte Excel de sábanas", xl,
                                f"SABANAS_{mes_sel[:3].upper()}{anio_sel}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary", use_container_width=True)
        with c2:
            buf_csv = io.StringIO()
            df_all.to_csv(buf_csv, index=False, encoding="utf-8-sig")
            st.download_button("⬇  Datos en CSV", buf_csv.getvalue().encode("utf-8-sig"),
                                f"SABANAS_{mes_sel[:3].upper()}{anio_sel}.csv",
                                mime="text/csv", use_container_width=True)

        if PSYCOPG2_OK and usar_bd:
            st.markdown("""
<div class="sec-hdr" style="margin-top:1.4rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">BD</span>
  <span class="sec-title">Carga a PostgreSQL</span>
</div>""", unsafe_allow_html=True)
            st.markdown('<div class="modo modo-s">🖥  Modo servidor — carga disponible</div>',
                        unsafe_allow_html=True)
            b1, b2, b3 = st.columns(3)
            with b1: nombre_per = st.text_input("Nombre periodo", value=f"NÓMINA {periodo}")
            with b2: fecha_bd   = st.date_input("Fecha de pago", value=datetime.today())
            with b3: tipo_pago  = st.selectbox("Tipo", ["QUINCENAL","MENSUAL","SEMANAL"])
            if st.button("📤  Cargar a PostgreSQL", type="primary", use_container_width=True):
                try:
                    con = bd_conectar(); cur = con.cursor()
                    id_per = bd_insertar_periodo(cur, nombre_per, str(fecha_bd),
                                                  tipo_pago, MESES.index(mes_sel)+1, anio_sel)
                    con.commit()
                    ok_cnt = skip = err_cnt = 0
                    bar = st.progress(0)
                    for idx_bd, (_, row) in enumerate(df_all.iterrows()):
                        try:
                            cur.execute("SAVEPOINT sp1")
                            id_emp = bd_insertar_empleado(cur, row)
                            if not id_emp:
                                cur.execute("RELEASE SAVEPOINT sp1"); skip += 1; continue
                            bd_insertar_nomina(cur, id_emp, id_per, row)
                            cur.execute("RELEASE SAVEPOINT sp1"); ok_cnt += 1
                        except:
                            err_cnt += 1
                            cur.execute("ROLLBACK TO SAVEPOINT sp1")
                            cur.execute("RELEASE SAVEPOINT sp1")
                        bar.progress((idx_bd+1)/len(df_all))
                    bar.empty(); con.commit(); cur.close(); con.close()
                    st.success(f"✅  Cargados: {ok_cnt:,}  ·  Sin ID: {skip}  ·  Errores: {err_cnt}")
                except Exception as e:
                    st.error(f"Error: {e}")


# ══════════════════════════════════════════════════════════════
# ══ TAB 2 · LAYOUTS DE DISPERSIÓN ════════════════════════════
# ══════════════════════════════════════════════════════════════
with TAB2:
    st.markdown("""
<div class="sec-hdr">
  <span class="sec-badge">02</span>
  <span class="sec-title">Layouts de Dispersión Bancaria</span>
  <span class="sec-desc">BBVA · Banorte · Santander</span>
</div>""", unsafe_allow_html=True)

    # ── Uploaders por banco ─────────────────────────────────────
    _c1, _c2, _c3 = st.columns(3)
    with _c1:
        st.markdown('''<div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.4rem">
          <div style="width:14px;height:14px;border-radius:50%;background:#003087"></div>
          <span class="up-lbl" style="margin:0">BBVA</span></div>''', unsafe_allow_html=True)
        ups_bbva = st.file_uploader("BBVA", type=["xls","xlsx","csv"],
                                     accept_multiple_files=True, key="up_lay_bbva",
                                     label_visibility="collapsed")
        if ups_bbva:
            st.markdown(f'<div class="up-ok">📎 {len(ups_bbva)} archivo(s)</div>', unsafe_allow_html=True)
    with _c2:
        st.markdown('''<div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.4rem">
          <div style="width:14px;height:14px;border-radius:50%;background:#E40028"></div>
          <span class="up-lbl" style="margin:0">BANORTE</span></div>''', unsafe_allow_html=True)
        ups_bnr = st.file_uploader("BANORTE", type=["xls","xlsx","csv"],
                                    accept_multiple_files=True, key="up_lay_bnr",
                                    label_visibility="collapsed")
        if ups_bnr:
            st.markdown(f'<div class="up-ok">📎 {len(ups_bnr)} archivo(s)</div>', unsafe_allow_html=True)
    with _c3:
        st.markdown('''<div style="display:flex;align-items:center;gap:.5rem;margin-bottom:.4rem">
          <div style="width:14px;height:14px;border-radius:50%;background:#EC0000"></div>
          <span class="up-lbl" style="margin:0">SANTANDER</span></div>''', unsafe_allow_html=True)
        ups_san = st.file_uploader("SANTANDER", type=["xls","xlsx","csv"],
                                    accept_multiple_files=True, key="up_lay_san",
                                    label_visibility="collapsed")
        if ups_san:
            st.markdown(f'<div class="up-ok">📎 {len(ups_san)} archivo(s)</div>', unsafe_allow_html=True)

    # Juntar todos bajo banco forzado por sección
    def _forzar_banco(files, banco_forzado):
        """Lee archivos y sobreescribe el banco detectado con el banco de la sección."""
        regs, log = [], []
        for f in (files or []):
            raw = f.read()
            r, t, e = leer_layout_bytes(raw, f.name)
            # Forzar banco al de la sección — el usuario los puso en esa columna a propósito
            for reg in r:
                reg['banco'] = banco_forzado
            regs.extend(r)
            log.append((f.name, len(r), t, e))
        return regs, log

    ups_lay = (ups_bbva or []) + (ups_bnr or []) + (ups_san or [])

    if ups_lay and st.button("▶  Analizar Layouts", type="primary",
                              use_container_width=True, key="btn_lay"):
        _regs_tmp, _log_tmp = [], []
        _bar = st.progress(0)
        _total_f = len(ups_lay)
        _idx_f = 0
        for _banco_f, _files_f in [("BBVA", ups_bbva), ("BANORTE", ups_bnr), ("SANTANDER", ups_san)]:
            for _f in (_files_f or []):
                _raw_f = _f.read()
                _r, _t, _e = leer_layout_bytes(_raw_f, _f.name)
                for _reg in _r: _reg["banco"] = _banco_f
                _lt_f = sum(x["importe"] for x in _r)
                _regs_tmp.extend(_r)
                _log_tmp.append((_f.name, len(_r), _lt_f, _e))
                _idx_f += 1
                _bar.progress(_idx_f / max(_total_f, 1))
        _bar.empty()
        st.session_state["layouts"] = {"regs": _regs_tmp, "log": _log_tmp}

    if "layouts" in st.session_state:
        import io as _iol
        from openpyxl import Workbook as _WBL
        from openpyxl.styles import Font as _FL, PatternFill as _PFL, Alignment as _ALL, Border as _BRL, Side as _SDL
        from openpyxl.utils import get_column_letter as _GCLL
        from collections import defaultdict as _ddL

        _ld = st.session_state["layouts"]
        _lr = _ld["regs"]; _ll = _ld["log"]
        _lt = sum(r["importe"] for r in _lr)
        _lb = {}
        for r in _lr: _lb[r["banco"]] = _lb.get(r["banco"],0) + r["importe"]
        BCOL_L = {"BBVA":"D6EAF8","BANORTE":"E2EFDA","SANTANDER":"FCE4D6","DESCONOCIDO":"F2F2F2"}

        st.markdown(f"""
<div class="kpi-row">
  <div class="kpi k-blue"><div class="kpi-v">{len(_lr):,}</div>
    <div class="kpi-l">Total operaciones</div><div class="kpi-s">{len(_ll)} archivos</div></div>
  <div class="kpi k-green"><div class="kpi-v">${_lt:,.2f}</div>
    <div class="kpi-l">Total a dispersar</div>
    <div class="kpi-s" style="font-size:.6rem">{"  ·  ".join(f"{b}: ${v:,.2f}" for b,v in sorted(_lb.items()))}</div>
  </div>
</div>""", unsafe_allow_html=True)

        if _lr:
            def _make_xl_lay(regs, log):
                wb=_WBL(); _s=_SDL(style="thin",color="CCCCCC")
                brd=lambda:_BRL(left=_s,right=_s,top=_s,bottom=_s)
                fl=lambda c:_PFL("solid",fgColor=c)
                fn=lambda bold=False,color="1A1A1A",size=9:_FL(bold=bold,color=color,size=size,name="Calibri")
                aln=lambda h="left",v="center":_ALL(horizontal=h,vertical=v)
                ws1=wb.active; ws1.title="RESUMEN LAYOUTS"; ws1.sheet_view.showGridLines=False
                ws1.merge_cells("A1:E1")
                c=ws1.cell(1,1,f"LAYOUTS DE DISPERSIÓN — {len(log)} archivos · {len(regs):,} operaciones")
                c.font=_FL(bold=True,color="FFFFFF",size=13,name="Calibri")
                c.fill=fl("1A5276"); c.alignment=_ALL(horizontal="center",vertical="center")
                ws1.row_dimensions[1].height=30
                pb=_ddL(lambda:{"ops":0,"total":0.0})
                for r in regs: pb[r["banco"]]["ops"]+=1; pb[r["banco"]]["total"]+=r["importe"]
                ws1.append([]); ws1.append(["BANCO","OPERACIONES","TOTAL","",""])
                fr=ws1.max_row
                for c in ws1[fr]: c.font=fn(bold=True,color="FFFFFF"); c.fill=fl("1A5276"); c.border=brd()
                for b,v in sorted(pb.items()):
                    ws1.append([b,v["ops"],v["total"],"",""]); fr=ws1.max_row
                    for c in ws1[fr]: c.border=brd(); c.font=fn(size=9); c.fill=fl(BCOL_L.get(b,"F2F2F2"))
                    ws1.cell(fr,3).number_format="$#,##0.00"; ws1.cell(fr,3).alignment=aln("right")
                ws1.append(["TOTAL",len(regs),sum(r["importe"] for r in regs),"",""]); fr=ws1.max_row
                for c in ws1[fr]: c.font=fn(bold=True,size=10); c.fill=fl("EEF4FF"); c.border=brd()
                ws1.cell(fr,3).number_format="$#,##0.00"; ws1.cell(fr,3).alignment=aln("right")
                ws1.append([]); ws1.append(["ARCHIVO","BANCO","OPS","TOTAL","ERROR"]); fr=ws1.max_row
                for c in ws1[fr]: c.font=fn(bold=True,color="FFFFFF"); c.fill=fl("1A5276"); c.border=brd()
                for arch,n,t,e in log:
                    ba=next((r["banco"] for r in regs if r["archivo"]==arch),"")
                    ws1.append([arch,ba,n,t,e or "OK"]); fr=ws1.max_row
                    for c in ws1[fr]: c.border=brd(); c.font=fn(size=9); c.fill=fl(BCOL_L.get(ba,"F2F2F2"))
                    ws1.cell(fr,4).number_format="$#,##0.00"; ws1.cell(fr,4).alignment=aln("right")
                for i,w in enumerate([42,12,10,16,30],1): ws1.column_dimensions[_GCLL(i)].width=w
                ws2=wb.create_sheet("DETALLE LAYOUTS"); ws2.sheet_view.showGridLines=False
                ws2.merge_cells("A1:G1")
                c=ws2.cell(1,1,f"DETALLE — {len(regs):,} ops  |  ${sum(r['importe'] for r in regs):,.2f}")
                c.font=_FL(bold=True,color="FFFFFF",size=12,name="Calibri"); c.fill=fl("1A5276")
                c.alignment=_ALL(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=28
                H=["BANCO","CUENTA","IMPORTE $","NOMBRE","ID EMPLEADO","REFERENCIA","ARCHIVO"]
                W=[12,18,14,32,14,20,42]
                for i,(h,w) in enumerate(zip(H,W),1):
                    c=ws2.cell(2,i,h); c.font=fn(bold=True,color="FFFFFF",size=9)
                    c.fill=fl("1A5276"); c.border=brd(); c.alignment=_ALL(horizontal="center",vertical="center")
                    ws2.column_dimensions[_GCLL(i)].width=w
                ws2.row_dimensions[2].height=26; ws2.freeze_panes="A3"
                for i,r in enumerate(sorted(regs,key=lambda x:(x["banco"],x["archivo"]))):
                    fi=i+3; cf=BCOL_L.get(r["banco"],"F2F2F2")
                    for j,val in enumerate([r["banco"],r["cuenta"],r["importe"],r["nombre"],r["empleado_id"],r["referencia"],r["archivo"]],1):
                        c=ws2.cell(fi,j); c.border=brd()
                        c.fill=fl("FAFBFD" if i%2==1 else cf); c.font=fn(size=9)
                        if j==3: c.value=float(val or 0); c.number_format="$#,##0.00"; c.alignment=aln("right")
                        else: c.value=str(val) if val else ""; c.alignment=aln("left","center")
                    ws2.row_dimensions[fi].height=20
                ft=len(regs)+3; ws2.cell(ft,1,"TOTAL").font=fn(bold=True,size=10)
                ws2.cell(ft,3,sum(r["importe"] for r in regs)); ws2.cell(ft,3).number_format="$#,##0.00"
                ws2.cell(ft,3).font=fn(bold=True,size=11); ws2.cell(ft,3).alignment=aln("right")
                for j in range(1,8): ws2.cell(ft,j).fill=fl("EEF4FF"); ws2.cell(ft,j).border=brd()
                buf=_iol.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

            _xl_lay=_make_xl_lay(_lr,_ll)
            st.download_button("⬇  Descargar análisis Excel layouts",_xl_lay,
                "LAYOUTS_DISPERSION.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",use_container_width=True,key="dl_lay")

        _html_log=""
        for arch,n,t,e in _ll:
            ok='<span style="color:#16A34A;font-weight:700">✔</span>' if not e else '<span style="color:#DC2626;font-weight:700">✘</span>'
            ba=next((r["banco"] for r in _lr if r["archivo"]==arch),"")
            _html_log+=(f'<tr><td class="c">{ok}</td><td class="mono">{arch}</td>' +
                f'<td class="c">{ba}</td><td class="r">{n:,}</td><td class="r">${t:,.2f}</td>' +
                f'<td style="font-size:.7rem;color:{"#DC2626" if e else "#16A34A"}">{e or "OK"}</td></tr>')
        st.markdown(f"""
<div class="t-wrap"><table class="t">
<thead><tr><th class="c"></th><th>Archivo</th><th class="c">Banco</th>
<th class="r">Operaciones</th><th class="r">Total</th><th>Estado</th>
</tr></thead><tbody>{_html_log}</tbody></table></div>""", unsafe_allow_html=True)

        if "df_sabanas" in st.session_state:
            _ds=st.session_state["df_sabanas"]
            _ts=float(_ds["total_quincenal"].sum()) if "total_quincenal" in _ds.columns else 0
            _diff=_ts-_lt; _dc2="pos" if abs(_diff)<1000 else ("neu" if abs(_diff)<50000 else "neg")
            st.markdown(f"""
<div class="exec" style="margin-top:1.2rem">
  <div class="exec-t">Sábanas vs Layouts — ¿Se dispersó lo calculado?</div>
  <div class="e-row hl"><span class="e-lbl"><b>Total sábanas</b></span>
    <span class="e-val">${_ts:,.2f}</span></div>
  <div class="e-row"><span class="e-lbl">Total layouts</span>
    <span class="e-val">${_lt:,.2f}</span></div>
  <div class="e-row"><span class="e-lbl"><b>Diferencia</b></span>
    <span class="e-val {_dc2}">${_diff:+,.2f}</span></div>
</div>""", unsafe_allow_html=True)
            if abs(_diff)<1.0: st.success("✅ Layouts coinciden con sábanas")
            elif abs(_diff)<50000: st.warning(f"⚠ Diferencia ${abs(_diff):,.2f}")
            else: st.error(f"❌ Diferencia significativa ${abs(_diff):,.2f}")


# ══════════════════════════════════════════════════════════════
# ══ TAB 3 · CRUCE NÓMINA vs BANCO ═══════════════════════════
# ══════════════════════════════════════════════════════════════
with TAB3:
    st.markdown("""
<div class="sec-hdr">
  <span class="sec-badge">02</span>
  <span class="sec-title">Nómina vs Dispersión Bancaria</span>
  <span class="sec-desc">BBVA · Banorte · Santander</span>
</div>""", unsafe_allow_html=True)

    with st.expander("⚙  Configuración de periodo", expanded=True):
        cp, ct = st.columns(2)
        with cp:
            st.markdown('<div class="up-lbl">Fecha de pago</div>', unsafe_allow_html=True)
            fecha_pago_str = st.text_input("Fecha", value="2026-01-15",
                                            label_visibility="collapsed")
        with ct:
            st.markdown('<div class="up-lbl">Tolerancia $</div>', unsafe_allow_html=True)
            tolerancia = st.number_input("Tolerancia", value=1.0, min_value=0.0,
                                          step=0.5, label_visibility="collapsed", key="tol_cruce")
    RANGO_1A = {'desde':(2026,1,1),'hasta':(2026,1,31)}
    RANGO_2A = {'desde':(2026,2,1),'hasta':(2026,2,28)}

    if PSYCOPG2_OK and usar_bd:
        st.markdown('<div class="modo modo-s">🖥  Modo servidor — nómina desde PostgreSQL · PDFs desde archivo</div>',
                    unsafe_allow_html=True)

        st.markdown('<div class="up-lbl">PDFs de dispersión bancaria</div>', unsafe_allow_html=True)
        ups_banco = st.file_uploader("PDFs bancarios", type=["pdf"],
                                      accept_multiple_files=True,
                                      key="up_banco_pdf", label_visibility="collapsed")
        if ups_banco:
            st.markdown(f'<div class="up-ok">📎 {len(ups_banco)} PDF(s) cargado(s)</div>',
                        unsafe_allow_html=True)

        if st.button("▶  Cargar nómina y preparar PDFs",
                     type="primary", use_container_width=True, key="btn_bd"):
            if not ups_banco:
                st.warning("⬆ Primero sube los PDFs bancarios")
            else:
                # ── Paso 1: obtener nómina ────────────────────────
                nom_rows = []
                fuente_nom = ""

                # Intentar desde BD
                try:
                    con = bd_conectar(); cur = con.cursor()
                    cur.execute("""
                        SELECT e.nss,e.nombre_completo,e.numero_cuenta,e.clave_interbancaria,
                               b.nombre_banco,n.total_quincenal,p.nombre_periodo,n.archivo_origen
                        FROM nomina n
                        JOIN empleados e ON e.id_empleado=n.id_empleado
                        JOIN periodos  p ON p.id_periodo=n.id_periodo
                        LEFT JOIN bancos b ON b.id_banco=n.id_banco
                        WHERE n.total_quincenal>0 AND p.fecha_pago=%s
                        ORDER BY b.nombre_banco,e.nombre_completo
                    """,(fecha_pago_str,))
                    nom_rows = cur.fetchall()
                    cur.close(); con.close()
                    fuente_nom = f"PostgreSQL ({len(nom_rows):,} registros)"
                except Exception as e_bd:
                    st.warning(f"⚠ No se pudo conectar a BD: {e_bd}")

                # Si BD falló o está vacía, usar sábanas del Módulo 1
                if not nom_rows:
                    if "df_sabanas" in st.session_state:
                        df_n = st.session_state["df_sabanas"].copy()
                        # Agrupar empleados semanales: mismo NSS+cuenta puede aparecer
                        # múltiples veces (s1, s2, s3) — sumar sus totales
                        grp_cols = ["nss","numero_cuenta","clave_interbancaria","banco",
                                    "quincena","nombre_completo"]
                        grp_cols_ok = [c for c in grp_cols if c in df_n.columns]
                        if "total_quincenal" in df_n.columns:
                            df_n["total_quincenal"] = pd.to_numeric(
                                df_n["total_quincenal"], errors="coerce").fillna(0)
                            df_agr = (df_n.groupby(grp_cols_ok, as_index=False, dropna=False)
                                         .agg(total_quincenal=("total_quincenal","sum"),
                                              archivo_origen=("archivo_origen","first")))
                        else:
                            df_agr = df_n
                        for _, r in df_agr.iterrows():
                            nom_rows.append((
                                str(r.get("nss","") or ""),
                                str(r.get("nombre_completo","") or ""),
                                str(r.get("numero_cuenta","") or "") if r.get("numero_cuenta") else "",
                                str(r.get("clave_interbancaria","") or "") if r.get("clave_interbancaria") else "",
                                str(r.get("banco","") or ""),
                                float(r.get("total_quincenal",0) or 0),
                                str(r.get("quincena",1) or 1),
                                str(r.get("archivo_origen","") or ""),
                            ))
                        fuente_nom = f"Sábanas Módulo 1 ({len(nom_rows):,} empleados agrupados)"
                    else:
                        st.error("❌ Sin datos de nómina. Procesa las sábanas en el Módulo 1 o verifica la fecha en PostgreSQL.")
                        st.stop()

                # ── Paso 2: leer PDFs bancarios ───────────────────
                pdf_rows_local = []
                errs_pdf = []
                bar = st.progress(0, "Leyendo PDFs bancarios…")
                for idx_f, f in enumerate(ups_banco):
                    try:
                        with pdfplumber.open(io.BytesIO(f.read())) as p:
                            for pag in p.pages:
                                txt = pag.extract_text() or ""
                                for linea in txt.split("\n"):
                                    m = re.search(r'(\d{10,18})\s+.*?\$?([\d,]{1,15}\.\d{2})', linea)
                                    if m:
                                        try: imp = float(m.group(2).replace(",",""))
                                        except: continue
                                        if imp > 0:
                                            pdf_rows_local.append(("", m.group(1), imp, "", "", f.name))
                    except Exception as ep:
                        errs_pdf.append(f"{f.name}: {ep}")
                    bar.progress((idx_f+1)/len(ups_banco))
                bar.empty()

                if not pdf_rows_local:
                    st.error("❌ No se encontraron registros en los PDFs. Verifica que sean PDFs de dispersión bancaria.")
                else:
                    st.session_state["nom_rows"] = nom_rows
                    st.session_state["pdf_rows"] = pdf_rows_local
                    st.success(f"✅  Nómina: {fuente_nom}  ·  PDFs: {len(pdf_rows_local):,} registros")
                    for e in errs_pdf: st.warning(e)
    else:
        st.markdown('<div class="modo modo-u">👤  Modo usuario — sube nómina y PDFs directamente</div>',
                    unsafe_allow_html=True)
        cx, cy = st.columns(2)
        with cx:
            st.markdown('<div class="up-lbl">Sábanas de nómina — Q1 y Q2 (todos los archivos)</div>',
                        unsafe_allow_html=True)
            if "df_sabanas" in st.session_state:
                n_s = len(st.session_state["df_sabanas"])
                st.markdown(f'<div class="up-ok">✔ Módulo 1 disponible — {n_s:,} empleados procesados</div>',
                            unsafe_allow_html=True)
                st.caption("Las sábanas del Módulo 1 se usarán automáticamente")
            else:
                up_sabs = st.file_uploader("Sábanas Q1 y Q2",
                                            type=["xlsx","xls"],
                                            accept_multiple_files=True,
                                            key="up_sabs_m2",
                                            label_visibility="collapsed")
                if up_sabs:
                    st.markdown(f'<div class="up-ok">📎 {len(up_sabs)} archivo(s)</div>',
                                unsafe_allow_html=True)
        with cy:
            st.markdown('<div class="up-lbl">Archivo dispersión (CSV/Excel de extraer_pdfs.py)</div>',
                        unsafe_allow_html=True)
            up_pdfs = st.file_uploader("Dispersión bancaria", type=["csv","xlsx"],
                                        accept_multiple_files=True,
                                        key="up_banco2", label_visibility="collapsed")
            if up_pdfs:
                st.markdown(f'<div class="up-ok">📋 {len(up_pdfs)} archivo(s)</div>',
                            unsafe_allow_html=True)

        st.markdown('<div style="height:.4rem"></div>', unsafe_allow_html=True)
        if st.button("▶  Preparar datos", use_container_width=True,
                     type="primary", key="btn_prep"):
            # ── Nómina ────────────────────────────────────────────
            df_n = None
            if "df_sabanas" in st.session_state:
                df_n = st.session_state["df_sabanas"]
            elif "up_sabs_m2" in st.session_state or (
                    "up_sabs" in dir() and up_sabs):
                ups_usar = up_sabs if up_sabs else []
                if ups_usar:
                    dfs_m2 = []
                    bar2 = st.progress(0, "Leyendo sábanas…")
                    for idx2, f2 in enumerate(ups_usar):
                        q2 = 1 if any(p in f2.name.lower()
                                      for p in ["1","primera","q1","1a","1er"]) else 2
                        df2, err2 = leer_sabana_bytes(f2.read(), f2.name, q2)
                        if df2 is not None and len(df2) > 0:
                            dfs_m2.append(df2)
                        if err2: st.warning(f"{f2.name}: {err2}")
                        bar2.progress((idx2+1)/len(ups_usar))
                    bar2.empty()
                    if dfs_m2:
                        df_n = pd.concat(dfs_m2, ignore_index=True)
                        st.session_state["df_sabanas"] = df_n

            if df_n is None or len(df_n) == 0:
                st.error("❌ Sin sábanas. Procésalas en el Módulo 1 o súbelas aquí."); st.stop()

            # Agrupar empleados semanales: mismo NSS+cuenta puede aparecer
            # múltiples veces (s1, s2, s3) — sumar sus totales
            df_n2 = df_n.copy()
            grp_cols = ["nss","numero_cuenta","clave_interbancaria",
                        "banco","quincena","nombre_completo"]
            grp_cols_ok = [c for c in grp_cols if c in df_n2.columns]
            if "total_quincenal" in df_n2.columns:
                df_n2["total_quincenal"] = pd.to_numeric(
                    df_n2["total_quincenal"], errors="coerce").fillna(0)
                df_n2 = (df_n2.groupby(grp_cols_ok, as_index=False, dropna=False)
                              .agg(total_quincenal=("total_quincenal","sum"),
                                   archivo_origen=("archivo_origen","first")))
            nom_rows = []
            for _, r in df_n2.iterrows():
                nom_rows.append((
                    str(r.get("nss","") or ""),
                    str(r.get("nombre_completo","") or ""),
                    str(r.get("numero_cuenta","") or "") if r.get("numero_cuenta") else "",
                    str(r.get("clave_interbancaria","") or "") if r.get("clave_interbancaria") else "",
                    str(r.get("banco","") or ""),
                    float(r.get("total_quincenal",0) or 0),
                    str(r.get("quincena",1) or 1),
                    str(r.get("archivo_origen","") or ""),
                ))

            if not up_pdfs:
                st.warning("⬆ Sube el CSV/Excel de dispersión bancaria"); st.stop()

            pdf_rows_xl, errs = [], []
            bar3 = st.progress(0, "Leyendo archivo de dispersión…")
            for idx3, f in enumerate(up_pdfs):
                try:
                    raw3 = f.read()
                    if f.name.lower().endswith(".csv"):
                        import csv as _c3
                        for _r3 in _c3.DictReader(raw3.decode("utf-8-sig").splitlines()):
                            try:
                                _i3=float(_r3.get("importe",0) or 0)
                                if _i3>0: pdf_rows_xl.append(("",_r3.get("cuenta",""),_i3,_r3.get("nombre",""),_r3.get("banco",""),_r3.get("archivo_pdf","") or f.name))
                            except: pass
                    else:
                        import openpyxl as _ox3
                        _wb3=_ox3.load_workbook(io.BytesIO(raw3),data_only=True,read_only=True)
                        _ws3=_wb3.active; _rows3=list(_ws3.iter_rows(values_only=True))
                        _hi3=0
                        for _ri3,_rw3 in enumerate(_rows3[:5]):
                            if any("CUENTA" in str(v or "").upper() for v in _rw3): _hi3=_ri3; break
                        _hdr3=[str(v or "").upper().strip() for v in _rows3[_hi3]]
                        _cib=next((i for i,h in enumerate(_hdr3) if "BANCO" in h),None)
                        _cic=next((i for i,h in enumerate(_hdr3) if "CUENTA" in h),None)
                        _cii=next((i for i,h in enumerate(_hdr3) if "IMPORTE" in h),None)
                        _cin=next((i for i,h in enumerate(_hdr3) if "NOMBRE" in h),None)
                        _cia=next((i for i,h in enumerate(_hdr3) if "ARCHIVO" in h),None)
                        for _rw3 in _rows3[_hi3+1:]:
                            if not _rw3 or not any(_rw3): continue
                            try:
                                _i3=float(_rw3[_cii] or 0) if _cii is not None else 0
                                if _i3<=0: continue
                                pdf_rows_xl.append(("",str(_rw3[_cic] or "") if _cic is not None else "",_i3,str(_rw3[_cin] or "") if _cin is not None else "",str(_rw3[_cib] or "") if _cib is not None else "",str(_rw3[_cia] or f.name) if _cia is not None else f.name))
                            except: pass
                        _wb3.close()
                except Exception as ep: errs.append(f"⚠ {f.name}: {ep}")
                bar3.progress((idx3+1)/len(up_pdfs))
            bar3.empty()
            if not pdf_rows_xl:
                st.error("❌ Sin registros. Sube CSV/Excel de extraer_pdfs.py"); st.stop()

            st.session_state["nom_rows"] = nom_rows
            st.session_state["pdf_rows"] = pdf_rows_xl
            st.success(f"✅  Nómina: {len(nom_rows):,} empleados  ·  PDFs: {len(pdf_rows_xl):,} registros")
            for e in errs: st.warning(e)

    st.markdown('<div style="height:.6rem"></div>', unsafe_allow_html=True)
    st.markdown('<hr style="border:none;border-top:1px solid #E2E8F0;margin:1rem 0">', unsafe_allow_html=True)

    hay_nom = "nom_rows" in st.session_state
    hay_pdf = "pdf_rows" in st.session_state

    if hay_nom and hay_pdf:
        n_nom = len(st.session_state["nom_rows"])
        n_pdf = len(st.session_state["pdf_rows"])
        st.markdown(f'<div style="font-size:.75rem;color:#16A34A;font-weight:600;margin-bottom:.5rem">'
                    f'✔ Listo para cruzar — Nómina: {n_nom:,} registros · PDFs: {n_pdf:,} registros</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div style="font-size:.75rem;color:#8A9BB0;margin-bottom:.5rem">'
                    'Carga la nómina y los PDFs para activar el cruce</div>',
                    unsafe_allow_html=True)

    if st.button("🔀  Ejecutar cruce nómina vs PDF banco",
                  type="primary",
                  use_container_width=True,
                  key="btn_cruce",
                  disabled=not (hay_nom and hay_pdf)):
        with st.spinner("Cruzando…"):
            res, stats, dobles, tot_pdf = hacer_cruce_banco(
                st.session_state["nom_rows"],
                st.session_state["pdf_rows"],
                tolerancia, RANGO_1A, RANGO_2A)
            st.session_state["cruce"] = dict(
                res=res, stats=stats, dobles=dobles,
                tot_pdf=tot_pdf, fecha=fecha_pago_str)
        st.success(f"✅  {len(res):,} registros procesados")

    if "cruce" in st.session_state:
        cb  = st.session_state["cruce"]
        res = cb["res"]; stats = cb["stats"]; dobles = cb["dobles"]
        tr  = len(res) or 1
        tn  = sum(r["imp_nom"] for r in res)
        tp  = sum(r["imp_pdf"] for r in res if r["imp_pdf"])
        df2 = tn - tp
        pok = stats["OK"] / tr * 100

        # KPIs
        st.markdown(f"""
<div class="kpi-row">
  <div class="kpi k-blue">
    <div class="kpi-v">{tr:,}</div>
    <div class="kpi-l">Empleados</div>
    <div class="kpi-s">{cb["fecha"]}</div>
  </div>
  <div class="kpi k-green">
    <div class="kpi-v">{stats["OK"]:,}</div>
    <div class="kpi-l">Confirmados OK</div>
    <div class="kpi-s">{pok:.1f}%</div>
  </div>
  <div class="kpi k-orange">
    <div class="kpi-v">{stats["DIFIERE"]:,}</div>
    <div class="kpi-l">Importe difiere</div>
    <div class="kpi-s">±${tolerancia:.2f}</div>
  </div>
  <div class="kpi k-red">
    <div class="kpi-v">{stats["NO EN PDF"]:,}</div>
    <div class="kpi-l">No en PDF</div>
    <div class="kpi-s">Verificar</div>
  </div>
  <div class="kpi k-purple">
    <div class="kpi-v">{len(dobles):,}</div>
    <div class="kpi-l">Doble abono</div>
    <div class="kpi-s">2+ pagos por cuenta</div>
  </div>
</div>""", unsafe_allow_html=True)

        # Resumen ejecutivo
        dc  = "pos" if abs(df2)<1000 else ("neu" if abs(df2)<50000 else "neg")
        sem = ("✔  Cuadra" if abs(df2)<1000 else
               "⚠  Revisar diferencia" if abs(df2)<50000 else "✘  Diferencia significativa")
        sem_bg = ("#F0FDF4" if abs(df2)<1000 else
                  "#FFFBEB" if abs(df2)<50000 else "#FEF2F2")
        sem_border = ("#BBF7D0" if abs(df2)<1000 else
                      "#FDE68A" if abs(df2)<50000 else "#FECACA")

        grp = defaultdict(lambda:{"nom":0.,"pdf":0.,"ok":0,"no":0,"dif":0})
        for r in res:
            k = r["banco"] or "SIN BANCO"
            grp[k]["nom"] += r["imp_nom"]; grp[k]["pdf"] += r["imp_pdf"] or 0
            if r["estatus"]=="OK": grp[k]["ok"]+=1
            elif r["estatus"]=="NO EN PDF": grp[k]["no"]+=1
            else: grp[k]["dif"]+=1

        banco_rows = ""
        for banco, v in sorted(grp.items()):
            bd3 = v["nom"]-v["pdf"]
            dc3 = "pos" if abs(bd3)<1000 else ("neu" if abs(bd3)<50000 else "neg")
            banco_rows += (
                f'<div class="e-row"><span class="e-lbl">{banco} '
                f'<span style="font-size:.67rem;color:#8A9BB0">'
                f'OK:{v["ok"]} · No PDF:{v["no"]} · Dif:{v["dif"]}</span></span>'
                f'<span class="e-val {dc3}" style="font-size:.75rem">'
                f'${v["nom"]:,.2f} → ${v["pdf"]:,.2f}&nbsp;({bd3:+,.2f})</span></div>')

        st.markdown(f"""
<div class="exec">
  <div class="exec-t">Resumen ejecutivo &nbsp;·&nbsp; {periodo} &nbsp;·&nbsp; {cb["fecha"]}</div>
  <div class="e-row hl">
    <span class="e-lbl"><b>Total nómina dispersada</b></span>
    <span class="e-val">${tn:,.2f}</span>
  </div>
  <div class="e-row hl">
    <span class="e-lbl"><b>Total confirmado en PDFs bancarios</b></span>
    <span class="e-val">${tp:,.2f}</span>
  </div>
  <div class="e-row" style="background:{sem_bg};border:1px solid {sem_border};border-radius:6px;padding:.6rem .75rem;margin:.2rem -.5rem">
    <span class="e-lbl" style="font-weight:700;color:#1A2A3A">{sem}</span>
    <span class="e-val {dc}" style="font-size:1rem">${df2:+,.2f}</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">Empleados confirmados</span>
    <span class="e-val pos">{stats["OK"]:,} &nbsp;({pok:.1f}%)</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">No encontrados en PDFs</span>
    <span class="e-val {"neg" if stats["NO EN PDF"]>0 else "pos"}">{stats["NO EN PDF"]:,}</span>
  </div>
  <div class="e-row">
    <span class="e-lbl">Cuentas con doble abono</span>
    <span class="e-val {"neg" if len(dobles)>0 else "pos"}">{len(dobles):,}</span>
  </div>
  {banco_rows}
</div>""", unsafe_allow_html=True)

        # ── DESCARGA EXCEL — visible antes de la tabla ──────
        st.markdown("""
<div class="sec-hdr" style="margin-top:1rem">
  <span class="sec-badge" style="background:#1A3A6B">⬇</span>
  <span class="sec-title">Descargar reporte Excel</span>
</div>""", unsafe_allow_html=True)
        _xl = None
        try:
            _xl = exportar_nomina_pdf_excel(
                res, stats, dobles, cb["tot_pdf"],
                "HIVICO / FIREWALL SEGURIDAD PRIVADA SA DE CV",
                "B61-79454-10-7", periodo, cb["fecha"], tolerancia)
        except Exception as _ex:
            st.error(f"Error generando Excel: {_ex}")
            import traceback; st.code(traceback.format_exc())
        if _xl:
            st.download_button(
                "⬇  Reporte completo Excel",
                _xl,
                f"CRUCE_NOMINA_PDF_{cb['fecha'].replace('-','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="dl_excel_top")

        # Filtros + tabla
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.4rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">↓</span>
  <span class="sec-title">Detalle de registros</span>
</div>""", unsafe_allow_html=True)

        f1, f2, f3 = st.columns([2,2,3])
        with f1:
            f_banco = st.multiselect("Banco",
                sorted(set(r["banco"] for r in res if r["banco"])), key="fb")
        with f2:
            f_est = st.multiselect("Estado",
                ["OK","NO EN PDF","DIFIERE"], default=["NO EN PDF","DIFIERE"], key="fe")
        with f3:
            f_bus = st.text_input("Buscar NSS o nombre",
                placeholder="Escriba para filtrar…", key="fbus")

        res_f = [r for r in res
                 if (not f_banco or r["banco"] in f_banco)
                 and (not f_est   or any(e in r["estatus"] for e in f_est))
                 and (not f_bus   or f_bus.lower() in (r["nss"]+r["nombre"]).lower())]

        st.markdown(
            f'<div style="font-size:.68rem;color:#8A9BB0;margin-bottom:.5rem">'
            f'Mostrando <b>{min(500,len(res_f)):,}</b> de <b>{len(res_f):,}</b> registros</div>',
            unsafe_allow_html=True)

        rows_t = ""
        for r in res_f[:500]:
            kb = ("ok" if r["estatus"]=="OK" else
                  "no" if r["estatus"]=="NO EN PDF" else "dif")
            doble = '<span class="badge b-doble">Doble</span>' if r["doble_abono"] else ""
            ip  = f"${r['imp_pdf']:,.2f}" if r["imp_pdf"] is not None else                   '<span style="color:#B0BCC8">—</span>'
            ds  = f"${r['diff']:+,.2f}" if r["diff"] is not None else "—"
            dc4 = ("color:#16A34A" if (r["diff"] is not None and abs(r["diff"])<=tolerancia)
                   else "color:#DC2626")
            rows_t += (
                f'<tr><td class="mono">{r["nss"]}</td>'
                f'<td style="color:#2A3A4A">{r["nombre"][:42]}</td>'
                f'<td class="c" style="color:#4A6A8A">{r["banco"]}</td>'
                f'<td class="r">${r["imp_nom"]:,.2f}</td>'
                f'<td class="r">{ip}</td>'
                f'<td class="r" style="{dc4}">{ds}</td>'
                f'<td><span class="badge b-{kb}">'
                f'{"OK" if kb=="ok" else "No en PDF" if kb=="no" else "Difiere"}'
                f'</span>{(" "+doble) if doble else ""}</td></tr>')

        st.markdown(f"""
<div class="t-wrap"><table class="t">
<thead><tr>
  <th>NSS</th><th>Nombre</th><th class="c">Banco</th>
  <th class="r">Nómina</th><th class="r">PDF banco</th>
  <th class="r">Diferencia</th><th>Estado</th>
</tr></thead>
<tbody>{rows_t}</tbody>
<tfoot><tr>
  <td colspan="3">Total visible &nbsp;({min(500,len(res_f)):,} registros)</td>
  <td class="r">${sum(r["imp_nom"] for r in res_f[:500]):,.2f}</td>
  <td class="r">${sum(r["imp_pdf"] for r in res_f[:500] if r["imp_pdf"]):,.2f}</td>
  <td colspan="2"></td>
</tr></tfoot>
</table></div>""", unsafe_allow_html=True)

        if len(res_f) > 500:
            st.caption("Se muestran 500 registros. El Excel contiene el total completo.")

        # Exportar
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.4rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">↓</span>
  <span class="sec-title">Exportar reporte</span>
</div>""", unsafe_allow_html=True)

        xl_banco = None
        try:
            xl_banco = exportar_nomina_pdf_excel(
                res, stats, dobles, cb["tot_pdf"],
                "HIVICO / FIREWALL SEGURIDAD PRIVADA SA DE CV",
                "B61-79454-10-7", periodo, cb["fecha"], tolerancia)
        except Exception as _e_xl:
            st.error(f"Error generando Excel: {_e_xl}")
            import traceback; st.code(traceback.format_exc())

        pdf_ej = None
        try:
            pdf_ej = generar_pdf_ejecutivo(
                res, stats, dobles, tn, tp, periodo, cb["fecha"])
        except Exception as _e_pdf:
            pass  # reportlab opcional

        e1, e2 = st.columns(2)
        with e1:
            if xl_banco:
                st.download_button(
                    "⬇  Reporte completo Excel", xl_banco,
                    f"CRUCE_NOMINA_PDF_{cb['fecha'].replace('-','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary", use_container_width=True)
            else:
                st.warning("No se pudo generar el Excel — revisa el error arriba.")
        with e2:
            if pdf_ej:
                st.download_button(
                    "⬇  Resumen ejecutivo PDF", pdf_ej,
                    f"RESUMEN_{cb['fecha'].replace('-','')}.pdf",
                    mime="application/pdf", use_container_width=True)
            else:
                st.markdown(
                    '<div style="font-size:.71rem;color:#8A9BB0;padding:.5rem 0">'
                    'Instala reportlab: <code>pip install reportlab</code></div>',
                    unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# PARSER DE CÉDULAS IMSS (SUA v3.7.x)
# ══════════════════════════════════════════════════════════════
def _pdfs_de_fuente(uploaded_files):
    """Acepta lista de PDFs directos o ZIPs con PDFs adentro (subcarpetas incluidas).
    Devuelve lista de (nombre, bytes)."""
    import zipfile as _zf
    resultado = []
    for f in (uploaded_files or []):
        raw = f.read()
        if f.name.lower().endswith('.zip'):
            try:
                with _zf.ZipFile(io.BytesIO(raw)) as z:
                    for nombre in sorted(z.namelist()):
                        if nombre.lower().endswith('.pdf') and not os.path.basename(nombre).startswith('.'):
                            with z.open(nombre) as pf:
                                resultado.append((os.path.basename(nombre), pf.read()))
            except Exception as _ze:
                st.warning(f"ZIP {f.name}: {_ze}")
        elif f.name.lower().endswith('.pdf'):
            resultado.append((f.name, raw))
    return resultado


def parsear_cedulas_imss(pdf_bytes, nombre_archivo):
    """Parser SUA v3.7.x — formato de texto extraído por pdfplumber.
    Estructura por línea:
      Par fila NSS:  DD-DD-DD-DDDD-D  NOMBRE  CURP  CLAVE_UBIC
      Par fila ISM:  ISM  FECHA  DIAS  SDI  0  0  0  PAT_CV  OBR_CV  SUMA_CV  SUM_CyV  APT_PAT  AMORT  SUMA_TOTAL
    Devuelve (registros, meta, error)
    """
    import re as _re

    RE_NSS   = _re.compile(r'^(\d{2}-\d{2}-\d{2}-\d{4}-\d)\s+(.+?)\s+([A-Z]{4}[\-]?\d{6}[A-Z0-9]{6,10}[\d]{0,2}|[A-Z]{2,4}-\d{6}-)\s*([A-Z]{2}\d{2}[A-Z]?)?\s*$')
    RE_ISM   = _re.compile(r'^(ISM|ISV|ISB|P/IV|P/CV|S/R|J/R|E/C)\s+(\d{2}/\d{2}/\d{4})\s+(\d{1,3})\s+([\d,]+\.\d{2})')
    RE_BAJA  = _re.compile(r'^Baja\s+(\d{2}/\d{2}/\d{4})')
    RE_NUM   = _re.compile(r'[\d,]+\.\d{2}')
    RE_META1 = _re.compile(r'Registro\s+Patronal[:\s]+([\w\-]+)')
    RE_META2 = _re.compile(r'Bimestre\s+de\s+Proceso[:\s]+(.+?)(?:Fecha|$)')
    RE_META3 = _re.compile(r'Nombre\s+o\s+Raz[oó]n\s+Social[:\s]+(.+?)(?:Delegaci|$)')

    meta = {'reg_patronal': '', 'bimestre': '', 'empresa': '', 'archivo': nombre_archivo}
    registros = []

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            todas_lineas = []
            for pag in pdf.pages:
                txt = pag.extract_text() or ''
                todas_lineas.extend(txt.split('\n'))

        # Meta del encabezado (primeras 15 líneas de cada página repetidas)
        enc = ' '.join(todas_lineas[:15])
        m = RE_META1.search(enc)
        if m: meta['reg_patronal'] = m.group(1).strip()
        m = RE_META2.search(enc)
        if m: meta['bimestre'] = m.group(1).strip()[:30]
        m = RE_META3.search(enc)
        if m: meta['empresa'] = m.group(1).strip()[:80]

        i = 0
        while i < len(todas_lineas):
            linea = todas_lineas[i].strip()

            m_nss = RE_NSS.match(linea)
            if not m_nss:
                i += 1
                continue

            nss_fmt  = m_nss.group(1)
            nss      = nss_fmt.replace('-', '')
            nombre   = m_nss.group(2).strip()
            curp     = m_nss.group(3).strip()
            clave_u  = (m_nss.group(4) or '').strip()

            # Siguiente línea ISM
            dias = 0; sdi = 0.0
            pat_cv = obr_cv = suma_cv = apt_pat = amort = suma_total = 0.0
            fecha_mov = ''; baja = False; fecha_baja = ''

            j = i + 1
            while j < len(todas_lineas) and j <= i + 3:
                sig = todas_lineas[j].strip()
                m_ism = RE_ISM.match(sig)
                if m_ism:
                    fecha_mov = m_ism.group(2)
                    try: dias = int(m_ism.group(3))
                    except: dias = 0
                    try: sdi = float(m_ism.group(4).replace(',',''))
                    except: sdi = 0.0
                    # Todos los números de la línea
                    nums = [float(n.replace(',','')) for n in RE_NUM.findall(sig)]
                    # Estructura: SDI ya capturado, luego 0 0 0, luego:
                    # PAT_CV OBR_CV SUMA_CV SUMA_RCV APT_PAT AMORT SUMA_TOTAL
                    # Filtrar solo los razonables (> 1)
                    vals = [n for n in nums if n > 1.0]
                    # SDI es el primero; quitarlo
                    if vals and abs(vals[0] - sdi) < 0.01:
                        vals = vals[1:]
                    if len(vals) >= 3:
                        pat_cv    = vals[0]
                        obr_cv    = vals[1]
                        suma_cv   = vals[2]
                    if len(vals) >= 5:
                        apt_pat   = vals[3]
                        amort     = vals[4] if len(vals) >= 6 else 0.0
                        suma_total= vals[-1]
                    elif len(vals) >= 4:
                        apt_pat   = vals[3]
                        suma_total= vals[3]
                    break
                j += 1

            # Buscar Baja en las 3 líneas siguientes
            for k in range(i+1, min(i+5, len(todas_lineas))):
                m_b = RE_BAJA.match(todas_lineas[k].strip())
                if m_b:
                    baja = True
                    fecha_baja = m_b.group(1)
                    break

            sueldo_imss = round(sdi * 30, 2) if sdi > 0 else 0.0

            registros.append({
                'nss'           : nss,
                'nss_fmt'       : nss_fmt,
                'nombre'        : nombre,
                'curp'          : curp,
                'clave_ubic'    : clave_u,
                'dias'          : dias,
                'sdi'           : sdi,
                'sueldo_imss'   : sueldo_imss,
                'cuota_patronal': pat_cv,
                'cuota_obrera'  : obr_cv,
                'suma_cv'       : suma_cv,
                'apt_patronal'  : apt_pat,
                'amortizacion'  : amort,
                'suma_total'    : suma_total,
                'baja'          : baja,
                'fecha_baja'    : fecha_baja,
                'fecha_mov'     : fecha_mov,
                'reg_patronal'  : meta['reg_patronal'],
                'bimestre'      : meta['bimestre'],
                'empresa_imss'  : meta['empresa'],
                'archivo'       : nombre_archivo,
            })
            i += 1

        return registros, meta, None

    except Exception as e:
        import traceback
        return [], {}, str(e) + '\n' + traceback.format_exc()


def exportar_imss_excel(df_cruce, meta_list, periodo):
    """Genera Excel del cruce IMSS."""
    wb = Workbook()
    CH = "1F4E79"; CV = "C6EFCE"; CR = "FFC7CE"; CA = "FFEB9C"

    # ── Hoja 1: Resumen ────────────────────────────────────────
    ws1 = wb.active; ws1.title = "RESUMEN IMSS"
    xl_titulo(ws1, f"CRUCE NÓMINA vs IMSS — {periodo}", 1, 6, CH)

    n_ok   = len(df_cruce[df_cruce['estatus'] == 'OK'])
    n_sdi  = len(df_cruce[df_cruce['estatus'] == 'SDI DIFIERE'])
    n_nim  = len(df_cruce[df_cruce['estatus'] == 'NO EN IMSS'])
    n_nsa  = len(df_cruce[df_cruce['estatus'] == 'NO EN SÁBANA'])
    n_baja = len(df_cruce[df_cruce['baja'] == True])
    total  = len(df_cruce)

    ws1.merge_cells("A2:F2")
    ws1["A2"].value = (f"Total trabajadores: {total:,}  |  OK: {n_ok:,}  |  "
                       f"SDI difiere: {n_sdi:,}  |  No en IMSS: {n_nim:,}  |  "
                       f"No en sábana: {n_nsa:,}  |  Bajas con pago: {n_baja:,}")
    ws1["A2"].font = xl_fn(size=9, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="center")

    hdrs = ["CONCEPTO", "CANTIDAD", "% DEL TOTAL", "", "", ""]
    ws1.append(hdrs); fr = ws1.max_row
    for c in ws1[fr]: c.font = xl_fn(bold=True, color="FFFFFF"); c.fill = xl_fill(CH); c.border = xl_brd()

    for concepto, n, color in [
        ("✔  Cuadra (NSS y SDI coinciden)", n_ok, CV),
        ("⚠  SDI difiere > 5%", n_sdi, CA),
        ("✘  En sábana pero NO en IMSS", n_nim, CR),
        ("⚠  En IMSS pero NO en sábana", n_nsa, "FFF3E0"),
        ("🔴  Baja IMSS pero aparece en sábana", n_baja, CR),
    ]:
        pct = n / total * 100 if total else 0
        ws1.append([concepto, n, pct, "", "", ""]); fr = ws1.max_row
        for c in ws1[fr]: c.border = xl_brd(); c.font = xl_fn(size=9); c.fill = xl_fill(color)
        ws1.cell(fr, 3).number_format = '0.0"%"'

    for col, w in zip("ABCDEF", [48, 12, 12, 6, 6, 6]):
        ws1.column_dimensions[col].width = w

    # ── Hoja 2: Detalle completo ───────────────────────────────
    ws2 = wb.create_sheet("DETALLE CRUCE")
    xl_titulo(ws2, f"DETALLE CRUCE NÓMINA vs IMSS — {periodo}", 1, 14, CH)

    hdrs2 = ["NSS", "NOMBRE SÁBANA", "NOMBRE IMSS", "CURP", "ESTATUS",
             "BAJA IMSS", "DÍAS IMSS", "SDI IMSS", "SUELDO IMSS*30",
             "SUELDO SÁBANA", "DIFER. SDI%", "CUOTA PATRONAL", "CUOTA OBRERA", "EMPRESA IMSS"]
    ws2.append(hdrs2); fr = ws2.max_row
    for c in ws2[fr]:
        c.font = xl_fn(bold=True, color="FFFFFF", size=8)
        c.fill = xl_fill(CH); c.border = xl_brd()
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws2.row_dimensions[fr].height = 32

    COLOR_EST = {"OK": CV, "SDI DIFIERE": CA, "NO EN IMSS": CR,
                 "NO EN SÁBANA": "FFF3E0", "BAJA CON PAGO": "FF0000"}
    for _, r in df_cruce.iterrows():
        bg = COLOR_EST.get(r.get('estatus', ''), "FFFFFF")
        ws2.append([
            r.get('nss',''), r.get('nombre_sabana',''), r.get('nombre_imss',''),
            r.get('curp',''), r.get('estatus',''),
            '⚠ BAJA' if r.get('baja') else '',
            r.get('dias', ''), r.get('sdi', ''), r.get('sueldo_imss', ''),
            r.get('sueldo_sabana', ''), r.get('dif_sdi_pct', ''),
            r.get('cuota_patronal', ''), r.get('cuota_obrera', ''),
            r.get('empresa_imss', ''),
        ])
        fr = ws2.max_row
        for c in ws2[fr]: c.border = xl_brd(); c.font = xl_fn(size=8); c.fill = xl_fill(bg)
        for ci in [8, 9, 10, 12, 13]:
            ws2.cell(fr, ci).number_format = '$#,##0.00'
        ws2.cell(fr, 11).number_format = '0.0"%"'
        if r.get('baja'):
            ws2.cell(fr, 6).fill = xl_fill("FF0000")
            ws2.cell(fr, 6).font = xl_fn(bold=True, size=8, color="FFFFFF")

    for i, w in enumerate([14, 32, 32, 20, 16, 12, 8, 10, 14, 14, 10, 14, 14, 28], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # ── Hojas filtradas ────────────────────────────────────────
    for est, titulo, color_h in [
        ("NO EN IMSS",    "NO EN IMSS",    "9C0006"),
        ("NO EN SÁBANA",  "NO EN SABANA",  "7F6000"),
        ("SDI DIFIERE",   "SDI DIFIERE",   "833C00"),
    ]:
        sub = df_cruce[df_cruce['estatus'] == est]
        if len(sub) == 0: continue
        ws = wb.create_sheet(titulo[:30])
        xl_titulo(ws, f"{titulo} — {len(sub):,} registros", 1, 14, color_h)
        ws.append(hdrs2); fr = ws.max_row
        for c in ws[fr]:
            c.font = xl_fn(bold=True, color="FFFFFF", size=8)
            c.fill = xl_fill(color_h); c.border = xl_brd()
        for _, r in sub.iterrows():
            ws.append([
                r.get('nss',''), r.get('nombre_sabana',''), r.get('nombre_imss',''),
                r.get('curp',''), r.get('estatus',''),
                '⚠ BAJA' if r.get('baja') else '',
                r.get('dias',''), r.get('sdi',''), r.get('sueldo_imss',''),
                r.get('sueldo_sabana',''), r.get('dif_sdi_pct',''),
                r.get('cuota_patronal',''), r.get('cuota_obrera',''),
                r.get('empresa_imss',''),
            ])
            fr = ws.max_row
            for c in ws[fr]: c.border = xl_brd(); c.font = xl_fn(size=8)
            for ci in [8, 9, 10, 12, 13]: ws.cell(fr, ci).number_format = '$#,##0.00'
            ws.cell(fr, 11).number_format = '0.0"%"'
        for i, w in enumerate([14, 32, 32, 20, 16, 12, 8, 10, 14, 14, 10, 14, 14, 28], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# ══ TAB 4 · CRUCE IMSS ═══════════════════════════════════════
# ══════════════════════════════════════════════════════════════
with TAB4:
    st.markdown("""
<div class="sec-hdr">
  <span class="sec-badge">04</span>
  <span class="sec-title">Cruce Nómina vs IMSS</span>
  <span class="sec-desc">Cédulas SUA · NSS · SDI · Días cotizados vs trabajados · Cuotas · Bajas</span>
</div>""", unsafe_allow_html=True)

    # ── Uploader: PDFs directos o ZIP con toda la carpeta ─────
    st.markdown('''<div class="up-lbl">Cédulas IMSS — PDFs individuales o ZIP con toda la carpeta del mes</div>''',
                unsafe_allow_html=True)
    ups_imss = st.file_uploader("Cédulas IMSS", type=["pdf","zip"],
                                 accept_multiple_files=True,
                                 key="up_cedulas_imss",
                                 label_visibility="collapsed")
    if ups_imss:
        n_zip = sum(1 for f in ups_imss if f.name.lower().endswith('.zip'))
        n_pdf = sum(1 for f in ups_imss if f.name.lower().endswith('.pdf'))
        partes = []
        if n_pdf: partes.append(f"{n_pdf} PDF(s)")
        if n_zip: partes.append(f"{n_zip} ZIP(s) con subcarpetas")
        st.markdown(f'<div class="up-ok">📎 {" · ".join(partes)}</div>', unsafe_allow_html=True)

    # ── Config ─────────────────────────────────────────────────
    _ci1, _ci2 = st.columns(2)
    with _ci1:
        tol_sdi  = st.slider("Tolerancia SDI (%)", 0, 20, 5, 1,
                              key="tol_sdi_imss",
                              help="Diferencia máxima entre sueldo sábana y SDI×30 para considerar OK")
    with _ci2:
        tol_dias = st.slider("Tolerancia días", 0, 5, 2, 1,
                              key="tol_dias_imss",
                              help="Diferencia máxima de días cotizados vs trabajados para considerar OK")

    # ── Info sábanas ────────────────────────────────────────────
    hay_sabanas = "df_sabanas" in st.session_state
    if hay_sabanas:
        n_sab = len(st.session_state["df_sabanas"])
        st.markdown(f'<div class="up-ok" style="margin:.4rem 0">✔ Sábanas Tab 01 disponibles — {n_sab:,} filas</div>',
                    unsafe_allow_html=True)
    else:
        st.warning("⬆ Procesa primero las sábanas en Tab 01.")

    st.markdown('<div style="height:.4rem"></div>', unsafe_allow_html=True)

    if ups_imss and st.button("▶  Extraer cédulas y cruzar", type="primary",
                               use_container_width=True, key="btn_imss"):
        if not hay_sabanas:
            st.error("❌ Sin sábanas — procésalas en Tab 01 primero.")
        else:
            # ── Paso 1: desempacar archivos (PDFs directos + ZIPs) ──
            pares_pdf = _pdfs_de_fuente(ups_imss)
            if not pares_pdf:
                st.error("❌ No se encontraron PDFs. Verifica que el ZIP contenga PDFs.")
                st.stop()

            st.info(f"📄 {len(pares_pdf)} PDFs encontrados — extrayendo…")

            # ── Paso 2: parsear cada cédula ────────────────────────
            todos_regs = []; todos_meta = []; errs_imss = []
            bar_i = st.progress(0, "Parseando cédulas IMSS…")
            for idx_i, (nom_i, raw_i) in enumerate(pares_pdf):
                regs_i, meta_i, err_i = parsear_cedulas_imss(raw_i, nom_i)
                if err_i:
                    errs_imss.append(f"⚠ {nom_i}: {str(err_i)[:100]}")
                todos_regs.extend(regs_i)
                todos_meta.append((nom_i, len(regs_i),
                                   meta_i.get("reg_patronal",""),
                                   meta_i.get("empresa",""), err_i))
                bar_i.progress((idx_i + 1) / len(pares_pdf))
            bar_i.empty()
            for e in errs_imss: st.warning(e)

            if not todos_regs:
                st.error("❌ Sin registros extraídos. Verifica que los PDFs sean de texto (SUA v3.7.x), no escaneados.")
                st.stop()

            df_imss = pd.DataFrame(todos_regs)
            df_imss["nss_norm"] = df_imss["nss"].astype(str).str.replace(r"\D","",regex=True)

            # Deduplicar IMSS por NSS (si mismo NSS aparece en 2 cédulas tomar la de mayor suma_total)
            df_imss_u = (df_imss.sort_values("suma_total", ascending=False)
                                .drop_duplicates("nss_norm", keep="first"))

            # ── Paso 3: preparar sábana — sumar días por NSS ───────
            df_sab = st.session_state["df_sabanas"].copy()
            df_sab["nss_norm"] = df_sab["nss"].astype(str).str.replace(r"\D","",regex=True)
            df_sab["dias_trabajados"] = pd.to_numeric(
                df_sab.get("dias_trabajados", 0), errors="coerce").fillna(0)
            df_sab["sueldo_mensual"]  = pd.to_numeric(
                df_sab.get("sueldo_mensual", 0), errors="coerce").fillna(0)
            df_sab["total_quincenal"] = pd.to_numeric(
                df_sab.get("total_quincenal", 0), errors="coerce").fillna(0)

            # Agrupar por NSS: sumar días de todos los archivos, tomar sueldo más alto
            df_sab_grp = (df_sab.groupby("nss_norm", as_index=False)
                          .agg(
                              nombre_sabana   =("nombre_completo","first"),
                              sueldo_sabana   =("sueldo_mensual","max"),
                              total_quincenal =("total_quincenal","sum"),
                              dias_nom        =("dias_trabajados","sum"),  # suma bimestre
                              banco_sabana    =("banco","first"),
                              n_archivos      =("archivo_origen","count"),
                              archivos_origen =("archivo_origen",lambda x: " | ".join(sorted(set(x.dropna())))),
                          ))
            # Días nómina: suma de todos los archivos del NSS, sin tope artificial
            # (puede pasar 32 si hay semanales — es válido)

            # ── Paso 4: cruce ───────────────────────────────────────
            nss_imss = set(df_imss_u["nss_norm"].dropna())
            nss_sab  = set(df_sab_grp["nss_norm"].dropna())

            filas = []

            # Trabajadores en IMSS
            for _, ri in df_imss_u.iterrows():
                n = ri["nss_norm"]
                baja   = bool(ri.get("baja", False))
                dias_i = int(ri.get("dias", 0) or 0)
                sdi    = float(ri.get("sdi", 0) or 0)
                sueldo_ims = float(ri.get("sueldo_imss", 0) or 0)

                if n not in nss_sab:
                    filas.append({**ri.to_dict(),
                                  "nombre_sabana":"","sueldo_sabana":0,
                                  "total_quincenal":0,"dias_nom":0,
                                  "dif_dias":None,"dif_sdi_pct":None,
                                  "banco_sabana":"","archivos_origen":"",
                                  "estatus":"NO EN SÁBANA"})
                    continue

                rs = df_sab_grp[df_sab_grp["nss_norm"] == n].iloc[0]
                sueldo_sab = float(rs.get("sueldo_sabana", 0) or 0)
                dias_nom   = int(rs.get("dias_nom", 0) or 0)

                # SDI%: diferencia entre sueldo sábana y SDI×30
                if sueldo_sab > 0 and sueldo_ims > 0:
                    dif_sdi = abs(sueldo_sab - sueldo_ims) / sueldo_sab * 100
                else:
                    dif_sdi = 0.0

                # Días: diferencia días cotizados IMSS vs días nómina
                dif_dias = abs(dias_i - dias_nom)

                # Estatus
                if baja:
                    estatus = "BAJA CON PAGO"
                elif dif_sdi > tol_sdi and dif_dias > tol_dias:
                    estatus = "SDI Y DÍAS DIFIEREN"
                elif dif_sdi > tol_sdi:
                    estatus = "SDI DIFIERE"
                elif dif_dias > tol_dias:
                    estatus = "DÍAS DIFIEREN"
                else:
                    estatus = "OK"

                filas.append({**ri.to_dict(),
                              "nombre_sabana":   rs.get("nombre_sabana",""),
                              "sueldo_sabana":   sueldo_sab,
                              "total_quincenal": rs.get("total_quincenal",0),
                              "dias_nom":        dias_nom,
                              "dif_dias":        dias_nom - dias_i,
                              "dif_sdi_pct":     round(dif_sdi, 1),
                              "banco_sabana":    rs.get("banco_sabana",""),
                              "archivos_origen": rs.get("archivos_origen",""),
                              "estatus":         estatus})

            # En sábana pero no en IMSS
            for _, rs in df_sab_grp[~df_sab_grp["nss_norm"].isin(nss_imss)].iterrows():
                filas.append({
                    "nss_norm":rs["nss_norm"],"nss":rs["nss_norm"],"nss_fmt":rs["nss_norm"],
                    "nombre":"","curp":"","clave_ubic":"",
                    "dias":0,"sdi":0,"sueldo_imss":0,
                    "cuota_patronal":0,"cuota_obrera":0,"suma_cv":0,
                    "apt_patronal":0,"amortizacion":0,"suma_total":0,
                    "baja":False,"fecha_baja":"","fecha_mov":"",
                    "reg_patronal":"","bimestre":"","empresa_imss":"","archivo":"",
                    "nombre_sabana":   rs.get("nombre_sabana",""),
                    "sueldo_sabana":   rs.get("sueldo_sabana",0),
                    "total_quincenal": rs.get("total_quincenal",0),
                    "dias_nom":        rs.get("dias_nom",0),
                    "dif_dias":        None,"dif_sdi_pct":None,
                    "banco_sabana":    rs.get("banco_sabana",""),
                    "archivos_origen": rs.get("archivos_origen",""),
                    "estatus":"NO EN IMSS"})

            df_cruce = pd.DataFrame(filas)
            st.session_state["cruce_imss"] = {
                "df": df_cruce, "meta": todos_meta,
                "n_imss": len(df_imss_u), "n_sab": len(df_sab_grp),
            }
            st.success(f"✅  IMSS: {len(df_imss_u):,} trabajadores · Sábana: {len(df_sab_grp):,} · Cruce: {len(df_cruce):,}")

    # ── Mostrar resultados ─────────────────────────────────────
    if "cruce_imss" in st.session_state:
        ci      = st.session_state["cruce_imss"]
        df_cr   = ci["df"]
        n_ok    = len(df_cr[df_cr["estatus"] == "OK"])
        n_sdi   = len(df_cr[df_cr["estatus"].str.contains("SDI", na=False)])
        n_dd    = len(df_cr[df_cr["estatus"].str.contains("DÍAS", na=False)])
        n_nim   = len(df_cr[df_cr["estatus"] == "NO EN IMSS"])
        n_nsa   = len(df_cr[df_cr["estatus"] == "NO EN SÁBANA"])
        n_baja  = len(df_cr[df_cr["baja"] == True])
        total   = len(df_cr) or 1
        pok     = n_ok / total * 100

        st.markdown(f"""
<div class="kpi-row">
  <div class="kpi k-blue"><div class="kpi-v">{total:,}</div>
    <div class="kpi-l">Total trabajadores</div>
    <div class="kpi-s">IMSS: {ci["n_imss"]:,} · Sábana: {ci["n_sab"]:,}</div></div>
  <div class="kpi k-green"><div class="kpi-v">{n_ok:,}</div>
    <div class="kpi-l">Coinciden OK</div>
    <div class="kpi-s">{pok:.1f}%</div></div>
  <div class="kpi k-orange"><div class="kpi-v">{n_sdi:,}</div>
    <div class="kpi-l">SDI difiere</div>
    <div class="kpi-s">Posible subdeclaración</div></div>
  <div class="kpi k-orange"><div class="kpi-v">{n_dd:,}</div>
    <div class="kpi-l">Días difieren</div>
    <div class="kpi-s">Nómina vs IMSS</div></div>
  <div class="kpi k-red"><div class="kpi-v">{n_nim:,}</div>
    <div class="kpi-l">Sin IMSS</div>
    <div class="kpi-s">En nómina, no en cédula</div></div>
  <div class="kpi k-purple"><div class="kpi-v">{n_nsa:,}</div>
    <div class="kpi-l">Sin sábana</div>
    <div class="kpi-s">En IMSS, no en nómina</div></div>
</div>""", unsafe_allow_html=True)

        if n_baja > 0:
            st.error(f"🔴  {n_baja:,} trabajador(es) con BAJA IMSS siguen en sábana — revisar urgente")

        # Log cédulas
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.2rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">LOG</span>
  <span class="sec-title">Cédulas procesadas</span>
</div>""", unsafe_allow_html=True)
        _log_html = ""
        for arch, n_r, reg_pat, empresa, err in ci["meta"]:
            ok_ic = ('<span style="color:#16A34A;font-weight:700">✔</span>' if not err
                     else '<span style="color:#DC2626;font-weight:700">✘</span>')
            _log_html += (f'<tr><td class="c">{ok_ic}</td>'
                          f'<td class="mono">{arch}</td>'
                          f'<td style="font-size:.71rem;color:#4A6A8A">{reg_pat}</td>'
                          f'<td style="font-size:.71rem;color:#4A6A8A">{str(empresa)[:40]}</td>'
                          f'<td class="r">{n_r:,}</td>'
                          f'<td style="font-size:.7rem;color:{"#DC2626" if err else "#16A34A"}">' +
                          (str(err)[:60] if err else "OK") + '</td></tr>')
        st.markdown(f"""
<div class="t-wrap"><table class="t">
<thead><tr><th class="c"></th><th>Archivo</th><th>Reg. Patronal</th>
<th>Empresa</th><th class="r">Trabajadores</th><th>Estado</th>
</tr></thead><tbody>{_log_html}</tbody></table></div>""", unsafe_allow_html=True)

        # Excel
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.2rem">
  <span class="sec-badge" style="background:#1A3A6B">⬇</span>
  <span class="sec-title">Descargar reporte Excel</span>
</div>""", unsafe_allow_html=True)
        try:
            _xl_imss = exportar_imss_excel(df_cr, ci["meta"], periodo)
            st.download_button("⬇  Reporte cruce IMSS Excel", _xl_imss,
                f"CRUCE_IMSS_{mes_sel[:3].upper()}{anio_sel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True, key="dl_imss")
        except Exception as _ex_xl:
            st.error(f"Error Excel: {_ex_xl}")

        # Filtros + tabla
        st.markdown("""
<div class="sec-hdr" style="margin-top:1.4rem">
  <span class="sec-badge" style="background:#F0F5FF;color:#1A3A6B">↓</span>
  <span class="sec-title">Detalle de registros</span>
</div>""", unsafe_allow_html=True)

        _fi1, _fi2, _fi3 = st.columns([2,2,3])
        with _fi1:
            f_est_i = st.multiselect("Estatus",
                ["OK","SDI DIFIERE","DÍAS DIFIEREN","SDI Y DÍAS DIFIEREN",
                 "NO EN IMSS","NO EN SÁBANA","BAJA CON PAGO"],
                default=["SDI DIFIERE","DÍAS DIFIEREN","SDI Y DÍAS DIFIEREN",
                         "NO EN IMSS","NO EN SÁBANA","BAJA CON PAGO"],
                key="fi_est_i")
        with _fi2:
            f_baja_i = st.checkbox("Solo bajas IMSS con pago", key="fi_baja_i")
        with _fi3:
            f_bus_i = st.text_input("Buscar NSS o nombre", key="fi_bus_i")

        df_f = df_cr.copy()
        if f_est_i: df_f = df_f[df_f["estatus"].isin(f_est_i)]
        if f_baja_i: df_f = df_f[df_f["baja"] == True]
        if f_bus_i:
            _b = f_bus_i.lower()
            df_f = df_f[df_f.apply(
                lambda r: _b in (str(r.get("nss",""))+str(r.get("nombre_sabana",""))+str(r.get("nombre",""))).lower(),
                axis=1)]

        st.markdown(f'<div style="font-size:.68rem;color:#8A9BB0;margin-bottom:.5rem">'
                    f'Mostrando <b>{min(500, len(df_f)):,}</b> de <b>{len(df_f):,}</b> registros</div>',
                    unsafe_allow_html=True)

        _CB = {"OK":"b-ok","SDI DIFIERE":"b-dif","DÍAS DIFIEREN":"b-dif",
               "SDI Y DÍAS DIFIEREN":"b-dif","NO EN IMSS":"b-no",
               "NO EN SÁBANA":"b-dif","BAJA CON PAGO":"b-no"}
        _rows_i = ""
        for _, r in df_f.head(500).iterrows():
            _est = r.get("estatus","")
            _bc  = _CB.get(_est,"")
            _nombre = str(r.get("nombre_sabana") or r.get("nombre",""))[:38]
            _sdi_s  = f"${r['sdi']:,.2f}" if r.get("sdi") else "—"
            _sue_s  = f"${r.get('sueldo_sabana',0):,.2f}" if r.get("sueldo_sabana") else "—"
            _pct_s  = f"{r['dif_sdi_pct']:.1f}%" if r.get("dif_sdi_pct") is not None else "—"
            _di     = int(r.get("dias") or 0)
            _dn     = int(r.get("dias_nom") or 0)
            _dd     = r.get("dif_dias")
            _dd_s   = (f'<span style="color:{"#DC2626" if _dd is not None and abs(_dd) > tol_dias else "#16A34A"}">' +
                       f'{_dd:+d}</span>') if _dd is not None else "—"
            _baja_s = '<span style="color:#DC2626;font-weight:700">⚠ BAJA</span>' if r.get("baja") else ""
            _rows_i += (f'<tr><td class="mono">{r.get("nss_fmt") or r.get("nss","")}</td>'
                        f'<td>{_nombre}</td>'
                        f'<td class="r">{_sdi_s}</td>'
                        f'<td class="r">{_sue_s}</td>'
                        f'<td class="r">{_pct_s}</td>'
                        f'<td class="c">{_di}</td>'
                        f'<td class="c">{_dn}</td>'
                        f'<td class="c">{_dd_s}</td>'
                        f'<td>{_baja_s}</td>'
                        f'<td><span class="badge {_bc}">{_est}</span></td></tr>')

        st.markdown(f"""
<div class="t-wrap"><table class="t">
<thead><tr>
  <th>NSS</th><th>Nombre</th>
  <th class="r">SDI IMSS</th><th class="r">Sueldo Sábana</th><th class="r">Dif SDI%</th>
  <th class="c">Días IMSS</th><th class="c">Días Nóm</th><th class="c">Dif Días</th>
  <th>Baja</th><th>Estatus</th>
</tr></thead>
<tbody>{_rows_i}</tbody>
</table></div>""", unsafe_allow_html=True)

        if len(df_f) > 500:
            st.caption("Se muestran 500 registros. El Excel contiene el total completo.")
